"""
Advanced Accounting System - Admin Interface
Complete admin for journals, ledgers, vouchers, and financial management
"""

from django.contrib import admin
from django.utils.html import format_html
from django.db.models import Sum, Q
from django.urls import reverse
from django.utils.safestring import mark_safe
from .models_accounting_advanced import (
    AccountCategory, FiscalYear, AccountingPeriod,
    Journal, AdvancedJournalEntry, AdvancedJournalEntryLine, AdvancedGeneralLedger,
    PaymentVoucher, ReceiptVoucher,
    RevenueCategory, Revenue,
    ExpenseCategory, Expense,
    AdvancedAccountsReceivable, AccountsPayable,
    BankAccount, BankTransaction,
    Budget, BudgetLine, TaxRate,
    AccountingAuditLog
)


# ==================== INLINE ADMINS ====================

class AdvancedJournalEntryLineInline(admin.TabularInline):
    model = AdvancedJournalEntryLine
    extra = 2
    fields = ['line_number', 'account', 'cost_center', 'description', 'debit_amount', 'credit_amount']
    
    def get_readonly_fields(self, request, obj=None):
        if obj and obj.status == 'posted':
            return ['line_number', 'account', 'cost_center', 'description', 'debit_amount', 'credit_amount']
        return []


class BudgetLineInline(admin.TabularInline):
    model = BudgetLine
    extra = 1
    fields = ['account', 'cost_center', 'budgeted_amount', 'actual_amount', 'variance', 'variance_percent']
    readonly_fields = ['actual_amount', 'variance', 'variance_percent']


# ==================== ACCOUNT MANAGEMENT ====================

@admin.register(AccountCategory)
class AccountCategoryAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'category_type', 'is_active']
    list_filter = ['category_type', 'is_active']
    search_fields = ['code', 'name']
    
    fieldsets = (
        (None, {
            'fields': ('code', 'name', 'category_type', 'description')
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )


@admin.register(FiscalYear)
class FiscalYearAdmin(admin.ModelAdmin):
    list_display = ['name', 'start_date', 'end_date', 'is_current_display', 'is_closed', 'closed_date']
    list_filter = ['is_closed', 'start_date']
    search_fields = ['name']
    readonly_fields = ['is_current_display']
    
    fieldsets = (
        ('Fiscal Year Information', {
            'fields': ('name', 'start_date', 'end_date')
        }),
        ('Status', {
            'fields': ('is_closed', 'closed_date', 'closed_by', 'is_current_display')
        }),
    )
    
    def is_current_display(self, obj):
        if obj.is_current:
            return format_html('<span style="color: green; font-weight: bold;">✓ Current</span>')
        return format_html('<span style="color: gray;">Not Current</span>')
    is_current_display.short_description = 'Status'


@admin.register(AccountingPeriod)
class AccountingPeriodAdmin(admin.ModelAdmin):
    list_display = ['name', 'fiscal_year', 'period_number', 'start_date', 'end_date', 'is_closed']
    list_filter = ['fiscal_year', 'is_closed', 'start_date']
    search_fields = ['name']
    
    fieldsets = (
        ('Period Information', {
            'fields': ('fiscal_year', 'period_number', 'name', 'start_date', 'end_date')
        }),
        ('Status', {
            'fields': ('is_closed', 'closed_date')
        }),
    )


# ==================== JOURNALS & LEDGERS ====================

@admin.register(Journal)
class JournalAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'journal_type', 'is_active']
    list_filter = ['journal_type', 'is_active']
    search_fields = ['code', 'name']
    
    fieldsets = (
        ('Journal Information', {
            'fields': ('code', 'name', 'journal_type', 'description')
        }),
        ('Default Accounts', {
            'fields': ('default_debit_account', 'default_credit_account'),
            'classes': ('collapse',)
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )


@admin.register(AdvancedJournalEntry)
class AdvancedJournalEntryAdmin(admin.ModelAdmin):
    list_display = ['entry_number', 'entry_date', 'journal', 'description_short', 'total_debit', 'total_credit', 'status_badge', 'is_balanced_display']
    list_filter = ['status', 'journal', 'entry_date', 'fiscal_year']
    search_fields = ['entry_number', 'description', 'reference']
    readonly_fields = ['entry_number', 'total_debit', 'total_credit', 'is_balanced_display', 'posted_by', 'posting_date']
    date_hierarchy = 'entry_date'
    
    inlines = [AdvancedJournalEntryLineInline]
    
    fieldsets = (
        ('Journal Entry Information', {
            'fields': ('entry_number', 'journal', 'entry_date', 'posting_date', 'reference')
        }),
        ('Description', {
            'fields': ('description', 'notes')
        }),
        ('Accounting Period', {
            'fields': ('fiscal_year', 'accounting_period')
        }),
        ('Totals', {
            'fields': ('total_debit', 'total_credit', 'is_balanced_display')
        }),
        ('Status & Tracking', {
            'fields': ('status', 'created_by', 'posted_by')
        }),
        ('Links', {
            'fields': ('invoice', 'reversed_entry'),
            'classes': ('collapse',)
        }),
    )
    
    def description_short(self, obj):
        return obj.description[:50] + '...' if len(obj.description) > 50 else obj.description
    description_short.short_description = 'Description'
    
    def status_badge(self, obj):
        colors = {
            'draft': 'secondary',
            'posted': 'success',
            'void': 'danger',
            'reversed': 'warning'
        }
        color = colors.get(obj.status, 'secondary')
        return format_html(
            '<span class="badge bg-{}">{}</span>',
            color,
            obj.get_status_display()
        )
    status_badge.short_description = 'Status'
    
    def is_balanced_display(self, obj):
        if obj.is_balanced:
            return format_html('<span style="color: green;">✓ Balanced</span>')
        return format_html('<span style="color: red;">✗ Not Balanced</span>')
    is_balanced_display.short_description = 'Balanced?'
    
    def get_readonly_fields(self, request, obj=None):
        readonly = list(self.readonly_fields)
        if obj and obj.status == 'posted':
            readonly.extend(['journal', 'entry_date', 'fiscal_year', 'accounting_period'])
        return readonly
    
    actions = ['post_entries', 'void_entries']
    
    def post_entries(self, request, queryset):
        count = 0
        for entry in queryset.filter(status='draft'):
            try:
                entry.post(request.user)
                count += 1
            except Exception as e:
                self.message_user(request, f"Error posting {entry.entry_number}: {e}", level='error')
        
        self.message_user(request, f"Successfully posted {count} journal entries", level='success')
    post_entries.short_description = "Post selected journal entries"
    
    def void_entries(self, request, queryset):
        count = queryset.filter(status='posted').update(status='void')
        self.message_user(request, f"Voided {count} journal entries", level='warning')
    void_entries.short_description = "Void selected journal entries"


@admin.register(AdvancedGeneralLedger)
class AdvancedGeneralLedgerAdmin(admin.ModelAdmin):
    list_display = ['transaction_date', 'account', 'description_short', 'debit_amount', 'credit_amount', 'balance', 'journal_entry_link']
    list_filter = ['account', 'transaction_date', 'fiscal_year', 'accounting_period', 'is_voided']
    search_fields = ['description', 'journal_entry__entry_number', 'account__account_code', 'account__account_name']
    readonly_fields = ['journal_entry', 'journal_entry_line', 'account', 'transaction_date', 'posting_date', 'debit_amount', 'credit_amount', 'balance']
    date_hierarchy = 'transaction_date'
    
    def has_add_permission(self, request):
        return False  # Ledger entries created automatically from journal entries
    
    def has_delete_permission(self, request, obj=None):
        return False  # Cannot delete ledger entries
    
    def description_short(self, obj):
        return obj.description[:40] + '...' if len(obj.description) > 40 else obj.description
    description_short.short_description = 'Description'
    
    def journal_entry_link(self, obj):
        url = reverse('admin:hospital_advancedjournalentry_change', args=[obj.journal_entry.pk])
        return format_html('<a href="{}">{}</a>', url, obj.journal_entry.entry_number)
    journal_entry_link.short_description = 'Journal Entry'


# ==================== PAYMENT & RECEIPT VOUCHERS ====================

@admin.register(PaymentVoucher)
class PaymentVoucherAdmin(admin.ModelAdmin):
    list_display = ['voucher_number', 'voucher_date', 'payee_name', 'payment_type', 'amount', 'status_badge', 'payment_method']
    list_filter = ['status', 'payment_type', 'payment_method', 'voucher_date']
    search_fields = ['voucher_number', 'payee_name', 'description', 'payment_reference']
    readonly_fields = ['voucher_number', 'journal_entry', 'approved_date', 'created', 'modified']
    date_hierarchy = 'voucher_date'
    autocomplete_fields = ['expense_account', 'payment_account']
    
    fieldsets = (
        ('Voucher Information', {
            'fields': ('voucher_number', 'voucher_date', 'payment_type', 'status'),
            'description': 'Basic voucher identification and type'
        }),
        ('Payee Details', {
            'fields': ('payee_name', 'payee_type', 'amount'),
            'description': 'Who is being paid and how much'
        }),
        ('Description', {
            'fields': ('description',),
        }),
        ('Payment Details', {
            'fields': ('payment_method', 'payment_date', 'payment_reference', 'bank_name', 'account_number', 'cheque_number'),
            'description': 'Payment processing information',
            'classes': ('collapse',)
        }),
        ('Accounting Links', {
            'fields': ('expense_account', 'payment_account', 'journal_entry'),
            'description': 'Linked accounting accounts and journal entries'
        }),
        ('Supporting Documents', {
            'fields': ('invoice_number', 'po_number'),
            'classes': ('collapse',)
        }),
        ('Approval Workflow', {
            'fields': ('requested_by', 'approved_by', 'approved_date', 'paid_by'),
            'description': 'Approval and payment audit trail'
        }),
        ('Additional Information', {
            'fields': ('notes', 'created', 'modified'),
            'classes': ('collapse',)
        }),
    )
    
    def get_readonly_fields(self, request, obj=None):
        """Make certain fields readonly for existing vouchers"""
        readonly = ['voucher_number', 'journal_entry', 'approved_date', 'created', 'modified']
        if obj and obj.status in ['paid', 'void']:
            # Can't edit paid or voided vouchers
            readonly.extend(['amount', 'payee_name', 'expense_account', 'payment_account'])
        return readonly
    
    def status_badge(self, obj):
        colors = {
            'draft': 'secondary',
            'pending_approval': 'warning',
            'approved': 'info',
            'paid': 'success',
            'rejected': 'danger',
            'void': 'dark'
        }
        color = colors.get(obj.status, 'secondary')
        return format_html(
            '<span class="badge bg-{}">{}</span>',
            color,
            obj.get_status_display()
        )
    status_badge.short_description = 'Status'
    
    actions = ['approve_vouchers', 'mark_as_paid', 'export_to_excel']
    
    list_per_page = 50
    
    def approve_vouchers(self, request, queryset):
        count = 0
        for voucher in queryset.filter(status='pending_approval'):
            try:
                voucher.approve(request.user)
                count += 1
            except Exception as e:
                self.message_user(request, f"Error approving {voucher.voucher_number}: {e}", level='error')
        
        self.message_user(request, f"Approved {count} payment vouchers", level='success')
    approve_vouchers.short_description = "Approve selected vouchers"
    
    def mark_as_paid(self, request, queryset):
        """Mark selected approved vouchers as paid"""
        count = 0
        for voucher in queryset.filter(status='approved'):
            try:
                # Simple mark as paid without journal entry (for now)
                voucher.status = 'paid'
                voucher.payment_date = timezone.now().date()
                voucher.paid_by = request.user
                voucher.save()
                count += 1
            except Exception as e:
                self.message_user(request, f"Error marking {voucher.voucher_number} as paid: {e}", level='error')
        
        if count > 0:
            self.message_user(request, f"✅ Marked {count} vouchers as paid successfully!", level='success')
    mark_as_paid.short_description = "✅ Mark selected as paid"
    
    def export_to_excel(self, request, queryset):
        """Export selected vouchers to Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from datetime import datetime
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payment Vouchers"
        
        # Headers
        headers = ['Voucher #', 'Date', 'Payee', 'Type', 'Amount (GHS)', 'Status', 'Payment Date', 'Reference']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        for voucher in queryset:
            ws.append([
                voucher.voucher_number,
                voucher.voucher_date.strftime('%Y-%m-%d'),
                voucher.payee_name,
                voucher.get_payment_type_display(),
                float(voucher.amount),
                voucher.get_status_display(),
                voucher.payment_date.strftime('%Y-%m-%d') if voucher.payment_date else '',
                voucher.payment_reference or '',
            ])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="payment_vouchers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    export_to_excel.short_description = "📊 Export to Excel"


@admin.register(ReceiptVoucher)
class ReceiptVoucherAdmin(admin.ModelAdmin):
    list_display = ['receipt_number', 'receipt_date', 'received_from', 'amount', 'payment_method', 'status_badge']
    list_filter = ['status', 'payment_method', 'receipt_date']
    search_fields = ['receipt_number', 'received_from', 'description', 'reference']
    readonly_fields = ['receipt_number', 'journal_entry', 'received_by']
    date_hierarchy = 'receipt_date'
    
    fieldsets = (
        ('Receipt Information', {
            'fields': ('receipt_number', 'receipt_date')
        }),
        ('From', {
            'fields': ('received_from', 'patient', 'description')
        }),
        ('Payment', {
            'fields': ('amount', 'payment_method', 'reference')
        }),
        ('Accounting', {
            'fields': ('revenue_account', 'cash_account', 'journal_entry', 'invoice')
        }),
        ('Status', {
            'fields': ('status', 'received_by')
        }),
    )
    
    def status_badge(self, obj):
        colors = {'draft': 'secondary', 'issued': 'success', 'void': 'danger'}
        color = colors.get(obj.status, 'secondary')
        return format_html('<span class="badge bg-{}">{}</span>', color, obj.get_status_display())
    status_badge.short_description = 'Status'


# ==================== REVENUE & EXPENSE ====================

@admin.register(RevenueCategory)
class RevenueCategoryAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'account', 'is_active']
    list_filter = ['is_active']
    search_fields = ['code', 'name']


@admin.register(Revenue)
class RevenueAdmin(admin.ModelAdmin):
    list_display = ['revenue_number', 'revenue_date', 'category', 'amount', 'patient', 'payment_method']
    list_filter = ['category', 'payment_method', 'revenue_date']
    search_fields = ['revenue_number', 'description', 'patient__first_name', 'patient__last_name']
    readonly_fields = ['revenue_number', 'journal_entry', 'recorded_by']
    date_hierarchy = 'revenue_date'
    
    fieldsets = (
        ('Revenue Information', {
            'fields': ('revenue_number', 'revenue_date', 'category', 'description')
        }),
        ('Amount', {
            'fields': ('amount',)
        }),
        ('Source', {
            'fields': ('patient', 'invoice')
        }),
        ('Payment', {
            'fields': ('payment_method', 'reference')
        }),
        ('Accounting', {
            'fields': ('journal_entry', 'receipt_voucher', 'recorded_by')
        }),
    )


@admin.register(ExpenseCategory)
class ExpenseCategoryAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'account', 'requires_approval', 'approval_limit', 'is_active']
    list_filter = ['requires_approval', 'is_active']
    search_fields = ['code', 'name']


@admin.register(Expense)
class ExpenseAdmin(admin.ModelAdmin):
    list_display = ['expense_number', 'expense_date', 'category', 'vendor_name', 'amount', 'status_badge']
    list_filter = ['status', 'category', 'expense_date']
    search_fields = ['expense_number', 'vendor_name', 'description', 'vendor_invoice_number']
    readonly_fields = ['expense_number', 'payment_voucher', 'journal_entry', 'approved_by']
    date_hierarchy = 'expense_date'
    
    fieldsets = (
        ('Expense Information', {
            'fields': ('expense_number', 'expense_date', 'category', 'description')
        }),
        ('Vendor', {
            'fields': ('vendor_name', 'vendor_invoice_number')
        }),
        ('Amount', {
            'fields': ('amount',)
        }),
        ('Status & Approval', {
            'fields': ('status', 'recorded_by', 'approved_by')
        }),
        ('Accounting', {
            'fields': ('payment_voucher', 'journal_entry')
        }),
    )
    
    def status_badge(self, obj):
        colors = {
            'draft': 'secondary',
            'pending': 'warning',
            'approved': 'info',
            'paid': 'success',
            'rejected': 'danger'
        }
        color = colors.get(obj.status, 'secondary')
        return format_html('<span class="badge bg-{}">{}</span>', color, obj.get_status_display())
    status_badge.short_description = 'Status'


# ==================== RECEIVABLES & PAYABLES ====================

@admin.register(AdvancedAccountsReceivable)
class AdvancedAccountsReceivableAdmin(admin.ModelAdmin):
    list_display = ['invoice', 'patient', 'invoice_amount', 'amount_paid', 'balance_due', 'due_date', 'aging_badge', 'overdue_days']
    list_filter = ['is_overdue', 'aging_bucket', 'due_date']
    search_fields = ['invoice__invoice_number', 'patient__first_name', 'patient__last_name']
    readonly_fields = ['balance_due', 'is_overdue', 'days_overdue', 'aging_bucket']
    date_hierarchy = 'due_date'
    
    def aging_badge(self, obj):
        colors = {
            'current': 'success',
            '0-30': 'info',
            '31-60': 'warning',
            '61-90': 'danger',
            '90+': 'dark'
        }
        color = colors.get(obj.aging_bucket, 'secondary')
        return format_html('<span class="badge bg-{}">{}</span>', color, obj.aging_bucket)
    aging_badge.short_description = 'Aging'
    
    def overdue_days(self, obj):
        if obj.is_overdue:
            return format_html('<span style="color: red;">{} days</span>', obj.days_overdue)
        return '-'
    overdue_days.short_description = 'Overdue'


@admin.register(AccountsPayable)
class AccountsPayableAdmin(admin.ModelAdmin):
    list_display = ['bill_number', 'vendor_name', 'bill_date', 'due_date', 'amount', 'balance_due', 'is_overdue', 'payment_voucher']
    list_filter = ['is_overdue', 'bill_date', 'due_date']
    search_fields = ['bill_number', 'vendor_name', 'vendor_invoice', 'description']
    readonly_fields = ['balance_due', 'is_overdue', 'days_overdue']
    date_hierarchy = 'due_date'


# ==================== BANK & CASH ====================

@admin.register(BankAccount)
class BankAccountAdmin(admin.ModelAdmin):
    list_display = ['account_name', 'bank_name', 'account_number', 'account_type', 'current_balance', 'currency', 'is_active']
    list_filter = ['bank_name', 'account_type', 'currency', 'is_active']
    search_fields = ['account_name', 'account_number', 'bank_name']
    
    fieldsets = (
        ('Account Information', {
            'fields': ('account_name', 'account_number', 'account_type', 'currency')
        }),
        ('Bank Details', {
            'fields': ('bank_name', 'branch')
        }),
        ('Balances', {
            'fields': ('opening_balance', 'current_balance')
        }),
        ('Accounting', {
            'fields': ('gl_account',)
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
    )


@admin.register(BankTransaction)
class BankTransactionAdmin(admin.ModelAdmin):
    list_display = ['bank_account', 'transaction_date', 'transaction_type', 'description_short', 'amount', 'reconciled_badge']
    list_filter = ['bank_account', 'transaction_type', 'is_reconciled', 'transaction_date']
    search_fields = ['description', 'reference']
    date_hierarchy = 'transaction_date'
    
    def description_short(self, obj):
        return obj.description[:40] + '...' if len(obj.description) > 40 else obj.description
    description_short.short_description = 'Description'
    
    def reconciled_badge(self, obj):
        if obj.is_reconciled:
            return format_html('<span class="badge bg-success">Reconciled</span>')
        return format_html('<span class="badge bg-warning">Pending</span>')
    reconciled_badge.short_description = 'Reconciliation'
    
    actions = ['mark_reconciled']
    
    def mark_reconciled(self, request, queryset):
        count = queryset.update(is_reconciled=True, reconciled_date=timezone.now().date())
        self.message_user(request, f"Marked {count} transactions as reconciled", level='success')
    mark_reconciled.short_description = "Mark as reconciled"


# ==================== BUDGETING ====================

@admin.register(Budget)
class BudgetAdmin(admin.ModelAdmin):
    list_display = ['name', 'fiscal_year', 'start_date', 'end_date', 'total_revenue_budget', 'total_expense_budget', 'is_approved']
    list_filter = ['is_approved', 'fiscal_year', 'start_date']
    search_fields = ['name', 'description']
    
    inlines = [BudgetLineInline]
    
    fieldsets = (
        ('Budget Information', {
            'fields': ('name', 'fiscal_year', 'accounting_period', 'start_date', 'end_date', 'description')
        }),
        ('Totals', {
            'fields': ('total_revenue_budget', 'total_expense_budget')
        }),
        ('Approval', {
            'fields': ('is_approved', 'approved_by')
        }),
    )


@admin.register(TaxRate)
class TaxRateAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'rate', 'account', 'effective_date', 'is_active']
    list_filter = ['is_active', 'effective_date']
    search_fields = ['code', 'name']


# ==================== AUDIT LOG ====================

@admin.register(AccountingAuditLog)
class AccountingAuditLogAdmin(admin.ModelAdmin):
    list_display = ['timestamp', 'user', 'action', 'model_name', 'object_repr']
    list_filter = ['action', 'model_name', 'timestamp']
    search_fields = ['object_repr', 'user__username']
    readonly_fields = ['user', 'action', 'timestamp', 'model_name', 'object_id', 'object_repr', 'changes', 'ip_address']
    date_hierarchy = 'timestamp'
    
    def has_add_permission(self, request):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False

