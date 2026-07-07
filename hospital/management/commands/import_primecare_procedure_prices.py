"""
Import procedure prices from a PrimeCare-style Excel workbook into ProcedureCatalog
and optional CashierChargeBundle rows (rows sharing a bundle group become one package).
"""
from __future__ import annotations

import hashlib
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from hospital.models import CashierChargeBundle
from hospital.models_advanced import ProcedureCatalog


def _letters_to_index(letters: str) -> int:
    s = letters.strip().upper()
    n = 0
    for ch in s:
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f'Invalid column letters: {letters!r}')
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n


def _cell(ws, row_1based: int, col_1based: int):
    if col_1based < 1:
        return None
    return ws.cell(row=row_1based, column=col_1based).value


def _norm_str(val) -> str:
    if val is None:
        return ''
    return str(val).strip()


def _parse_money(val) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, Decimal):
        return val
    s = _norm_str(val)
    if not s:
        return None
    s = re.sub(r'[^\d.\-]', '', s.replace(',', ''))
    if not s or s == '-':
        return None
    try:
        return Decimal(s).quantize(Decimal('0.01'))
    except InvalidOperation:
        return None


def _stable_procedure_code(raw_code: str, name: str) -> str:
    base = _norm_str(raw_code)
    if base:
        safe = re.sub(r'[^A-Za-z0-9_-]', '-', base)[:44]
        if safe:
            return safe[:50]
    h = hashlib.sha1(name.encode('utf-8')).hexdigest()[:8].upper()
    slug = re.sub(r'[^A-Za-z0-9]+', '-', name.upper())[:20].strip('-') or 'PROC'
    return f'{slug}-{h}'[:50]


_GENERATED_CODE_SUFFIX_RE = re.compile(r'-([0-9A-F]{8})$')


def procedure_code_looks_generated_from_import(code: str) -> bool:
    """True for codes emitted by `_stable_procedure_code` when Excel has no explicit CODE column."""
    return bool((code or '').strip() and _GENERATED_CODE_SUFFIX_RE.search(code.strip()))


def _prune_orphan_generated_procedures(
    *,
    canonical_codes: set[str],
    dry_run: bool,
    stdout,
    style,
) -> int:
    """
    Soft-delete catalog rows produced by hashed PrimeCare imports that no longer appear
    on the tariff sheet. Leaves explicit spreadsheet codes (no hash suffix) and all
    other hospital-managed codes untouched.
    """
    n = 0
    qs = ProcedureCatalog.objects.filter(is_deleted=False, is_active=True).only(
        'id', 'code', 'name'
    )
    to_clear: list[str] = []
    preview_limit = 30
    for obj in qs:
        code = (obj.code or '').strip()
        if not procedure_code_looks_generated_from_import(code):
            continue
        if code in canonical_codes:
            continue
        to_clear.append(str(obj.pk))
        n += 1
        if dry_run and n <= preview_limit:
            stdout.write(style.WARNING(f'  [prune] would deactivate {code} | {obj.name[:80]}'))
    if dry_run and n > preview_limit:
        stdout.write(
            style.WARNING(f'  [prune] ... and {n - preview_limit} more (not listed)')
        )
    if not dry_run and to_clear:
        ProcedureCatalog.objects.filter(pk__in=to_clear).update(
            is_deleted=True,
            is_active=False,
            modified=timezone.now(),
        )
    return n


def _stable_bundle_code(group_label: str, prefix: str = 'PRIME') -> str:
    pfx = re.sub(r'[^A-Za-z0-9_-]+', '', (prefix or 'PRIME').strip())[:16] or 'PRIME'
    slug = re.sub(r'[^A-Za-z0-9]+', '-', group_label.strip())[:50].strip('-') or 'PACK'
    h = hashlib.sha1(group_label.encode('utf-8')).hexdigest()[:6].upper()
    code = f'{pfx}-{slug}-{h}'.upper()
    return code[:80]


def _line_billing_code(bundle_code: str, idx: int, desc: str) -> str:
    part = re.sub(r'[^A-Za-z0-9]+', '', desc)[:24].upper() or 'LINE'
    raw = f'{bundle_code[:30]}-{part}-{idx + 1}'
    if len(raw) <= 80:
        return raw
    return f'{bundle_code[:40]}-L{idx + 1}'[:80]


def _auto_detect_header_row(ws, *, scan_rows: int = 30) -> int:
    """First row containing both a procedure/name column and a price column (titles vary by workbook)."""
    max_r = min(scan_rows, int(ws.max_row or scan_rows))
    max_c = min(int(ws.max_column or 26), 40)
    for r in range(1, max_r + 1):
        header_cells = [(c, _norm_str(_cell(ws, r, c))) for c in range(1, max_c + 1)]
        if not any(t.strip() for _, t in header_cells):
            continue
        auto = _find_header_columns(header_cells)
        proc_col = auto['procedure']
        cash_col = auto['cash']
        # Ignore banner rows where one merged cell falsely matches both (e.g. "… PROCEDURES … CASH …")
        if proc_col and cash_col and proc_col != cash_col:
            return r
    return 1


def _col_is_total_money_header(key_lc: str) -> bool:
    """True when this column holds line/invoice totals (not unit price)."""
    if not key_lc.strip():
        return False
    if 'total' in key_lc:
        return True
    # Net payable row headers sometimes appear up top on templates
    if 'payable' in key_lc or 'net amount' in key_lc:
        return True
    if 'quantity' in key_lc or key_lc.strip() == 'qty':
        return True
    return False


def _pick_unit_price_column(title_to_col: dict[str, int]) -> int | None:
    """
    Prefer unit / list prices. Avoid TOTAL AMOUNT: legacy heuristic used substring
    ``amount``, which matched 'total amount (ghs)' before 'unit price (ghs)'.
    """
    items = list(title_to_col.items())
    for key, col in items:
        if _col_is_total_money_header(key):
            continue
        if 'total amount' in key or 'grand total' in key or 'net amount' in key:
            continue
        if (
            'unit price' in key
            or 'unit rate' in key
            or 'list price' in key
            or ('unit' in key and 'price' in key)
        ):
            return col
    for key, col in items:
        if _col_is_total_money_header(key):
            continue
        if 'total amount' in key or 'grand total' in key:
            continue
        if 'price' in key:
            return col
    for key, col in items:
        if _col_is_total_money_header(key):
            continue
        if key.strip() in ('cash', 'ghs', 'cedis'):
            return col
        for needle in ('fee', 'rate', 'cash'):
            if needle in key:
                return col
    return None


def _find_header_columns(header_row_cells: list[tuple[int, str | None]]) -> dict[str, int]:
    """
    Map canonical keys to 1-based column index from lowercase header titles.
    """
    title_to_col: dict[str, int] = {}
    for col_idx, raw in header_row_cells:
        t = _norm_str(raw).lower()
        if not t:
            continue
        title_to_col[t] = col_idx

    def pick(candidates: tuple[str, ...], *, skip_price_total: bool = False) -> int | None:
        for c in candidates:
            for key, col in title_to_col.items():
                if c not in key:
                    continue
                if skip_price_total and _col_is_total_money_header(key):
                    continue
                if skip_price_total and c == 'amount' and ('total' in key or 'payable' in key):
                    continue
                return col
        return None

    cash_col = _pick_unit_price_column(title_to_col)
    if cash_col is None:
        cash_col = pick(
            ('cash', 'price', 'fee', 'rate', 'ghs', 'cedi'),
            skip_price_total=True,
        )

    return {
        'procedure': (
            pick(
                (
                    'procedure name',
                    'procedure',
                    'service',
                    'description',
                    'item',
                    'name',
                )
            )
            or 0
        ),
        'cash': cash_col or 0,
        'code': pick(('code', 'sku', 'item code', 'procedure code')) or 0,
        'bundle': pick(
            ('bundle', 'package', 'group', 'pack id', 'bundle id', 'group id')
        )
        or 0,
        'category': pick(('category', 'type', 'class')) or 0,
    }


class Command(BaseCommand):
    help = (
        'Import PrimeCare procedure prices from .xlsx into ProcedureCatalog and optional '
        'CashierChargeBundle. When the workbook has a tab named "Cash Procedures Only", '
        'that tariff list is used by default (not the active sheet). Invoice-style tabs '
        'use UNIT PRICE, not TOTAL AMOUNT, as the price column.'
    )

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str, help='Path to the Excel workbook')
        parser.add_argument(
            '--sheet',
            type=int,
            default=None,
            help='Zero-based worksheet index (omit with default tariff selection)',
        )
        parser.add_argument(
            '--sheet-name',
            type=str,
            default=None,
            help='Exact worksheet tab name (overrides --sheet when set)',
        )
        parser.add_argument(
            '--use-active-sheet',
            action='store_true',
            help='Use the workbook\'s active sheet; by default imports from tab '
            '"Cash Procedures Only" when that tab exists (PrimeCare tariff list).',
        )
        parser.add_argument(
            '--header-row',
            type=int,
            default=None,
            help='1-based header row (default: auto-detect rows with procedure + price columns)',
        )
        parser.add_argument('--data-start-row', type=int, default=None, help='1-based first data row (default: header+1)')
        parser.add_argument('--dry-run', action='store_true', help='Parse and print planned changes only')
        parser.add_argument(
            '--skip-prune-generated-orphans',
            action='store_true',
            help='After import, skip soft-deleting hashed PrimeCare catalog rows that no longer '
            'appear on this tariff sheet (normally only on "Cash Procedures Only" with no bundle groups).',
        )
        parser.add_argument(
            '--prefix-bundle',
            type=str,
            default='PRIME',
            help='Prefix for generated bundle codes (slug + hash appended)',
        )

        parser.add_argument(
            '--col-procedure',
            type=str,
            default=None,
            help='Procedure/name column: Excel letters e.g. A, or 1-based index',
        )
        parser.add_argument(
            '--col-cash',
            type=str,
            default=None,
            help='Cash price column: letters or 1-based index',
        )
        parser.add_argument(
            '--col-code',
            type=str,
            default=None,
            help='Optional explicit procedure code column: letters or 1-based index',
        )
        parser.add_argument(
            '--col-bundle-group',
            type=str,
            default=None,
            help='Bundle / package grouping column (same value = one multi-line bundle). Letters or index.',
        )
        parser.add_argument(
            '--col-category',
            type=str,
            default=None,
            help='Optional category column mapped to ProcedureCatalog category choices',
        )

    def handle(self, *args, **options):
        path = Path(options['xlsx_path']).expanduser().resolve()
        if not path.is_file():
            raise CommandError(f'File not found: {path}')

        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise CommandError('openpyxl is required') from e

        # read_only breaks max_row on some workbooks; sheets are usually small for tariff imports
        wb = load_workbook(path, read_only=False, data_only=True)
        sheet_name = (options.get('sheet_name') or '').strip()
        sheet_idx = options['sheet']
        std_tariff = 'Cash Procedures Only'
        try:
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise CommandError(f'Unknown worksheet name "{sheet_name}". Choices: {wb.sheetnames!r}')
                ws = wb[sheet_name]
            elif sheet_idx is not None:
                ws = wb.worksheets[sheet_idx]
            elif not options['use_active_sheet'] and std_tariff in wb.sheetnames:
                ws = wb[std_tariff]
            else:
                ws = wb.active
        except IndexError as e:
            raise CommandError(f'Worksheet index {sheet_idx} not found') from e

        self.stdout.write(self.style.NOTICE(f'Using worksheet: {ws.title}'))

        header_row = options['header_row']
        if header_row is None:
            header_row = _auto_detect_header_row(ws)
        data_start = options['data_start_row'] or (header_row + 1)

        def parse_col_opt(val: str | None) -> int:
            if not val:
                return 0
            s = val.strip()
            if s.isdigit():
                return int(s)
            return _letters_to_index(s)

        col_proc = parse_col_opt(options['col_procedure'])
        col_cash = parse_col_opt(options['col_cash'])
        col_code = parse_col_opt(options['col_code'])
        col_bundle = parse_col_opt(options['col_bundle_group'])
        col_cat = parse_col_opt(options['col_category'])

        header_cells: list[tuple[int, str | None]] = []
        max_col = ws.max_column or 26
        for c in range(1, max_col + 1):
            header_cells.append((c, _cell(ws, header_row, c)))

        auto = _find_header_columns([(a, _norm_str(b) if b is not None else None) for a, b in header_cells])
        if not col_proc:
            col_proc = auto['procedure']
        if not col_cash:
            col_cash = auto['cash']
        if not col_code:
            col_code = auto['code']
        if not col_bundle:
            col_bundle = auto['bundle']
        if not col_cat:
            col_cat = auto['category']

        if not col_proc or not col_cash:
            raise CommandError(
                'Could not detect procedure/name and cash columns. '
                'Set --col-procedure and --col-cash (Excel letters or 1-based indexes). '
                f'Auto-detected: procedure={col_proc}, cash={col_cash}, bundle={col_bundle}'
            )

        singles: list[dict] = []
        bundles: dict[str, list[dict]] = defaultdict(list)

        allowed_cats = {k for k, _ in ProcedureCatalog.PROCEDURE_CATEGORIES}

        row = data_start
        while row <= (ws.max_row or row + 1):
            name = _norm_str(_cell(ws, row, col_proc))
            if not name:
                row += 1
                continue
            cash_amt = _parse_money(_cell(ws, row, col_cash))
            if cash_amt is None or cash_amt < 0:
                self.stderr.write(self.style.WARNING(f'Row {row}: skip "{name}" — invalid cash'))
                row += 1
                continue
            raw_code = _norm_str(_cell(ws, row, col_code)) if col_code else ''
            cat_raw = (_norm_str(_cell(ws, row, col_cat)).lower().replace(' ', '_') if col_cat else '')
            cat = cat_raw if cat_raw in allowed_cats else 'other'

            bundle_key = _norm_str(_cell(ws, row, col_bundle)) if col_bundle else ''
            payload = {
                'name': name,
                'cash': cash_amt,
                'code_hint': raw_code,
                'category': cat,
                'row': row,
            }
            if bundle_key:
                bundles[bundle_key].append(payload)
            else:
                singles.append(payload)
            row += 1

        self.stdout.write(
            self.style.NOTICE(
                f'Parsed {len(singles)} single procedure(s), {len(bundles)} bundle group(s) '
                f'(columns: procedure={col_proc}, cash={col_cash}, bundle={col_bundle})'
            )
        )

        canonical_codes = {_stable_procedure_code(p['code_hint'], p['name']) for p in singles}
        prune_generated = (
            not options['skip_prune_generated_orphans']
            and ws.title == std_tariff
            and not bundles
            and bool(canonical_codes)
        )

        if options['dry_run']:
            for p in singles[:30]:
                code = _stable_procedure_code(p['code_hint'], p['name'])
                self.stdout.write(f'  [single] {code} | {p["name"]} | GHS {p["cash"]}')
            if len(singles) > 30:
                self.stdout.write(f'  ... and {len(singles) - 30} more singles')
            for gkey, lines in list(bundles.items())[:20]:
                self.stdout.write(f'  [bundle] {gkey!r} ({len(lines)} lines)')
                for ln in lines:
                    self.stdout.write(f'      - {ln["name"]} GHS {ln["cash"]}')
            if len(bundles) > 20:
                self.stdout.write(f'  ... and {len(bundles) - 20} more bundles')
            if prune_generated:
                np = _prune_orphan_generated_procedures(
                    canonical_codes=canonical_codes,
                    dry_run=True,
                    stdout=self.stdout,
                    style=self.style,
                )
                self.stdout.write(
                    self.style.NOTICE(
                        f'Dry-run prune: {np} hashed procedure row(s) in DB are not on this sheet '
                        f'(run import without --dry-run to deactivate them).'
                    )
                )
            return

        created_p, updated_p = 0, 0
        created_b, updated_b = 0, 0
        pruned = 0

        with transaction.atomic():
            for p in singles:
                code = _stable_procedure_code(p['code_hint'], p['name'])
                defaults = {
                    'name': p['name'][:200],
                    'category': p['category'],
                    'price': p['cash'],
                    'cash_price': p['cash'],
                    'is_active': True,
                    'is_deleted': False,
                }
                obj, was_created = ProcedureCatalog.objects.update_or_create(
                    code=code,
                    defaults=defaults,
                )
                if was_created:
                    created_p += 1
                else:
                    updated_p += 1

            for gkey, lines in bundles.items():
                if not lines:
                    continue
                bundle_code = _stable_bundle_code(gkey, options['prefix_bundle'])
                json_lines = []
                for i, ln in enumerate(lines):
                    desc = ln['name'][:200]
                    bc = _line_billing_code(bundle_code, i, desc)
                    json_lines.append(
                        {
                            'billing_code': bc,
                            'description': desc,
                            'amount_cash': str(ln['cash']),
                            'amount_insurance': None,
                            'amount_corporate': None,
                        }
                    )
                label = gkey[:200] if len(gkey) <= 200 else gkey[:197] + '...'
                _, was_created = CashierChargeBundle.objects.update_or_create(
                    bundle_code=bundle_code,
                    defaults={
                        'label': label,
                        'lines': json_lines,
                        'is_active': True,
                    },
                )
                if was_created:
                    created_b += 1
                else:
                    updated_b += 1

            if prune_generated:
                pruned = _prune_orphan_generated_procedures(
                    canonical_codes=canonical_codes,
                    dry_run=False,
                    stdout=self.stdout,
                    style=self.style,
                )
        prune_msg = f'; pruned orphan generated catalog row(s): {pruned}' if prune_generated else ''
        self.stdout.write(
            self.style.SUCCESS(
                f'Done. ProcedureCatalog created={created_p}, updated={updated_p}; '
                f'CashierChargeBundle created={created_b}, updated={updated_b}{prune_msg}.'
            )
        )
