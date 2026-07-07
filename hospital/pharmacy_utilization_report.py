"""
Pharmacy utilization analytics — shared report context and export builders.
"""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime, time, timedelta
from decimal import Decimal
from io import BytesIO

from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone

from .models import Drug
from .pharmacy_consumption_estimate import (
    DEFAULT_COVER_ALERT_DAYS,
    DEFAULT_MOVEMENT_WINDOW_DAYS,
    TOP_DRUG_RANK_LIMIT,
    compute_pharmacy_drug_movement_metrics,
    daily_global_outflow_counts,
    drug_ids_for_utilization_report,
    global_outflow_by_channel_since,
    parse_positive_int,
    top_expensive_formulary_drugs,
    top_moving_drugs_ranked,
)

HEADER_TEAL = '0D9488'
HEADER_LIGHT = '14B8A6'
LABEL_BG = 'CCFBF1'
BORDER_HEX = 'D1D5DB'
ZEBRA = 'F0FDFA'


def build_pharmacy_utilization_report_context(request, *, include_chart_data: bool = True) -> dict:
    """Build utilization report context from GET params (shared by HTML view and exports)."""
    movement_window_days = parse_positive_int(
        request.GET.get('movement_days'), DEFAULT_MOVEMENT_WINDOW_DAYS, max_val=365
    )
    cover_alert_days = parse_positive_int(
        request.GET.get('cover_alert_days'), DEFAULT_COVER_ALERT_DAYS, max_val=180
    )
    include_catalog = (request.GET.get('scope') or 'consumed').strip() == 'with_stock'
    query = (request.GET.get('q') or '').strip()
    category_filter = (request.GET.get('category') or '').strip()
    sort_key = (request.GET.get('sort') or 'consumed_desc').strip()
    period = (request.GET.get('period') or 'rolling').strip()
    if period not in ('rolling', 'calendar_month'):
        period = 'rolling'

    now = timezone.now()
    tz = timezone.get_current_timezone()
    if period == 'calendar_month':
        today = timezone.localdate()
        month_start = today.replace(day=1)
        start_dt = timezone.make_aware(datetime.combine(month_start, time.min), tz)
        end_dt = now
        window_label = (
            f'Calendar month to date ({month_start.strftime("%b %d")} – '
            f'{timezone.localtime(end_dt, tz).strftime("%b %d, %Y")})'
        )
    else:
        end_dt = now
        start_dt = end_dt - timedelta(days=movement_window_days)
        window_label = f'Rolling last {movement_window_days} days'

    drug_ids = drug_ids_for_utilization_report(
        start_dt,
        include_stock_catalog=include_catalog,
        end_dt=end_dt,
    )
    if query:
        drug_ids = list(
            Drug.objects.filter(
                Q(name__icontains=query) | Q(generic_name__icontains=query),
                id__in=drug_ids,
            ).values_list('id', flat=True)
        )
    if category_filter:
        drug_ids = list(
            Drug.objects.filter(
                id__in=drug_ids,
                category=category_filter,
            ).values_list('id', flat=True)
        )

    if period == 'calendar_month':
        metrics = compute_pharmacy_drug_movement_metrics(
            drug_ids,
            movement_window_days=movement_window_days,
            cover_alert_days=cover_alert_days,
            window_start=start_dt,
            window_end=end_dt,
        )
    else:
        metrics = compute_pharmacy_drug_movement_metrics(
            drug_ids,
            movement_window_days=movement_window_days,
            cover_alert_days=cover_alert_days,
        )

    row_pairs: list[tuple] = []
    for did, m in metrics.items():
        if not include_catalog and int(m.get('total_out_window', 0) or 0) <= 0:
            continue
        row_pairs.append((did, m))

    drug_map = {
        d.id: d
        for d in Drug.objects.filter(id__in=[p[0] for p in row_pairs]).select_related()
    }

    def _sort_key(item):
        did, m = item
        d = drug_map.get(did)
        name = (d.name if d else '').lower()
        outv = int(m.get('total_out_window', 0) or 0)
        risk = 1 if m.get('is_runout_risk') else 0
        sugg = int(m.get('suggested_order_qty', 0) or 0)
        if sort_key == 'name':
            return (name, -outv)
        if sort_key == 'consumed_asc':
            return (outv, name)
        if sort_key == 'risk':
            return (-risk, -outv, name)
        if sort_key == 'suggest_desc':
            return (-sugg, -outv, name)
        return (-outv, name)

    row_pairs.sort(key=_sort_key)

    utilization_rows = [
        {'drug': drug_map[did], 'm': m}
        for did, m in row_pairs
        if drug_map.get(did) is not None
    ]
    total_units_consumed = sum(int(r['m'].get('total_out_window', 0) or 0) for r in utilization_rows)
    drugs_with_consumption = sum(
        1 for r in utilization_rows if int(r['m'].get('total_out_window', 0) or 0) > 0
    )

    effective_window_days = movement_window_days
    if utilization_rows:
        effective_window_days = int(utilization_rows[0]['m'].get('window_days') or movement_window_days)

    global_channel = global_outflow_by_channel_since(start_dt, end_dt=end_dt)
    daily_series = daily_global_outflow_counts(start_dt, end_dt=end_dt)
    gc_total = max(1, int(global_channel.get('total') or 0))
    pct_rx_g = 100.0 * int(global_channel.get('rx', 0)) / gc_total
    pct_walk_g = 100.0 * int(global_channel.get('walk_in', 0)) / gc_total
    pct_loss_g = 100.0 * int(global_channel.get('loss', 0)) / gc_total

    cat_lookup = dict(Drug.CATEGORIES)
    cat_agg = defaultdict(
        lambda: {'units': 0, 'skus': 0, 'at_risk': 0, 'suggest_sum': 0}
    )
    for r in utilization_rows:
        m = r['m']
        d = r['drug']
        key = (d.category or '').strip() or '__none__'
        cat_agg[key]['units'] += int(m.get('total_out_window', 0) or 0)
        cat_agg[key]['skus'] += 1
        if m.get('is_runout_risk'):
            cat_agg[key]['at_risk'] += 1
        cat_agg[key]['suggest_sum'] += int(m.get('suggested_order_qty', 0) or 0)

    category_rollup = []
    for code, ag in sorted(cat_agg.items(), key=lambda x: -x[1]['units']):
        label = 'Uncategorized' if code == '__none__' else cat_lookup.get(code, code)
        category_rollup.append(
            {
                'code': code,
                'label': label,
                'units': ag['units'],
                'skus': ag['skus'],
                'at_risk': ag['at_risk'],
                'suggest_sum': ag['suggest_sum'],
            }
        )

    top_cat_chart = category_rollup[:10]
    other_units = sum(c['units'] for c in category_rollup[10:])
    if other_units > 0:
        top_cat_chart = list(top_cat_chart) + [
            {
                'code': '_other',
                'label': 'Other categories',
                'units': other_units,
                'skus': 0,
                'at_risk': 0,
                'suggest_sum': 0,
            }
        ]

    by_consume = sorted(
        utilization_rows,
        key=lambda r: -int(r['m'].get('total_out_window', 0) or 0),
    )
    pareto_rows = []
    cum = 0
    for r in by_consume[:TOP_DRUG_RANK_LIMIT]:
        u = int(r['m'].get('total_out_window', 0) or 0)
        cum += u
        pct = (100.0 * cum / float(total_units_consumed)) if total_units_consumed else 0.0
        name = r['drug'].name
        if len(name) > 32:
            name = name[:29] + '…'
        pareto_rows.append(
            {
                'drug_id': r['drug'].id,
                'label': name,
                'units': u,
                'cumulative_pct': round(pct, 1),
            }
        )

    top_expensive_drugs = top_expensive_formulary_drugs(limit=TOP_DRUG_RANK_LIMIT)
    if period == 'calendar_month':
        top_moving_drugs = top_moving_drugs_ranked(
            start_dt,
            end_dt=end_dt,
            movement_window_days=movement_window_days,
            window_start=start_dt,
            window_end=end_dt,
            limit=TOP_DRUG_RANK_LIMIT,
        )
    else:
        top_moving_drugs = top_moving_drugs_ranked(
            start_dt,
            end_dt=end_dt,
            movement_window_days=movement_window_days,
            limit=TOP_DRUG_RANK_LIMIT,
        )

    risk_candidates = [r for r in utilization_rows if r['m'].get('is_runout_risk')]
    risk_candidates.sort(
        key=lambda r: (
            -int(r['m'].get('suggested_order_qty', 0) or 0),
            -int(r['m'].get('total_out_window', 0) or 0),
        )
    )
    risk_portfolio = risk_candidates[:25]
    risk_sku_count = len(risk_candidates)
    risk_suggest_total = sum(int(r['m'].get('suggested_order_qty', 0) or 0) for r in risk_candidates)

    scope_label = (
        'Drugs with consumption + in-stock zero movement'
        if include_catalog
        else 'Drugs with consumption only'
    )
    sort_labels = {
        'consumed_desc': 'Highest consumption',
        'consumed_asc': 'Lowest consumption',
        'risk': 'Run-out risk first',
        'suggest_desc': 'Largest suggest first',
        'name': 'Name A–Z',
    }
    category_filter_label = (
        dict(Drug.CATEGORIES).get(category_filter, category_filter)
        if category_filter
        else 'All categories'
    )

    context = {
        'utilization_rows': utilization_rows,
        'movement_window_days': movement_window_days,
        'effective_window_days': effective_window_days,
        'cover_alert_days': cover_alert_days,
        'query': query,
        'category_filter': category_filter,
        'category_filter_label': category_filter_label,
        'drug_categories': Drug.CATEGORIES,
        'scope': 'with_stock' if include_catalog else 'consumed',
        'scope_label': scope_label,
        'sort': sort_key,
        'sort_label': sort_labels.get(sort_key, sort_key),
        'period': period,
        'period_label': 'Calendar month to date' if period == 'calendar_month' else 'Rolling window',
        'window_label': window_label,
        'total_units_consumed': total_units_consumed,
        'sku_count': len(utilization_rows),
        'drugs_with_consumption': drugs_with_consumption,
        'window_start': start_dt,
        'window_end': end_dt,
        'global_channel': global_channel,
        'pct_rx_global': round(pct_rx_g, 1),
        'pct_walk_in_global': round(pct_walk_g, 1),
        'pct_loss_global': round(pct_loss_g, 1),
        'daily_series': daily_series,
        'category_rollup': category_rollup,
        'pareto_rows': pareto_rows,
        'top_expensive_drugs': top_expensive_drugs,
        'top_moving_drugs': top_moving_drugs,
        'top_drug_rank_limit': TOP_DRUG_RANK_LIMIT,
        'risk_portfolio': risk_portfolio,
        'risk_sku_count': risk_sku_count,
        'risk_suggest_total': risk_suggest_total,
        'generated_at': now,
    }

    if include_chart_data:
        context['utilization_chart_data'] = {
            'meta': {
                'period': period,
                'windowDays': effective_window_days,
                'windowLabel': window_label,
            },
            'channelDoughnut': {
                'labels': [
                    'Inpatient / Rx dispensing',
                    'Walk-in OTC (dispensed)',
                    'Recorded stock losses',
                ],
                'values': [
                    int(global_channel.get('rx', 0)),
                    int(global_channel.get('walk_in', 0)),
                    int(global_channel.get('loss', 0)),
                ],
            },
            'dailyTrend': {
                'labels': [d['date_iso'] for d in daily_series],
                'rx': [d['rx'] for d in daily_series],
                'walk_in': [d['walk_in'] for d in daily_series],
                'loss': [d['loss'] for d in daily_series],
                'total': [d['total'] for d in daily_series],
            },
            'categoryBars': {
                'labels': [c['label'] for c in top_cat_chart],
                'values': [c['units'] for c in top_cat_chart],
            },
            'pareto': {
                'labels': [p['label'] for p in pareto_rows],
                'cumulativePct': [p['cumulative_pct'] for p in pareto_rows],
            },
        }

    return context


def _export_filename_stem(ctx: dict) -> str:
    ts = timezone.localtime(ctx['generated_at']).strftime('%Y%m%d_%H%M')
    return f'pharmacy_utilization_{ts}'


def pharmacy_utilization_pdf_response(ctx: dict, hospital_name: str = '') -> HttpResponse:
    """Multi-section PDF: summary, top expensive, top movers, category rollup, detail."""
    try:
        from xml.sax.saxutils import escape

        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return HttpResponse(
            'ReportLab is required for PDF export. Please install it and try again.',
            status=500,
        )

    def _cell(val, max_len=40):
        s = '' if val is None else str(val).replace('\r', ' ').replace('\n', ' ')
        return s[:max_len] if len(s) > max_len else s

    def _section(title: str, subtitle: str = ''):
        elements.append(Paragraph(f'<b>{escape(title)}</b>', section_style))
        if subtitle:
            elements.append(Paragraph(escape(subtitle), meta_style))
        elements.append(Spacer(1, 0.08 * inch))

    def _table(data, col_widths, header_hex=HEADER_TEAL):
        if len(data) <= 1:
            data = data + [['—'] * (len(data[0]) if data else 1)]
        t = Table(data, repeatRows=1, colWidths=col_widths)
        t.setStyle(
            TableStyle(
                [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{header_hex}')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 7),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdfa')]),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]
            )
        )
        elements.append(t)
        elements.append(Spacer(1, 0.14 * inch))

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=0.4 * inch,
        bottomMargin=0.4 * inch,
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
    )
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'UtilTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor(f'#{HEADER_TEAL}'),
        spaceAfter=4,
        alignment=TA_CENTER,
    )
    section_style = ParagraphStyle(
        'UtilSection',
        parent=styles['Heading2'],
        fontSize=11,
        textColor=colors.HexColor(f'#{HEADER_TEAL}'),
        spaceBefore=4,
        spaceAfter=2,
    )
    meta_style = ParagraphStyle(
        'UtilMeta',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_LEFT,
    )

    gen_str = timezone.localtime(ctx['generated_at']).strftime('%Y-%m-%d %H:%M')
    elements.append(Paragraph('<b>Pharmacy utilization analytics</b>', title_style))
    if hospital_name:
        elements.append(Paragraph(escape(hospital_name), meta_style))
    elements.append(Paragraph(f'Generated: {gen_str}', meta_style))
    elements.append(Spacer(1, 0.1 * inch))

    summary_lines = [
        ('Reporting window', ctx['window_label']),
        ('Period type', ctx['period_label']),
        ('Units out (period)', f"{ctx['total_units_consumed']:,}"),
        ('Drugs with use', f"{ctx['drugs_with_consumption']:,}"),
        ('Table rows', f"{ctx['sku_count']:,}"),
        ('Channel — Rx', f"{ctx['pct_rx_global']:.1f}%"),
        ('Channel — Walk-in', f"{ctx['pct_walk_in_global']:.1f}%"),
        ('Channel — Losses', f"{ctx['pct_loss_global']:.1f}%"),
        ('Run-out risk SKUs', f"{ctx['risk_sku_count']:,}"),
        ('Scope', ctx['scope_label']),
        ('Sort', ctx['sort_label']),
        ('Category filter', ctx['category_filter_label']),
    ]
    if ctx.get('query'):
        summary_lines.append(('Search', ctx['query']))

    sum_data = [['Metric', 'Value']]
    sum_data.extend([[a, b] for a, b in summary_lines])
    _table(sum_data, [2.2 * inch, 8.5 * inch])

    # Top expensive
    _section(
        f"Top {ctx['top_drug_rank_limit']} expensive drugs",
        'Ranked by unit selling price (formulary-wide)',
    )
    exp_header = ['#', 'Drug', 'Strength/Form', 'Unit price (GHS)', 'Cost (GHS)', 'On hand', 'Stock value (GHS)']
    exp_data = [exp_header]
    for row in ctx['top_expensive_drugs']:
        d = row['drug']
        exp_data.append(
            [
                str(row['rank']),
                _cell(d.name, 34),
                _cell(f'{d.strength} {d.form}'.strip(), 22),
                f"{row['unit_price']:.2f}",
                f"{row['cost_price']:.2f}",
                f"{row['total_on_hand']:,}",
                f"{row['stock_value']:.2f}",
            ]
        )
    _table(
        exp_data,
        [0.35 * inch, 2.0 * inch, 1.2 * inch, 0.95 * inch, 0.85 * inch, 0.65 * inch, 1.0 * inch],
    )

    elements.append(PageBreak())

    # Top movers
    _section(
        f"Top {ctx['top_drug_rank_limit']} fast-moving drugs",
        ctx['window_label'],
    )
    mov_header = [
        '#',
        'Drug',
        'Strength/Form',
        f"{ctx['effective_window_days']}d units",
        'Avg/day',
        'On hand',
        'Cover',
        'Rx',
        'Walk-in',
        'Loss',
    ]
    mov_data = [mov_header]
    for row in ctx['top_moving_drugs']:
        d = row['drug']
        m = row['m']
        mov_data.append(
            [
                str(row['rank']),
                _cell(d.name, 34),
                _cell(f'{d.strength} {d.form}'.strip(), 22),
                f"{int(m['total_out_window']):,}",
                m['avg_daily_out_display'],
                f"{int(m['total_on_hand']):,}",
                m['days_cover_display'],
                f"{int(m['out_rx']):,}",
                f"{int(m['out_walk_in']):,}",
                f"{int(m['out_loss']):,}",
            ]
        )
    _table(
        mov_data,
        [
            0.35 * inch,
            1.85 * inch,
            1.1 * inch,
            0.75 * inch,
            0.55 * inch,
            0.55 * inch,
            0.55 * inch,
            0.55 * inch,
            0.55 * inch,
            0.55 * inch,
        ],
    )

    # Category rollup
    if ctx['category_rollup']:
        _section('Roll-up by drug category')
        cat_header = ['Category', 'Units out', 'SKUs', 'At risk', 'Sum of suggest']
        cat_data = [cat_header]
        for c in ctx['category_rollup']:
            cat_data.append(
                [
                    _cell(c['label'], 48),
                    f"{c['units']:,}",
                    f"{c['skus']:,}",
                    f"{c['at_risk']:,}",
                    f"{c['suggest_sum']:,}",
                ]
            )
        _table(cat_data, [3.5 * inch, 1.0 * inch, 0.8 * inch, 0.8 * inch, 1.1 * inch])

    # Utilization detail
    if ctx['utilization_rows']:
        elements.append(PageBreak())
        _section(
            'Drug utilization detail',
            f'{len(ctx["utilization_rows"])} drugs matching current filters',
        )
        det_header = [
            'Drug',
            'Strength/Form',
            f"{ctx['effective_window_days']}d total",
            'Rx',
            'Walk-in',
            'Loss',
            'Avg/d',
            'OH',
            'Reorder',
            'Cover',
            'Suggest',
            'Risk',
        ]
        det_data = [det_header]
        for row in ctx['utilization_rows']:
            d = row['drug']
            m = row['m']
            det_data.append(
                [
                    _cell(d.name, 32),
                    _cell(f'{d.strength} {d.form}'.strip(), 18),
                    f"{int(m['total_out_window']):,}",
                    f"{int(m['out_rx']):,}",
                    f"{int(m['out_walk_in']):,}",
                    f"{int(m['out_loss']):,}",
                    m['avg_daily_out_display'],
                    f"{int(m['total_on_hand']):,}",
                    f"{int(m['reorder_point']):,}",
                    m['days_cover_display'],
                    f"{int(m['suggested_order_qty']):,}",
                    'Yes' if m.get('is_runout_risk') else '',
                ]
            )
        _table(
            det_data,
            [
                1.6 * inch,
                0.95 * inch,
                0.62 * inch,
                0.48 * inch,
                0.52 * inch,
                0.45 * inch,
                0.48 * inch,
                0.42 * inch,
                0.52 * inch,
                0.48 * inch,
                0.52 * inch,
                0.38 * inch,
            ],
        )

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_export_filename_stem(ctx)}.pdf"'
    return response


def pharmacy_utilization_excel_response(ctx: dict, hospital_name: str = '') -> HttpResponse:
    """Multi-sheet Excel: summary, top expensive, top movers, categories, detail, at-risk."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            'openpyxl is required for Excel export. Please install it and try again.',
            status=500,
        )

    thin = Side(style='thin', color=BORDER_HEX)
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color=HEADER_TEAL, end_color=HEADER_TEAL, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    zebra_fill = PatternFill(start_color=ZEBRA, end_color=ZEBRA, fill_type='solid')
    label_fill = PatternFill(start_color=LABEL_BG, end_color=LABEL_BG, fill_type='solid')

    gen_str = timezone.localtime(ctx['generated_at']).strftime('%Y-%m-%d %H:%M:%S')
    wb = Workbook()

    def _style_header_row(ws, row_num: int, n_cols: int):
        for col in range(1, n_cols + 1):
            c = ws.cell(row_num, col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = grid

    def _write_data_rows(ws, start_row: int, rows: list[list], money_cols: set | None = None, int_cols: set | None = None):
        money_cols = money_cols or set()
        int_cols = int_cols or set()
        for idx, vals in enumerate(rows):
            r = start_row + idx
            zebra = idx % 2 == 1
            for col, val in enumerate(vals, 1):
                c = ws.cell(r, col, val)
                c.border = grid
                c.alignment = Alignment(vertical='center', wrap_text=True)
                if zebra:
                    c.fill = zebra_fill
                if col in money_cols and isinstance(val, (int, float, Decimal)):
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal='right', vertical='center')
                if col in int_cols and isinstance(val, (int, float)):
                    c.number_format = '#,##0'
                    c.alignment = Alignment(horizontal='right', vertical='center')

    # ----- Summary -----
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_properties.tabColor = HEADER_LIGHT
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Pharmacy utilization analytics'
    ws['A1'].font = Font(bold=True, size=20, color=HEADER_TEAL)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 36
    if hospital_name:
        ws.merge_cells('A2:F2')
        ws['A2'] = hospital_name
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(size=12, color='374151')
        meta_start = 4
    else:
        meta_start = 3

    meta = [
        ('Generated', gen_str),
        ('Reporting window', ctx['window_label']),
        ('Period', ctx['period_label']),
        ('Units out (period)', ctx['total_units_consumed']),
        ('Drugs with use', ctx['drugs_with_consumption']),
        ('Table rows', ctx['sku_count']),
        ('Channel Rx %', ctx['pct_rx_global']),
        ('Channel Walk-in %', ctx['pct_walk_in_global']),
        ('Channel Losses %', ctx['pct_loss_global']),
        ('Run-out risk SKUs', ctx['risk_sku_count']),
        ('Suggested restock (at-risk sum)', ctx['risk_suggest_total']),
        ('Scope', ctx['scope_label']),
        ('Sort', ctx['sort_label']),
        ('Category filter', ctx['category_filter_label']),
        ('Search', ctx.get('query') or '—'),
    ]
    for i, (label, value) in enumerate(meta):
        r = meta_start + i
        c1 = ws.cell(r, 1, label)
        c1.font = Font(bold=True)
        c1.fill = label_fill
        c1.border = grid
        c2 = ws.cell(r, 2, value)
        c2.border = grid
        c2.alignment = Alignment(wrap_text=True)
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 48

    # ----- Top expensive -----
    ws_e = wb.create_sheet('Top expensive', 1)
    ws_e.sheet_properties.tabColor = HEADER_TEAL
    exp_headers = ['Rank', 'Drug', 'Generic', 'Strength', 'Form', 'Unit price (GHS)', 'Cost (GHS)', 'On hand', 'Stock value (GHS)']
    for col, h in enumerate(exp_headers, 1):
        ws_e.cell(1, col, h)
    _style_header_row(ws_e, 1, len(exp_headers))
    exp_rows = []
    for row in ctx['top_expensive_drugs']:
        d = row['drug']
        exp_rows.append(
            [
                row['rank'],
                d.name,
                d.generic_name or '',
                d.strength,
                d.form,
                float(row['unit_price']),
                float(row['cost_price']),
                row['total_on_hand'],
                float(row['stock_value']),
            ]
        )
    _write_data_rows(ws_e, 2, exp_rows, money_cols={6, 7, 9}, int_cols={8})
    ws_e.auto_filter.ref = f'A1:{get_column_letter(len(exp_headers))}{max(1, 1 + len(exp_rows))}'
    ws_e.freeze_panes = 'A2'
    for i, w in enumerate([6, 28, 22, 12, 12, 14, 14, 10, 16], 1):
        ws_e.column_dimensions[get_column_letter(i)].width = w

    # ----- Top movers -----
    ws_m = wb.create_sheet('Top movers', 2)
    ws_m.sheet_properties.tabColor = HEADER_LIGHT
    mov_headers = [
        'Rank',
        'Drug',
        'Generic',
        'Strength',
        'Form',
        f"{ctx['effective_window_days']}d units",
        'Avg/day',
        'On hand',
        'Cover',
        'Rx',
        'Walk-in',
        'Loss',
        'Run-out risk',
    ]
    for col, h in enumerate(mov_headers, 1):
        ws_m.cell(1, col, h)
    _style_header_row(ws_m, 1, len(mov_headers))
    mov_rows = []
    for row in ctx['top_moving_drugs']:
        d = row['drug']
        m = row['m']
        mov_rows.append(
            [
                row['rank'],
                d.name,
                d.generic_name or '',
                d.strength,
                d.form,
                int(m['total_out_window']),
                float(m['avg_daily_out']),
                int(m['total_on_hand']),
                m['days_cover_display'],
                int(m['out_rx']),
                int(m['out_walk_in']),
                int(m['out_loss']),
                'Yes' if m.get('is_runout_risk') else '',
            ]
        )
    _write_data_rows(ws_m, 2, mov_rows, int_cols={6, 8, 10, 11, 12})
    ws_m.cell(2, 7).number_format = '0.00'
    ws_m.auto_filter.ref = f'A1:{get_column_letter(len(mov_headers))}{max(1, 1 + len(mov_rows))}'
    ws_m.freeze_panes = 'A2'
    for i, w in enumerate([6, 28, 22, 12, 12, 12, 10, 10, 10, 10, 10, 10, 12], 1):
        ws_m.column_dimensions[get_column_letter(i)].width = w

    # ----- Category rollup -----
    ws_c = wb.create_sheet('By category', 3)
    cat_headers = ['Category', 'Units out', 'SKUs', 'At risk', 'Sum of suggest']
    for col, h in enumerate(cat_headers, 1):
        ws_c.cell(1, col, h)
    _style_header_row(ws_c, 1, len(cat_headers))
    cat_rows = [[c['label'], c['units'], c['skus'], c['at_risk'], c['suggest_sum']] for c in ctx['category_rollup']]
    _write_data_rows(ws_c, 2, cat_rows, int_cols={2, 3, 4, 5})
    ws_c.auto_filter.ref = f'A1:E{max(1, 1 + len(cat_rows))}'
    ws_c.freeze_panes = 'A2'
    ws_c.column_dimensions['A'].width = 42
    for col in 'BCDE':
        ws_c.column_dimensions[col].width = 14

    # ----- Utilization detail -----
    ws_d = wb.create_sheet('Utilization detail', 4)
    det_headers = [
        'Drug',
        'Generic',
        'Strength',
        'Form',
        'Category',
        f"{ctx['effective_window_days']}d total",
        'Rx',
        'Walk-in',
        'Loss',
        'Avg/day',
        'On hand',
        'Reorder',
        'Cover',
        'Suggest',
        'Run-out risk',
    ]
    for col, h in enumerate(det_headers, 1):
        ws_d.cell(1, col, h)
    _style_header_row(ws_d, 1, len(det_headers))
    det_rows = []
    for row in ctx['utilization_rows']:
        d = row['drug']
        m = row['m']
        det_rows.append(
            [
                d.name,
                d.generic_name or '',
                d.strength,
                d.form,
                d.get_category_display(),
                int(m['total_out_window']),
                int(m['out_rx']),
                int(m['out_walk_in']),
                int(m['out_loss']),
                float(m['avg_daily_out']),
                int(m['total_on_hand']),
                int(m['reorder_point']),
                m['days_cover_display'],
                int(m['suggested_order_qty']),
                'Yes' if m.get('is_runout_risk') else '',
            ]
        )
    _write_data_rows(ws_d, 2, det_rows, int_cols={6, 7, 8, 9, 11, 12, 14})
    ws_d.auto_filter.ref = f'A1:{get_column_letter(len(det_headers))}{max(1, 1 + len(det_rows))}'
    ws_d.freeze_panes = 'A2'
    for i, w in enumerate([28, 22, 12, 12, 28, 12, 10, 10, 10, 10, 10, 10, 10, 10, 12], 1):
        ws_d.column_dimensions[get_column_letter(i)].width = w

    # ----- At-risk portfolio -----
    if ctx['risk_portfolio']:
        ws_r = wb.create_sheet('At-risk portfolio', 5)
        risk_headers = ['Drug', 'Strength', 'Form', 'On hand', 'Cover', 'Suggest', f"{ctx['effective_window_days']}d units"]
        for col, h in enumerate(risk_headers, 1):
            ws_r.cell(1, col, h)
        _style_header_row(ws_r, 1, len(risk_headers))
        risk_rows = []
        for row in ctx['risk_portfolio']:
            d = row['drug']
            m = row['m']
            risk_rows.append(
                [
                    d.name,
                    d.strength,
                    d.form,
                    int(m['total_on_hand']),
                    m['days_cover_display'],
                    int(m['suggested_order_qty']),
                    int(m['total_out_window']),
                ]
            )
        _write_data_rows(ws_r, 2, risk_rows, int_cols={4, 6, 7})
        ws_r.auto_filter.ref = f'A1:G{max(1, 1 + len(risk_rows))}'
        ws_r.freeze_panes = 'A2'
        for i, w in enumerate([28, 12, 12, 10, 10, 10, 12], 1):
            ws_r.column_dimensions[get_column_letter(i)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{_export_filename_stem(ctx)}.xlsx"'
    return response
