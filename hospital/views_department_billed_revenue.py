"""
Department-facing billed revenue (invoice lines by payer) — cash + insurance + corporate.

Uses Invoice.issued_at window; complements PaymentReceipt-based cash reports.
"""
from __future__ import annotations

import csv
import io

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from .decorators import role_required
from .services.department_billed_revenue_service import (
    DEPARTMENT_KEYS,
    build_department_report,
)
from .services.invoice_line_revenue_query import parse_ymd

_DEPT_LABEL = {
    'pharmacy': 'Pharmacy',
    'lab': 'Laboratory',
    'imaging': 'Imaging / radiology',
    'consultation': 'Consultation (OPD)',
}


def _parse_dept(raw: str) -> str:
    v = (raw or 'pharmacy').strip().lower()
    return v if v in DEPARTMENT_KEYS else 'pharmacy'


@login_required
@role_required(
    'pharmacist',
    'lab_technician',
    'radiologist',
    'doctor',
    'accountant',
    'senior_account_officer',
    'admin',
)
def department_billed_revenue_report(request):
    today = timezone.now().date()
    default_from = today.replace(day=1)
    date_from = parse_ymd(request.GET.get('date_from'), default_from)
    date_to = parse_ymd(request.GET.get('date_to'), today)
    if date_to < date_from:
        date_from, date_to = date_to, date_from

    dept = _parse_dept(request.GET.get('dept'))
    include_writeoff = request.GET.get('include_writeoff') in ('1', 'true', 'on', 'yes')

    payer_type = (request.GET.get('payer_type') or 'all').strip().lower()
    if payer_type not in ('all', 'cash', 'nhis', 'private', 'corporate', 'insurance'):
        payer_type = 'all'

    report = build_department_report(
        date_from=date_from,
        date_to=date_to,
        include_writeoff_period=include_writeoff,
        dept=dept,
        payer_type=payer_type if payer_type != 'all' else None,
    )
    payer_summary = report['payer_summary']
    top_services = report['top_services']

    export = (request.GET.get('export') or '').lower()
    if export == 'csv':
        return _export_csv(
            dept,
            payer_summary,
            top_services,
            date_from,
            date_to,
            include_writeoff,
            payer_type,
        )
    if export == 'xlsx':
        return _export_xlsx(
            dept,
            payer_summary,
            top_services,
            date_from,
            date_to,
            include_writeoff,
            payer_type,
        )

    return render(
        request,
        'hospital/department_billed_revenue.html',
        {
            'date_from': date_from,
            'date_to': date_to,
            'dept': dept,
            'dept_label': _DEPT_LABEL.get(dept, dept),
            'include_writeoff': include_writeoff,
            'payer_type': payer_type,
            'payer_summary': payer_summary,
            'top_services': top_services,
            'department_keys': sorted(DEPARTMENT_KEYS),
        },
    )


def _export_csv(dept, payer_summary, top_services, date_from, date_to, include_writeoff, payer_type):
    buf = io.StringIO()
    w = csv.writer(buf)
    label = _DEPT_LABEL.get(dept, dept)
    w.writerow([f'Billed revenue — {label} (invoice lines)'])
    w.writerow([f'Period {date_from.isoformat()} to {date_to.isoformat()}'])
    w.writerow([f'Issued-at basis · write-off invoices {"included" if include_writeoff else "excluded"}'])
    w.writerow([f'Payer filter: {payer_type}'])
    w.writerow([])
    w.writerow(['Summary by payer type', 'Lines', 'Billed (GHS)'])
    b = payer_summary['by_bucket']
    w.writerow(['Cash', '', f"{b['cash']:.2f}"])
    w.writerow(['Corporate', '', f"{b['corporate']:.2f}"])
    w.writerow(['NHIS', '', f"{b['nhis']:.2f}"])
    w.writerow(['Private insurance', '', f"{b['private']:.2f}"])
    w.writerow(['Other insurance', '', f"{b['insurance_other']:.2f}"])
    w.writerow(['Unknown / other payer', '', f"{b['unknown']:.2f}"])
    w.writerow(['All insurance (NHIS + private + other)', '', f"{payer_summary['insurance_all']:.2f}"])
    w.writerow(['Grand total', str(payer_summary['total_lines']), f"{payer_summary['total_billed']:.2f}"])
    w.writerow([])
    w.writerow(['Payer type (aggregate)', 'Lines', 'Billed (GHS)'])
    for r in payer_summary['rows']:
        w.writerow([r['payer_type'], r['line_count'], f"{r['billed']:.2f}"])
    w.writerow([])
    w.writerow(['Top services by billed amount'])
    w.writerow(['Code', 'Description', 'Category', 'Lines', 'Billed (GHS)'])
    for r in top_services:
        w.writerow(
            [
                r['service_code__code'],
                r['service_code__description'],
                r['service_code__category'],
                r['line_count'],
                f"{r['billed']:.2f}",
            ]
        )
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    safe_dept = dept.replace('/', '-')
    resp['Content-Disposition'] = (
        f'attachment; filename="dept_billed_{safe_dept}_{date_from.strftime("%Y%m%d")}_{date_to.strftime("%Y%m%d")}.csv"'
    )
    return resp


def _export_xlsx(dept, payer_summary, top_services, date_from, date_to, include_writeoff, payer_type):
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        return HttpResponse('Excel export requires openpyxl.', status=500)

    label = _DEPT_LABEL.get(dept, dept)
    wb = openpyxl.Workbook()

    ws0 = wb.active
    ws0.title = 'Summary'
    ws0.append([f'Billed revenue — {label}'])
    ws0.append([f'{date_from} to {date_to} · issued-at · payer filter: {payer_type}'])
    ws0.append([])
    ws0.append(['Bucket', 'Billed (GHS)'])
    b = payer_summary['by_bucket']
    bold = Font(bold=True)
    for title, val in [
        ('Cash', b['cash']),
        ('Corporate', b['corporate']),
        ('NHIS', b['nhis']),
        ('Private insurance', b['private']),
        ('Other insurance', b['insurance_other']),
        ('Unknown / other', b['unknown']),
        ('All insurance (NHIS + private + other)', payer_summary['insurance_all']),
    ]:
        ws0.append([title, float(val)])
    ws0.append([])
    ws0.append(['Total invoice lines', int(payer_summary['total_lines'])])
    ws0.append(['Total billed (GHS)', float(payer_summary['total_billed'])])
    ws0['A1'].font = Font(size=14, bold=True)

    ws1 = wb.create_sheet('By payer type')
    ws1.append(['Payer type', 'Lines', 'Billed (GHS)'])
    for c in ws1[1]:
        c.font = bold
    for r in payer_summary['rows']:
        ws1.append([r['payer_type'], r['line_count'], float(r['billed'])])

    ws2 = wb.create_sheet('By service code')
    ws2.append(['Code', 'Description', 'Category', 'Lines', 'Billed (GHS)'])
    for c in ws2[1]:
        c.font = bold
    for r in top_services:
        ws2.append(
            [
                r['service_code__code'],
                r['service_code__description'],
                r['service_code__category'],
                r['line_count'],
                float(r['billed']),
            ]
        )
    for ws in (ws1, ws2):
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = '#,##0.00'

    bio = io.BytesIO()
    wb.save(bio)
    safe_dept = dept.replace('/', '-')
    fname = f'dept_billed_{safe_dept}_{date_from.strftime("%Y%m%d")}_{date_to.strftime("%Y%m%d")}.xlsx'
    resp = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def department_billed_summary_context(*, date_from, date_to, dept: str, include_writeoff: bool = False):
    """
    Lightweight KPI dict for embedding on dashboards (month-to-date etc.).
    """
    dept = _parse_dept(dept)
    report = build_department_report(
        date_from=date_from,
        date_to=date_to,
        include_writeoff_period=include_writeoff,
        dept=dept,
        payer_type=None,
    )
    ps = report['payer_summary']
    return {
        'dept': dept,
        'total_billed': ps['total_billed'],
        'total_lines': ps['total_lines'],
        'by_bucket': ps['by_bucket'],
        'insurance_all': ps['insurance_all'],
    }
