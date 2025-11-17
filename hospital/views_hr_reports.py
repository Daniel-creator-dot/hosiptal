"""
HR Reports and Analytics Views
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Q, Sum, Avg
from django.utils import timezone
from datetime import date, datetime, time, timedelta
from decimal import Decimal
import csv
from io import BytesIO
from collections import defaultdict

from .models import Staff, Department
from .models_hr import (
    PayrollPeriod, Payroll, LeaveBalance, PerformanceReview, 
    TrainingRecord, StaffContract, PayGrade
)
from .models_advanced import LeaveRequest, Attendance
from .models_login_tracking import LoginHistory, SecurityAlert

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def is_hr_or_admin(user):
    """Check if user is HR or Admin"""
    # Allow superusers, staff users, and users in Admin/HR groups
    if user.is_superuser or user.is_staff:
        return True
    return user.groups.filter(name__in=['Admin', 'HR']).exists()


@login_required
def hr_reports_dashboard(request):
    """Main HR Reports Dashboard"""
    today = date.today()
    current_month_start = date(today.year, today.month, 1)
    
    # Staff Statistics
    total_staff = Staff.objects.filter(is_deleted=False, is_active=True).count()
    staff_by_department = Staff.objects.filter(
        is_deleted=False, is_active=True
    ).values('department__name').annotate(count=Count('id')).order_by('-count')
    
    staff_by_profession = Staff.objects.filter(
        is_deleted=False, is_active=True
    ).values('profession').annotate(count=Count('id')).order_by('-count')
    
    # Leave Statistics
    pending_leaves = LeaveRequest.objects.filter(
        status='pending', is_deleted=False
    ).count()
    
    approved_leaves_this_month = LeaveRequest.objects.filter(
        status='approved',
        approved_at__gte=current_month_start,
        is_deleted=False
    ).count()
    
    staff_on_leave_today = LeaveRequest.objects.filter(
        status='approved',
        start_date__lte=today,
        end_date__gte=today,
        is_deleted=False
    ).count()
    
    # Contract Expiry
    contracts_expiring_soon = StaffContract.objects.filter(
        end_date__gte=today,
        end_date__lte=today + timedelta(days=90),
        is_deleted=False,
        is_active=True
    ).count()
    
    # Birthday Statistics
    staff_with_birthdays = Staff.objects.filter(
        is_deleted=False,
        is_active=True,
        date_of_birth__isnull=False
    ).count()
    
    # Training Statistics
    trainings_this_year = TrainingRecord.objects.filter(
        start_date__year=today.year,
        is_deleted=False
    ).count()
    
    # Performance Reviews
    reviews_this_year = PerformanceReview.objects.filter(
        review_date__year=today.year,
        is_deleted=False
    ).count()
    
    # Payroll Statistics
    try:
        latest_payroll = PayrollPeriod.objects.filter(
            is_deleted=False
        ).order_by('-end_date').first()
        
        if latest_payroll:
            total_payroll = Payroll.objects.filter(
                period=latest_payroll,
                is_deleted=False
            ).aggregate(
                total=Sum('net_pay')
            )['total'] or Decimal('0.00')
        else:
            total_payroll = Decimal('0.00')
    except:
        total_payroll = Decimal('0.00')
    
    # Leave breakdown by status
    all_leaves = LeaveRequest.objects.filter(is_deleted=False)
    rejected_leaves_count = all_leaves.filter(status='rejected').count()
    cancelled_leaves_count = all_leaves.filter(status='cancelled').count()
    
    # Gender distribution
    male_count = Staff.objects.filter(
        is_deleted=False, is_active=True, gender='male'
    ).count()
    female_count = Staff.objects.filter(
        is_deleted=False, is_active=True, gender='female'
    ).count()
    
    # Employment status
    permanent_count = Staff.objects.filter(
        is_deleted=False, is_active=True, employment_status='permanent'
    ).count()
    contract_count = Staff.objects.filter(
        is_deleted=False, is_active=True, employment_status='contract'
    ).count()
    probation_count = Staff.objects.filter(
        is_deleted=False, is_active=True, employment_status='probation'
    ).count()
    
    # Recent Activities
    recent_leaves = LeaveRequest.objects.filter(
        is_deleted=False
    ).select_related('staff__user', 'staff__department').order_by('-created')[:10]
    
    recent_reviews = PerformanceReview.objects.filter(
        is_deleted=False
    ).select_related('staff__user', 'reviewed_by__user').order_by('-review_date')[:10]
    
    context = {
        'total_staff': total_staff,
        'staff_by_department': staff_by_department,
        'staff_by_profession': staff_by_profession,
        'pending_leaves': pending_leaves,
        'approved_leaves_this_month': approved_leaves_this_month,
        'rejected_leaves_count': rejected_leaves_count,
        'cancelled_leaves_count': cancelled_leaves_count,
        'staff_on_leave_today': staff_on_leave_today,
        'contracts_expiring_soon': contracts_expiring_soon,
        'staff_with_birthdays': staff_with_birthdays,
        'trainings_this_year': trainings_this_year,
        'reviews_this_year': reviews_this_year,
        'total_payroll': total_payroll,
        'male_count': male_count,
        'female_count': female_count,
        'permanent_count': permanent_count,
        'contract_count': contract_count,
        'probation_count': probation_count,
        'recent_leaves': recent_leaves,
        'recent_reviews': recent_reviews,
        'excel_available': EXCEL_AVAILABLE,
    }
    
    return render(request, 'hospital/hr_reports_dashboard.html', context)


@login_required
def staff_list_report(request):
    """Staff List Report"""
    department_filter = request.GET.get('department', '')
    profession_filter = request.GET.get('profession', '')
    status_filter = request.GET.get('status', 'active')
    export_format = request.GET.get('export', '')
    
    staff = Staff.objects.filter(is_deleted=False).select_related('user', 'department')
    
    if status_filter == 'active':
        staff = staff.filter(is_active=True)
    elif status_filter == 'inactive':
        staff = staff.filter(is_active=False)
    
    if department_filter:
        staff = staff.filter(department_id=department_filter)
    
    if profession_filter:
        staff = staff.filter(profession=profession_filter)
    
    staff = staff.order_by('department__name', 'user__last_name')
    
    if export_format == 'csv':
        return export_staff_csv(staff)
    elif export_format == 'excel' and EXCEL_AVAILABLE:
        return export_staff_excel(staff)
    
    departments = Department.objects.filter(is_deleted=False).order_by('name')
    professions = Staff.PROFESSION_CHOICES
    
    context = {
        'staff_list': staff,
        'departments': departments,
        'professions': professions,
        'department_filter': department_filter,
        'profession_filter': profession_filter,
        'status_filter': status_filter,
        'excel_available': EXCEL_AVAILABLE,
    }
    
    return render(request, 'hospital/reports/staff_list_report.html', context)


@login_required
def leave_report(request):
    """Leave Report"""
    status_filter = request.GET.get('status', '')
    leave_type_filter = request.GET.get('leave_type', '')
    department_filter = request.GET.get('department', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    export_format = request.GET.get('export', '')
    
    leaves = LeaveRequest.objects.filter(
        is_deleted=False
    ).select_related('staff__user', 'staff__department', 'approved_by__user')
    
    if status_filter:
        leaves = leaves.filter(status=status_filter)
    
    if leave_type_filter:
        leaves = leaves.filter(leave_type=leave_type_filter)
    
    if department_filter:
        leaves = leaves.filter(staff__department_id=department_filter)
    
    if date_from:
        leaves = leaves.filter(start_date__gte=date_from)
    
    if date_to:
        leaves = leaves.filter(end_date__lte=date_to)
    
    leaves = leaves.order_by('-start_date')
    
    # Statistics
    total_days = leaves.aggregate(total=Sum('days_requested'))['total'] or 0
    approved_count = leaves.filter(status='approved').count()
    pending_count = leaves.filter(status='pending').count()
    rejected_count = leaves.filter(status='rejected').count()
    
    if export_format == 'csv':
        return export_leave_csv(leaves)
    elif export_format == 'excel' and EXCEL_AVAILABLE:
        return export_leave_excel(leaves)
    
    departments = Department.objects.filter(is_deleted=False).order_by('name')
    
    context = {
        'leaves': leaves,
        'departments': departments,
        'status_filter': status_filter,
        'leave_type_filter': leave_type_filter,
        'department_filter': department_filter,
        'date_from': date_from,
        'date_to': date_to,
        'total_days': total_days,
        'approved_count': approved_count,
        'pending_count': pending_count,
        'rejected_count': rejected_count,
        'leave_types': LeaveRequest.LEAVE_TYPES,
        'excel_available': EXCEL_AVAILABLE,
    }
    
    return render(request, 'hospital/reports/leave_report.html', context)


@login_required
def attendance_report(request):
    """Attendance Report"""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    department_filter = request.GET.get('department', '')
    export_format = request.GET.get('export', '')
    
    attendance = Attendance.objects.filter(
        is_deleted=False
    ).select_related('staff__user', 'staff__department')
    
    if date_from:
        attendance = attendance.filter(date__gte=date_from)
    
    if date_to:
        attendance = attendance.filter(date__lte=date_to)
    
    if department_filter:
        attendance = attendance.filter(staff__department_id=department_filter)
    
    attendance = attendance.order_by('-date', 'staff__user__last_name')
    
    # Statistics
    present_count = attendance.filter(status='present').count()
    absent_count = attendance.filter(status='absent').count()
    late_count = attendance.filter(status='late').count()
    
    if export_format == 'csv':
        return export_attendance_csv(attendance)
    
    departments = Department.objects.filter(is_deleted=False).order_by('name')
    
    context = {
        'attendance': attendance,
        'departments': departments,
        'date_from': date_from,
        'date_to': date_to,
        'department_filter': department_filter,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'excel_available': EXCEL_AVAILABLE,
    }
    
    return render(request, 'hospital/reports/attendance_report.html', context)


@login_required
@user_passes_test(is_hr_or_admin)
def login_attendance_dashboard(request):
    """Real-time attendance derived from staff login sessions"""
    date_param = request.GET.get('date', '')
    department_filter = request.GET.get('department', '')
    suspicious_only = request.GET.get('suspicious', '') == 'yes'
    
    try:
        selected_date = date.fromisoformat(date_param) if date_param else date.today()
    except ValueError:
        selected_date = date.today()
    
    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime.combine(selected_date, time.min), tz)
    end_dt = timezone.make_aware(datetime.combine(selected_date, time.max), tz)
    
    login_qs = LoginHistory.objects.filter(
        login_time__gte=start_dt,
        login_time__lte=end_dt,
        status='success',
        is_deleted=False
    ).select_related('user', 'staff__department')
    
    if department_filter:
        login_qs = login_qs.filter(staff__department_id=department_filter)
    
    if suspicious_only:
        login_qs = login_qs.filter(is_suspicious=True)
    
    attendance_map = {}
    for record in login_qs:
        staff = record.staff
        if not staff:
            continue
        
        key = staff.pk
        session_end = record.logout_time or record.login_time
        entry = attendance_map.get(key)
        
        if not entry:
            attendance_map[key] = {
                'staff': staff,
                'department': staff.department,
                'first_login': record.login_time,
                'last_activity': session_end,
                'total_sessions': 1,
                'open_sessions': 0 if record.logout_time else 1,
                'locations': {record.location_display},
                'device_types': {record.device_type or 'Unknown'},
                'latest_ip': record.ip_address,
                'suspicious': record.is_suspicious,
            }
        else:
            entry['first_login'] = min(entry['first_login'], record.login_time)
            entry['last_activity'] = max(entry['last_activity'], session_end)
            entry['total_sessions'] += 1
            entry['locations'].add(record.location_display)
            entry['device_types'].add(record.device_type or 'Unknown')
            entry['latest_ip'] = record.ip_address or entry['latest_ip']
            entry['suspicious'] = entry['suspicious'] or record.is_suspicious
            if not record.logout_time:
                entry['open_sessions'] += 1
    
    attendance_rows = sorted(
        attendance_map.values(),
        key=lambda item: (
            item['department'].name if item['department'] else '',
            item['staff'].user.last_name,
            item['staff'].user.first_name,
        )
    )
    
    total_logged_in = len(attendance_rows)
    currently_online = sum(1 for row in attendance_rows if row['open_sessions'] > 0)
    suspicious_staff = sum(1 for row in attendance_rows if row['suspicious'])
    
    avg_first_login = None
    if attendance_rows:
        total_minutes = sum(row['first_login'].hour * 60 + row['first_login'].minute for row in attendance_rows)
        avg_minutes = total_minutes // len(attendance_rows)
        avg_hour = avg_minutes // 60
        avg_minute = avg_minutes % 60
        avg_first_login = f"{avg_hour:02d}:{avg_minute:02d}"
    
    recent_alerts = SecurityAlert.objects.filter(
        alert_time__date=selected_date,
        is_deleted=False
    ).select_related('user').order_by('-alert_time')[:8]
    
    recent_logins = login_qs.order_by('-login_time')[:10]
    departments = Department.objects.filter(is_deleted=False).order_by('name')
    
    for row in attendance_rows:
        row['locations_display'] = ', '.join(sorted(filter(None, row['locations'])))
        row['devices_display'] = ', '.join(sorted(filter(None, row['device_types'])))
        row['first_login_time'] = timezone.localtime(row['first_login'])
        row['last_activity_time'] = timezone.localtime(row['last_activity'])
    
    context = {
        'title': 'Login Attendance Monitor',
        'attendance_rows': attendance_rows,
        'selected_date': selected_date,
        'date_param': selected_date.isoformat(),
        'departments': departments,
        'department_filter': department_filter,
        'suspicious_only': suspicious_only,
        'stats': {
            'total_logged_in': total_logged_in,
            'currently_online': currently_online,
            'suspicious_staff': suspicious_staff,
            'avg_first_login': avg_first_login,
        },
        'recent_alerts': recent_alerts,
        'recent_logins': recent_logins,
        'has_data': bool(attendance_rows),
    }
    
    return render(request, 'hospital/hr/login_attendance_dashboard.html', context)


@login_required
@user_passes_test(is_hr_or_admin)
def live_session_monitor(request):
    """
    Real-time unit session monitor derived from login history
    """
    now = timezone.now()
    window_minutes = int(request.GET.get('window', 15) or 15)
    window_minutes = max(5, min(window_minutes, 180))
    window_delta = timedelta(minutes=window_minutes)
    idle_delta = timedelta(minutes=window_minutes * 2)
    
    lookback_hours = int(request.GET.get('lookback', 8) or 8)
    lookback_hours = max(2, min(lookback_hours, 24))
    lookback_start = now - timedelta(hours=lookback_hours)
    
    departments = Department.objects.filter(is_deleted=False).order_by('name')
    priority_units = {'pharmacy', 'laboratory', 'cashier', 'inventory', 'front desk', 'reception', 'radiology'}
    
    def init_unit(dept=None, label=None):
        name = label or (dept.name if dept else 'General / Cross-Department')
        return {
            'department': dept,
            'name': name,
            'is_priority': name.lower() in priority_units,
            'online_users': set(),
            'recent_users': [],
            'locations': set(),
            'last_seen': None,
            'total_sessions': 0,
        }
    
    unit_map = {}
    for dept in departments:
        unit_map[dept.pk] = init_unit(dept)
    unit_map['unassigned'] = init_unit(None)
    
    recent_logins = LoginHistory.objects.filter(
        login_time__gte=lookback_start,
        status='success',
        is_deleted=False
    ).select_related('staff__user', 'staff__department').order_by('-login_time')
    
    live_sessions = []
    
    for login in recent_logins:
        staff = login.staff
        dept_id = staff.department_id if staff and staff.department else 'unassigned'
        unit = unit_map.get(dept_id)
        if unit is None:
            unit = init_unit(label=f"Dept #{dept_id}")
            unit_map[dept_id] = unit
        
        staff_user = staff.user if staff and staff.user else login.user
        staff_name = staff_user.get_full_name() or staff_user.username
        
        last_activity = login.logout_time or login.login_time
        is_online = (login.logout_time is None) or (login.logout_time >= now - window_delta)
        
        unit['total_sessions'] += 1
        if last_activity and (unit['last_seen'] is None or last_activity > unit['last_seen']):
            unit['last_seen'] = last_activity
        if login.location_display:
            unit['locations'].add(login.location_display)
        
        if staff_name:
            unit['recent_users'].append({
                'name': staff_name,
                'time': login.login_time,
                'is_online': is_online,
            })
            if is_online:
                unit['online_users'].add(staff_name)
                live_sessions.append({
                    'staff': staff_name,
                    'department': staff.department.name if staff and staff.department else 'General / Cross-Department',
                    'login_time': login.login_time,
                    'device': login.device_type or 'Unknown',
                    'location': login.location_display or 'Unknown',
                    'ip_address': login.ip_address or '—',
                })
    
    unit_cards = []
    for unit in unit_map.values():
        online_count = len(unit['online_users'])
        if online_count > 0:
            status = 'online'
        elif unit['last_seen'] and unit['last_seen'] >= now - idle_delta:
            status = 'idle'
        else:
            status = 'offline'
        
        unit_cards.append({
            'name': unit['name'],
            'department': unit['department'],
            'is_priority': unit['is_priority'],
            'online_count': online_count,
            'recent_names': [u['name'] for u in unit['recent_users'][:3]],
            'total_sessions': unit['total_sessions'],
            'last_seen': unit['last_seen'],
            'status': status,
            'locations_display': ', '.join(sorted(unit['locations'])) if unit['locations'] else '—',
        })
    
    unit_cards.sort(key=lambda card: (0 if card['is_priority'] else 1, card['name']))
    
    stats = {
        'units_online': sum(1 for card in unit_cards if card['status'] == 'online'),
        'units_idle': sum(1 for card in unit_cards if card['status'] == 'idle'),
        'units_offline': sum(1 for card in unit_cards if card['status'] == 'offline'),
        'staff_online': sum(card['online_count'] for card in unit_cards),
    }
    
    live_sessions = sorted(live_sessions, key=lambda item: item['login_time'], reverse=True)[:25]
    
    context = {
        'title': 'Live System Sessions',
        'unit_cards': unit_cards,
        'stats': stats,
        'window_minutes': window_minutes,
        'lookback_hours': lookback_hours,
        'live_sessions': live_sessions,
        'now': now,
    }
    return render(request, 'hospital/hr/live_session_monitor.html', context)


@login_required
def payroll_report(request):
    """Payroll Report"""
    period_id = request.GET.get('period', '')
    department_filter = request.GET.get('department', '')
    export_format = request.GET.get('export', '')
    
    periods = PayrollPeriod.objects.filter(is_deleted=False).order_by('-end_date')
    
    if period_id:
        selected_period = PayrollPeriod.objects.filter(pk=period_id, is_deleted=False).first()
    else:
        selected_period = periods.first()
    
    payrolls = []
    total_gross = Decimal('0.00')
    total_deductions = Decimal('0.00')
    total_net = Decimal('0.00')
    
    if selected_period:
        payrolls = Payroll.objects.filter(
            period=selected_period,
            is_deleted=False
        ).select_related('staff__user', 'staff__department')
        
        if department_filter:
            payrolls = payrolls.filter(staff__department_id=department_filter)
        
        payrolls = payrolls.order_by('staff__user__last_name')
        
        # Calculate totals
        totals = payrolls.aggregate(
            total_gross=Sum('gross_pay'),
            total_deductions=Sum('total_deductions'),
            total_net=Sum('net_pay')
        )
        
        total_gross = totals['total_gross'] or Decimal('0.00')
        total_deductions = totals['total_deductions'] or Decimal('0.00')
        total_net = totals['total_net'] or Decimal('0.00')
    
    if export_format == 'csv':
        return export_payroll_csv(payrolls, selected_period)
    
    departments = Department.objects.filter(is_deleted=False).order_by('name')
    
    context = {
        'periods': periods,
        'selected_period': selected_period,
        'payrolls': payrolls,
        'departments': departments,
        'period_id': period_id,
        'department_filter': department_filter,
        'total_gross': total_gross,
        'total_deductions': total_deductions,
        'total_net': total_net,
        'excel_available': EXCEL_AVAILABLE,
    }
    
    return render(request, 'hospital/reports/payroll_report.html', context)


@login_required
def training_report(request):
    """Training Report"""
    year_filter = request.GET.get('year', str(date.today().year))
    department_filter = request.GET.get('department', '')
    export_format = request.GET.get('export', '')
    
    trainings = TrainingRecord.objects.filter(
        is_deleted=False
    ).select_related('staff__user', 'staff__department')
    
    if year_filter:
        trainings = trainings.filter(start_date__year=year_filter)
    
    if department_filter:
        trainings = trainings.filter(staff__department_id=department_filter)
    
    trainings = trainings.order_by('-start_date')
    
    # Statistics
    total_trainings = trainings.count()
    completed = trainings.filter(status='completed').count()
    
    if export_format == 'csv':
        return export_training_csv(trainings)
    
    departments = Department.objects.filter(is_deleted=False).order_by('name')
    years = range(date.today().year, date.today().year - 5, -1)
    
    context = {
        'trainings': trainings,
        'departments': departments,
        'years': years,
        'year_filter': year_filter,
        'department_filter': department_filter,
        'total_trainings': total_trainings,
        'completed': completed,
        'excel_available': EXCEL_AVAILABLE,
    }
    
    return render(request, 'hospital/reports/training_report.html', context)


@login_required
def performance_report(request):
    """Performance Review Report"""
    year_filter = request.GET.get('year', str(date.today().year))
    department_filter = request.GET.get('department', '')
    export_format = request.GET.get('export', '')
    
    reviews = PerformanceReview.objects.filter(
        is_deleted=False
    ).select_related('staff__user', 'staff__department', 'reviewed_by__user')
    
    if year_filter:
        reviews = reviews.filter(review_date__year=year_filter)
    
    if department_filter:
        reviews = reviews.filter(staff__department_id=department_filter)
    
    reviews = reviews.order_by('-review_date')
    
    # Statistics
    total_reviews = reviews.count()
    avg_score = reviews.aggregate(avg=Avg('overall_rating'))['avg'] or 0
    
    if export_format == 'csv':
        return export_performance_csv(reviews)
    
    departments = Department.objects.filter(is_deleted=False).order_by('name')
    years = range(date.today().year, date.today().year - 5, -1)
    
    context = {
        'reviews': reviews,
        'departments': departments,
        'years': years,
        'year_filter': year_filter,
        'department_filter': department_filter,
        'total_reviews': total_reviews,
        'avg_score': round(avg_score, 2),
        'excel_available': EXCEL_AVAILABLE,
    }
    
    return render(request, 'hospital/reports/performance_report.html', context)


# Export Functions
def export_staff_csv(staff_queryset):
    """Export staff list to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="staff_list_{date.today()}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Employee ID', 'Name', 'Department', 'Profession', 'Email', 'Phone', 'Date Joined', 'Status'])
    
    for staff in staff_queryset:
        writer.writerow([
            staff.employee_id or '-',
            staff.user.get_full_name(),
            staff.department.name if staff.department else '-',
            staff.get_profession_display(),
            staff.user.email,
            getattr(staff, 'phone_number', '-'),
            staff.date_of_joining.strftime('%Y-%m-%d') if staff.date_of_joining else '-',
            'Active' if staff.is_active else 'Inactive'
        ])
    
    return response


def export_staff_excel(staff_queryset):
    """Export staff list to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff List"
    
    # Headers
    headers = ['Employee ID', 'Name', 'Department', 'Profession', 'Email', 'Phone', 'Date Joined', 'Status']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for staff in staff_queryset:
        ws.append([
            staff.employee_id or '-',
            staff.user.get_full_name(),
            staff.department.name if staff.department else '-',
            staff.get_profession_display(),
            staff.user.email,
            getattr(staff, 'phone_number', '-'),
            staff.date_of_joining.strftime('%Y-%m-%d') if staff.date_of_joining else '-',
            'Active' if staff.is_active else 'Inactive'
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="staff_list_{date.today()}.xlsx"'
    
    return response


def export_leave_csv(leave_queryset):
    """Export leave report to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="leave_report_{date.today()}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Request #', 'Staff Name', 'Department', 'Leave Type', 'Start Date', 'End Date', 'Days', 'Status', 'Approved By', 'Approved At'])
    
    for leave in leave_queryset:
        writer.writerow([
            leave.request_number or '-',
            leave.staff.user.get_full_name(),
            leave.staff.department.name if leave.staff.department else '-',
            leave.get_leave_type_display(),
            leave.start_date.strftime('%Y-%m-%d'),
            leave.end_date.strftime('%Y-%m-%d'),
            leave.days_requested,
            leave.get_status_display(),
            leave.approved_by.user.get_full_name() if leave.approved_by else '-',
            leave.approved_at.strftime('%Y-%m-%d %H:%M') if leave.approved_at else '-'
        ])
    
    return response


def export_leave_excel(leave_queryset):
    """Export leave report to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Report"
    
    # Headers
    headers = ['Request #', 'Staff Name', 'Department', 'Leave Type', 'Start Date', 'End Date', 'Days', 'Status', 'Approved By', 'Approved At']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for leave in leave_queryset:
        ws.append([
            leave.request_number or '-',
            leave.staff.user.get_full_name(),
            leave.staff.department.name if leave.staff.department else '-',
            leave.get_leave_type_display(),
            leave.start_date.strftime('%Y-%m-%d'),
            leave.end_date.strftime('%Y-%m-%d'),
            leave.days_requested,
            leave.get_status_display(),
            leave.approved_by.user.get_full_name() if leave.approved_by else '-',
            leave.approved_at.strftime('%Y-%m-%d %H:%M') if leave.approved_at else '-'
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="leave_report_{date.today()}.xlsx"'
    
    return response


def export_attendance_csv(attendance_queryset):
    """Export attendance report to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{date.today()}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date', 'Staff Name', 'Department', 'Check In', 'Check Out', 'Status', 'Notes'])
    
    for att in attendance_queryset:
        writer.writerow([
            att.date.strftime('%Y-%m-%d'),
            att.staff.user.get_full_name(),
            att.staff.department.name if att.staff.department else '-',
            att.check_in_time.strftime('%H:%M') if att.check_in_time else '-',
            att.check_out_time.strftime('%H:%M') if att.check_out_time else '-',
            att.get_status_display(),
            att.notes or '-'
        ])
    
    return response


def export_payroll_csv(payroll_queryset, period):
    """Export payroll report to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="payroll_report_{date.today()}.csv"'
    
    writer = csv.writer(response)
    if period:
        writer.writerow([f'Payroll Report - {period.period_name}'])
        writer.writerow([])
    
    writer.writerow(['Employee ID', 'Staff Name', 'Department', 'Gross Pay', 'Total Deductions', 'Net Pay', 'Payment Status'])
    
    for payroll in payroll_queryset:
        writer.writerow([
            payroll.staff.employee_id or '-',
            payroll.staff.user.get_full_name(),
            payroll.staff.department.name if payroll.staff.department else '-',
            f'{payroll.gross_pay:.2f}',
            f'{payroll.total_deductions:.2f}',
            f'{payroll.net_pay:.2f}',
            payroll.get_payment_status_display()
        ])
    
    return response


def export_training_csv(training_queryset):
    """Export training report to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="training_report_{date.today()}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Staff Name', 'Department', 'Training Title', 'Training Type', 'Start Date', 'End Date', 'Duration (Hours)', 'Provider', 'Status'])
    
    for training in training_queryset:
        writer.writerow([
            training.staff.user.get_full_name(),
            training.staff.department.name if training.staff.department else '-',
            training.training_title,
            training.get_training_type_display(),
            training.start_date.strftime('%Y-%m-%d') if training.start_date else '-',
            training.end_date.strftime('%Y-%m-%d') if training.end_date else '-',
            training.duration_hours or '-',
            training.provider or '-',
            training.get_status_display()
        ])
    
    return response


def export_performance_csv(review_queryset):
    """Export performance review report to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="performance_report_{date.today()}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Staff Name', 'Department', 'Review Date', 'Period', 'Overall Rating', 'Reviewer', 'Recommendation'])
    
    for review in review_queryset:
        writer.writerow([
            review.staff.user.get_full_name(),
            review.staff.department.name if review.staff.department else '-',
            review.review_date.strftime('%Y-%m-%d'),
            review.get_review_period_display(),
            review.overall_rating,
            review.reviewed_by.user.get_full_name() if review.reviewed_by else '-',
            review.recommendation or '-'
        ])
    
    return response

