"""
Export lab results to Excel; lab financial report exports (date range).
"""
import json
from io import BytesIO
from datetime import timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Q, Count, Sum, Value, Max
from django.db.models.functions import Coalesce
from django.db.models.fields import DecimalField
from django.utils import timezone
from .models import LabResult

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@login_required
def export_lab_results_excel(request):
    """Export lab results to Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse(
            "Excel export requires openpyxl. Please install it: pip install openpyxl",
            content_type='text/plain',
            status=500
        )
    
    # Get filters from request
    status_filter = request.GET.get('status', '')
    query = request.GET.get('q', '')
    
    # Get lab results with same filters as list view
    results = LabResult.objects.filter(is_deleted=False).select_related(
        'test', 'order__encounter__patient', 'verified_by', 'order__encounter'
    )
    
    if status_filter:
        results = results.filter(status=status_filter)
    
    if query:
        results = results.filter(
            Q(test__name__icontains=query) |
            Q(order__encounter__patient__first_name__icontains=query) |
            Q(order__encounter__patient__last_name__icontains=query) |
            Q(order__encounter__patient__mrn__icontains=query)
        )
    
    # Get all results (not limited like list view)
    results_list = results.order_by('-created')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lab Results'
    
    # Define styles
    header_fill = PatternFill(start_color='7B68EE', end_color='7B68EE', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = 'LABORATORY RESULTS EXPORT'
    title_cell.font = Font(bold=True, size=14, color='7B68EE')
    title_cell.alignment = Alignment(horizontal='center')
    
    # Metadata
    ws['A2'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'] = f'Total Records: {results_list.count()}'
    ws['A4'] = f'Status Filter: {status_filter if status_filter else "All"}'
    ws['A5'] = f'Search Query: {query if query else "None"}'
    
    # Headers
    headers = [
        'Date', 'Time', 'Patient Name', 'MRN', 'Test Name', 'Test Code',
        'Status', 'Result Value', 'Units', 'Abnormal', 'Verified By', 'Notes'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row_num = 8
    for result in results_list:
        patient = result.order.encounter.patient if result.order and result.order.encounter else None
        patient_name = patient.full_name if patient else 'N/A'
        mrn = patient.mrn if patient else 'N/A'
        
        # Format result value
        result_value = result.value or ''
        if result.details:
            # For panel tests, combine details
            detail_str = ', '.join([f"{k}: {v}" for k, v in result.details.items()])
            if detail_str:
                result_value = detail_str
        if result.qualitative_result:
            result_value = result.qualitative_result
        
        # Format date and time
        result_date = result.created.date() if result.created else ''
        result_time = result.created.time() if result.created else ''
        
        data = [
            result_date,
            result_time,
            patient_name,
            mrn,
            result.test.name if result.test else 'N/A',
            result.test.code if result.test else 'N/A',
            result.get_status_display(),
            result_value,
            result.units or '',
            'Yes' if result.is_abnormal else 'No',
            result.verified_by.user.get_full_name() if result.verified_by and result.verified_by.user else 'N/A',
            result.notes or ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Highlight abnormal results
            if result.is_abnormal and col_num == 10:  # Abnormal column
                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        
        row_num += 1
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'lab_results_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def _parse_lab_report_export_dates(request):
    from datetime import datetime as dt_datetime

    today = timezone.now().date()
    date_from = today - timedelta(days=30)
    date_to = today
    raw_from = (request.GET.get('date_from') or '').strip()
    raw_to = (request.GET.get('date_to') or '').strip()
    if raw_from:
        try:
            date_from = dt_datetime.strptime(raw_from[:10], '%Y-%m-%d').date()
        except ValueError:
            pass
    if raw_to:
        try:
            date_to = dt_datetime.strptime(raw_to[:10], '%Y-%m-%d').date()
        except ValueError:
            pass
    if date_to < date_from:
        date_from, date_to = date_to, date_from
    max_days = 366
    if (date_to - date_from).days > max_days:
        date_from = date_to - timedelta(days=max_days)
    return date_from, date_to


def _parse_lab_report_export_basis(request):
    v = (request.GET.get('basis') or 'verified').strip().lower()
    if v in ('created', 'activity', 'all', 'total'):
        return 'created'
    return 'verified'


def _lab_report_export_queryset(date_from, date_to, basis):
    if basis == 'created':
        return (
            LabResult.objects.filter(
                created__date__gte=date_from,
                created__date__lte=date_to,
                is_deleted=False,
            )
            .select_related('test', 'order__encounter__patient', 'verified_by__user')
            .order_by('-created')
        )
    return (
        LabResult.objects.filter(
            status='completed',
            verified_at__isnull=False,
            verified_at__date__gte=date_from,
            verified_at__date__lte=date_to,
            is_deleted=False,
        )
        .select_related('test', 'order__encounter__patient', 'verified_by__user')
        .order_by('-verified_at')
    )


def _lab_report_dec_zero():
    return Value(Decimal('0'), output_field=DecimalField(max_digits=14, decimal_places=2))


def build_laboratory_financial_report_context(request):
    """Shared context for lab financial HTML report (date range + basis filters)."""
    date_from, date_to = _parse_lab_report_export_dates(request)
    basis = _parse_lab_report_export_basis(request)
    dec_zero = _lab_report_dec_zero()
    period_days = max(1, (date_to - date_from).days + 1)

    if basis == 'created':
        base_qs = LabResult.objects.filter(
            created__date__gte=date_from,
            created__date__lte=date_to,
            is_deleted=False,
        )
        total_activity = base_qs.count()
        completed_qs = base_qs.filter(status='completed')
        completed_totals = completed_qs.aggregate(
            n=Count('id'),
            revenue=Coalesce(Sum('test__price'), dec_zero),
            abnormal_n=Count('id', filter=Q(is_abnormal=True)),
        )
        n_completed = completed_totals['n'] or 0
        n_pending = base_qs.filter(status='pending').count()
        n_in_progress = base_qs.filter(status='in_progress').count()
        n_cancelled = base_qs.filter(status='cancelled').count()
        total_n = total_activity
        total_revenue = completed_totals['revenue'] or Decimal('0')
        abnormal_count = completed_totals['abnormal_n'] or 0
        date_field = 'created__date'
        table_qs = base_qs
    else:
        base_qs = LabResult.objects.filter(
            status='completed',
            verified_at__isnull=False,
            verified_at__date__gte=date_from,
            verified_at__date__lte=date_to,
            is_deleted=False,
        )
        totals = base_qs.aggregate(
            n=Count('id'),
            revenue=Coalesce(Sum('test__price'), dec_zero),
            abnormal_n=Count('id', filter=Q(is_abnormal=True)),
        )
        total_activity = 0
        n_completed = totals['n'] or 0
        n_pending = n_in_progress = n_cancelled = 0
        total_n = n_completed
        total_revenue = totals['revenue'] or Decimal('0')
        abnormal_count = totals['abnormal_n'] or 0
        date_field = 'verified_at__date'
        table_qs = base_qs

    normal_count = max(0, n_completed - abnormal_count)
    pct_abnormal = round(100.0 * abnormal_count / n_completed, 1) if n_completed else 0.0
    avg_revenue_per_day = (total_revenue / Decimal(period_days)).quantize(Decimal('0.01'))
    avg_activity_per_day = total_activity / period_days if basis == 'created' else 0
    avg_completed_per_day = total_n / period_days if basis == 'verified' else 0

    by_test = list(
        base_qs.values('test_id')
        .annotate(
            test__name=Max('test__name'),
            test__code=Max('test__code'),
            n=Count('id'),
            rev=Coalesce(
                Sum('test__price', filter=Q(status='completed')) if basis == 'created' else Sum('test__price'),
                dec_zero,
            ),
        )
        .order_by('-n')[:50]
    )

    denominator_frequency = total_n or 1
    frequent_tests_ranked = []
    for rank, row in enumerate(by_test, start=1):
        n = row['n']
        pct = round(100.0 * n / denominator_frequency, 1) if denominator_frequency else 0.0
        frequent_tests_ranked.append({
            'rank': rank,
            'name': row.get('test__name') or '—',
            'code': row.get('test__code') or '',
            'n': n,
            'pct': pct,
            'rev': row['rev'] or Decimal('0'),
        })

    daily_counts = {
        r[date_field]: r['n']
        for r in base_qs.values(date_field).annotate(n=Count('id'))
    }
    completed_daily_revenue = {
        r[date_field]: r['rev']
        for r in (
            completed_qs if basis == 'created' else base_qs
        ).values(date_field).annotate(rev=Coalesce(Sum('test__price'), dec_zero))
    }

    daily_breakdown_rows = []
    daily_chart = []
    day = date_from
    while day <= date_to:
        count = daily_counts.get(day, 0)
        revenue = completed_daily_revenue.get(day, Decimal('0'))
        daily_breakdown_rows.append({
            'date_display': day.strftime('%b %d, %Y'),
            'count': count,
            'revenue': revenue,
        })
        daily_chart.append({
            'date': day.strftime('%b %d'),
            'count': count,
            'revenue': float(revenue),
        })
        day += timedelta(days=1)

    top10 = by_test[:10]
    bar_labels = [(r.get('test__name') or '—')[:28] for r in top10]
    bar_counts = [r['n'] for r in top10]

    pie_top = by_test[:5]
    pie_other = sum(r['n'] for r in by_test[5:])
    pie_labels = [(r.get('test__name') or '—')[:20] for r in pie_top]
    pie_counts = [r['n'] for r in pie_top]
    if pie_other:
        pie_labels.append('Other')
        pie_counts.append(pie_other)

    table_limit = 250
    table_total = table_qs.count()
    table_rows = list(
        table_qs.select_related(
            'test', 'order__encounter__patient', 'verified_by__user',
        ).order_by('-created' if basis == 'created' else '-verified_at')[:table_limit]
    )

    narrative_lines = []
    if total_n:
        mode = 'created in' if basis == 'created' else 'verified in'
        narrative_lines.append(
            f'{total_n:,} lab record{"s" if total_n != 1 else ""} {mode} this period '
            f'({date_from.strftime("%b %d, %Y")} – {date_to.strftime("%b %d, %Y")}).'
        )
        if basis == 'created':
            narrative_lines.append(
                f'Of those, {n_completed:,} completed, {n_pending:,} pending, '
                f'{n_in_progress:,} in progress, {n_cancelled:,} cancelled.'
            )
        narrative_lines.append(
            f'Catalog list value on completed tests: GHS {total_revenue:,.2f} '
            f'(~GHS {avg_revenue_per_day:,.2f} per day).'
        )
        if abnormal_count:
            narrative_lines.append(
                f'{abnormal_count:,} completed result{"s" if abnormal_count != 1 else ""} '
                f'flagged abnormal ({pct_abnormal}%).'
            )
        if by_test:
            top = by_test[0]
            top_pct = round(100.0 * top['n'] / denominator_frequency, 1)
            narrative_lines.append(
                f'Most frequent test: {(top.get("test__name") or "—")} — '
                f'{top["n"]:,} times ({top_pct}% of rows in this view).'
            )

    lab_billed_invoice = None
    try:
        from .services.department_billed_revenue_service import build_department_report

        billed = build_department_report(
            date_from=date_from,
            date_to=date_to,
            include_writeoff_period=False,
            dept='lab',
            payer_type=None,
        )
        payer_summary = billed['payer_summary']
        if payer_summary['total_lines']:
            lab_billed_invoice = payer_summary
    except Exception:
        pass

    from .views_hod_shift_monitoring import is_hod

    return {
        'report_basis': basis,
        'date_from': date_from,
        'date_to': date_to,
        'has_data': total_n > 0,
        'narrative_lines': narrative_lines,
        'total_activity': total_activity,
        'avg_activity_per_day': avg_activity_per_day,
        'n_completed': n_completed,
        'n_pending': n_pending,
        'n_in_progress': n_in_progress,
        'n_cancelled': n_cancelled,
        'total_completed': n_completed,
        'avg_completed_per_day': avg_completed_per_day,
        'total_revenue': total_revenue,
        'avg_revenue_per_day': avg_revenue_per_day,
        'abnormal_count': abnormal_count,
        'normal_count': normal_count,
        'pct_abnormal': pct_abnormal,
        'lab_billed_invoice': lab_billed_invoice,
        'daily_chart_json': json.dumps(daily_chart),
        'bar_labels_json': json.dumps(bar_labels),
        'bar_counts_json': json.dumps(bar_counts),
        'pie_labels_json': json.dumps(pie_labels),
        'pie_counts_json': json.dumps(pie_counts),
        'daily_breakdown_rows': daily_breakdown_rows,
        'frequent_tests_ranked': frequent_tests_ranked,
        'denominator_frequency': total_n,
        'by_test': by_test,
        'table_rows': table_rows,
        'table_truncated': table_total > table_limit,
        'is_lab_hod': is_hod(request.user),
    }


@login_required
def laboratory_financial_report(request):
    """Laboratory period report: workload, catalog revenue, and test frequency."""
    context = build_laboratory_financial_report_context(request)
    return render(request, 'hospital/laboratory_financial_report.html', context)


@login_required
def export_laboratory_financial_excel(request):
    """Export completed lab results in date range (verified date) with revenue columns."""
    if not EXCEL_AVAILABLE:
        return HttpResponse(
            "Excel export requires openpyxl. Please install it: pip install openpyxl",
            content_type='text/plain',
            status=500,
        )
    date_from, date_to = _parse_lab_report_export_dates(request)
    basis = _parse_lab_report_export_basis(request)
    qs = _lab_report_export_queryset(date_from, date_to, basis)[:5000]

    dec_z = Value(Decimal('0'), output_field=DecimalField(max_digits=14, decimal_places=2))
    if basis == 'created':
        scope = LabResult.objects.filter(
            created__date__gte=date_from,
            created__date__lte=date_to,
            is_deleted=False,
        )
        total_rows = scope.count()
        completed_scope = scope.filter(status='completed')
        totals = completed_scope.aggregate(
            n=Count('id'),
            rev=Coalesce(Sum('test__price'), dec_z),
        )
        total_n = total_rows
        total_rev = totals['rev'] or Decimal('0')
        n_completed = totals['n'] or 0
    else:
        scope = LabResult.objects.filter(
            status='completed',
            verified_at__isnull=False,
            verified_at__date__gte=date_from,
            verified_at__date__lte=date_to,
            is_deleted=False,
        )
        totals = scope.aggregate(
            n=Count('id'),
            rev=Coalesce(Sum('test__price'), dec_z),
        )
        total_n = totals['n'] or 0
        total_rev = totals['rev'] or Decimal('0')
        n_completed = total_n

    if basis == 'created':
        headers = [
            'Created date',
            'Created time',
            'Status',
            'Verified date',
            'Verified time',
            'Patient',
            'MRN',
            'Test',
            'Code',
            'Price (GHS)',
            'Abnormal',
            'Verified by',
            'Notes',
        ]
    else:
        headers = [
            'Verified date',
            'Verified time',
            'Patient',
            'MRN',
            'Test',
            'Code',
            'Price (GHS)',
            'Abnormal',
            'Verified by',
            'Status',
            'Notes',
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lab report'

    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    end_col = get_column_letter(len(headers))
    ws.merge_cells(f'A1:{end_col}1')
    t = ws['A1']
    t.value = 'LABORATORY REPORT & FINANCIAL EXPORT'
    t.font = Font(bold=True, size=14, color='1565C0')
    t.alignment = Alignment(horizontal='center')

    mode_label = 'Total activity (created date)' if basis == 'created' else 'Completed (verified date)'
    ws['A2'] = f'Period: {date_from} to {date_to} — {mode_label}'
    ws['A3'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
    if basis == 'created':
        ws['A4'] = f'Rows in export: {qs.count()} (max 5000). Records in period: {total_n}. Completed subset: {n_completed}.'
    else:
        ws['A4'] = f'Rows in export: {qs.count()} (max 5000). Completed in period: {total_n}.'
    ws['A5'] = f'Catalog value (completed rows in period): GHS {total_rev:.2f}'

    start_row = 7
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row_num = start_row + 1
    for result in qs:
        patient = (
            result.order.encounter.patient
            if result.order and result.order.encounter
            else None
        )
        name = patient.full_name if patient else 'N/A'
        mrn = patient.mrn if patient else 'N/A'
        price = Decimal('0')
        if result.test and getattr(result.test, 'price', None) is not None:
            price = result.test.price
        vb = ''
        if result.verified_by and result.verified_by.user:
            vb = result.verified_by.user.get_full_name() or result.verified_by.user.username
        if basis == 'created':
            data = [
                result.created.date() if result.created else '',
                result.created.time() if result.created else '',
                result.get_status_display(),
                result.verified_at.date() if result.verified_at else '',
                result.verified_at.time() if result.verified_at else '',
                name,
                mrn,
                result.test.name if result.test else 'N/A',
                result.test.code if result.test else '',
                float(price),
                'Yes' if result.is_abnormal else 'No',
                vb or 'N/A',
                (result.notes or '')[:500],
            ]
        else:
            data = [
                result.verified_at.date() if result.verified_at else '',
                result.verified_at.time() if result.verified_at else '',
                name,
                mrn,
                result.test.name if result.test else 'N/A',
                result.test.code if result.test else '',
                float(price),
                'Yes' if result.is_abnormal else 'No',
                vb or 'N/A',
                result.get_status_display(),
                (result.notes or '')[:500],
            ]
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
        row_num += 1

    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = min(18, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fn = f'lab_financial_{date_from}_{date_to}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{fn}"'
    wb.save(response)
    return response


@login_required
def export_laboratory_financial_pdf(request):
    """PDF summary of lab completed volume and revenue for a date range."""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import inch
    except ImportError:
        return HttpResponse(
            'ReportLab is required for PDF export. Install reportlab and try again.',
            status=500,
            content_type='text/plain',
        )

    date_from, date_to = _parse_lab_report_export_dates(request)
    basis = _parse_lab_report_export_basis(request)
    dec_zero = Value(Decimal('0'), output_field=DecimalField(max_digits=14, decimal_places=2))

    if basis == 'created':
        base_qs = LabResult.objects.filter(
            created__date__gte=date_from,
            created__date__lte=date_to,
            is_deleted=False,
        )
        total_activity = base_qs.count()
        completed_qs = base_qs.filter(status='completed')
        totals = completed_qs.aggregate(
            n=Count('id'),
            revenue=Coalesce(Sum('test__price'), dec_zero),
            abnormal_n=Count('id', filter=Q(is_abnormal=True)),
        )
        total_n = total_activity
        total_rev = totals['revenue'] or Decimal('0')
        abnormal_n = totals['abnormal_n'] or 0
        distinct_tests = base_qs.values('test_id').distinct().count() if total_activity else 0
        by_test = list(
            base_qs.values('test_id')
            .annotate(
                test__name=Max('test__name'),
                n=Count('id'),
                rev=Coalesce(Sum('test__price', filter=Q(status='completed')), dec_zero),
            )
            .order_by('-n')[:25]
        )
    else:
        base_qs = LabResult.objects.filter(
            status='completed',
            verified_at__isnull=False,
            verified_at__date__gte=date_from,
            verified_at__date__lte=date_to,
            is_deleted=False,
        )
        totals = base_qs.aggregate(
            n=Count('id'),
            revenue=Coalesce(Sum('test__price'), dec_zero),
            abnormal_n=Count('id', filter=Q(is_abnormal=True)),
        )
        total_n = totals['n'] or 0
        total_rev = totals['revenue'] or Decimal('0')
        abnormal_n = totals['abnormal_n'] or 0
        distinct_tests = base_qs.values('test_id').distinct().count() if total_n else 0
        by_test = list(
            base_qs.values('test_id')
            .annotate(
                test__name=Max('test__name'),
                n=Count('id'),
                rev=Coalesce(Sum('test__price'), dec_zero),
            )
            .order_by('-n')[:25]
        )

    period_days = max(1, (date_to - date_from).days + 1)
    avg_rev_day = (total_rev / Decimal(period_days)).quantize(Decimal('0.01'))

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - inch
    pdf.setFont('Helvetica-Bold', 14)
    pdf.drawString(inch, y, 'Laboratory report & financial summary')
    y -= 0.35 * inch
    pdf.setFont('Helvetica', 10)
    mode = 'Created date (all activity)' if basis == 'created' else 'Verified date (completed only)'
    summary_lines = [
        f'Period: {date_from} to {date_to} ({period_days} calendar days) — {mode}',
        f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}',
    ]
    if basis == 'created':
        summary_lines.append(f'Lab records created in period: {total_n}')
        summary_lines.append(f'Completed subset: {totals["n"] or 0}')
    else:
        summary_lines.append(f'Completed tests: {total_n}')
    summary_lines.extend([
        f'Distinct test types: {distinct_tests}',
        f'Catalog value (completed rows): GHS {total_rev:.2f}',
        f'Average catalog value per day: GHS {avg_rev_day:.2f}',
        f'Abnormal (completed): {abnormal_n}',
    ])
    denom_pdf = total_n or 1
    if by_test and total_n:
        top = by_test[0]
        pct = round(100.0 * top['n'] / denom_pdf, 1)
        summary_lines.append(
            f'Most frequent test: {(top.get("test__name") or "")[:60]} — {top["n"]} times ({pct}% of rows in this view)'
        )
    for line in summary_lines:
        pdf.drawString(inch, y, line)
        y -= 0.22 * inch

    y -= 0.15 * inch
    pdf.setFont('Helvetica-Bold', 11)
    pdf.drawString(inch, y, 'Tests by frequency (this view)')
    y -= 0.25 * inch
    pdf.setFont('Helvetica', 9)
    pdf.drawString(inch, y, 'Test')
    pdf.drawString(3.5 * inch, y, 'Count')
    pdf.drawString(4.15 * inch, y, '%')
    pdf.drawString(4.65 * inch, y, 'GHS')
    y -= 0.18 * inch
    for row in by_test:
        if y < inch:
            pdf.showPage()
            y = height - inch
            pdf.setFont('Helvetica', 9)
        name = (row.get('test__name') or '')[:42]
        n = row['n']
        pct = round(100.0 * n / denom_pdf, 1) if denom_pdf else 0.0
        pdf.drawString(inch, y, name)
        pdf.drawString(3.5 * inch, y, str(n))
        pdf.drawString(4.15 * inch, y, f'{pct}%')
        pdf.drawString(4.65 * inch, y, f"{row['rev'] or 0:.2f}")
        y -= 0.18 * inch

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="lab_financial_{date_from}_{date_to}.pdf"'
    return response










