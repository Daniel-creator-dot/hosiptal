"""
Management reporting for finance: billed revenue by catalog service (InvoiceLine / ServiceCode).
"""
from __future__ import annotations

import csv
import io
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from .decorators import role_required
from .services.invoice_line_revenue_query import (
    aggregate_service_rows as _aggregate_service_rows,
    category_rollup as _category_rollup,
    parse_ymd as _parse_ymd,
    service_revenue_line_queryset as _service_revenue_line_queryset,
)


@login_required
@role_required('accountant', 'senior_account_officer')
def management_reports_hub(request):
    today = timezone.now().date()
    return render(
        request,
        'hospital/accountant/management_reports_hub.html',
        {
            'today': today,
            'start_of_month': today.replace(day=1),
        },
    )


@login_required
@role_required('accountant', 'senior_account_officer')
def management_service_revenue_report(request):
    today = timezone.now().date()
    default_from = today.replace(day=1)
    date_from = _parse_ymd(request.GET.get('date_from'), default_from)
    date_to = _parse_ymd(request.GET.get('date_to'), today)
    include_writeoff = request.GET.get('include_writeoff') in ('1', 'true', 'on', 'yes')

    payer_type = (request.GET.get('payer_type') or 'all').strip().lower()
    if payer_type not in ('all', 'cash', 'nhis', 'private', 'corporate', 'insurance'):
        payer_type = 'all'

    category = (request.GET.get('category') or '').strip()
    search_q = (request.GET.get('q') or '').strip()

    min_billed = None
    raw_min = (request.GET.get('min_billed') or '').strip()
    if raw_min:
        try:
            min_billed = Decimal(raw_min)
        except Exception:
            min_billed = None

    base_kw = dict(
        date_from=date_from,
        date_to=date_to,
        include_writeoff_period=include_writeoff,
        payer_type=payer_type if payer_type != 'all' else None,
        search_q=search_q or None,
    )
    qs_base = _service_revenue_line_queryset(category_icontains=None, **base_kw)
    category_choices = [
        c
        for c in qs_base.values_list('service_code__category', flat=True).distinct().order_by('service_code__category')
        if c
    ]

    qs = qs_base
    if category:
        qs = qs.filter(service_code__category=category)

    rows, total_billed = _aggregate_service_rows(qs, min_billed)
    by_category = _category_rollup(qs_base)

    export = (request.GET.get('export') or '').lower()
    if export == 'csv':
        return _service_revenue_csv_response(rows, total_billed, date_from, date_to)
    if export == 'xlsx':
        return _service_revenue_xlsx_response(rows, total_billed, date_from, date_to)

    return render(
        request,
        'hospital/accountant/management_service_revenue.html',
        {
            'date_from': date_from,
            'date_to': date_to,
            'include_writeoff': include_writeoff,
            'payer_type': payer_type,
            'category': category,
            'q': search_q,
            'min_billed': raw_min,
            'rows': rows,
            'total_billed': total_billed,
            'by_category': by_category,
            'category_choices': category_choices,
        },
    )


def _service_revenue_csv_response(rows, total_billed, date_from, date_to):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['Billed revenue by service (catalog)'])
    w.writerow([f'Period {date_from.isoformat()} to {date_to.isoformat()}'])
    w.writerow([])
    w.writerow(['Service code', 'Description', 'Category', 'Line count', 'Quantity', 'Billed (GHS)', 'Share %'])
    for r in rows:
        w.writerow(
            [
                r['service_code__code'],
                r['service_code__description'],
                r['service_code__category'],
                r['line_count'],
                str(r['qty']),
                f"{r['billed']:.2f}",
                f"{r['pct']:.2f}",
            ]
        )
    w.writerow([])
    w.writerow(['Total', '', '', '', '', f'{total_billed:.2f}', '100.00'])
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = (
        f'attachment; filename="management_service_revenue_{date_from.strftime("%Y%m%d")}_{date_to.strftime("%Y%m%d")}.csv"'
    )
    return resp


def _service_revenue_xlsx_response(rows, total_billed, date_from, date_to):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        return HttpResponse('Excel export requires openpyxl.', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Service revenue'
    ws.append([f'Billed revenue by service — {date_from} to {date_to}'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])
    hdr_row = 3
    ws.append(['Service code', 'Description', 'Category', 'Lines', 'Qty', 'Billed (GHS)', 'Share %'])
    for c in ws[hdr_row]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(
            [
                r['service_code__code'],
                r['service_code__description'],
                r['service_code__category'],
                r['line_count'],
                float(r['qty']),
                float(r['billed']),
                float(r['pct']),
            ]
        )
    total_row = ws.max_row + 1
    ws.append(['Total', '', '', '', '', float(total_billed), 100.0])
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'F{total_row}'].font = Font(bold=True)
    for i in range(hdr_row + 1, ws.max_row + 1):
        ws[f'F{i}'].number_format = '#,##0.00'
        ws[f'G{i}'].number_format = '0.00'
    for col in ws.columns:
        letter = col[0].column_letter
        maxlen = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[letter].width = min(55, max(10, maxlen + 2))
    bio = io.BytesIO()
    wb.save(bio)
    fname = f'management_service_revenue_{date_from.strftime("%Y%m%d")}_{date_to.strftime("%Y%m%d")}.xlsx'
    resp = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp
