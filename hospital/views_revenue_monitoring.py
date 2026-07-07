"""
Revenue Stream Monitoring Views
Monitor where revenue is coming from
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.conf import settings
from django.http import JsonResponse, HttpResponse, FileResponse
from decimal import Decimal
from datetime import timedelta
from collections import defaultdict
import io

from .models import Department
from .models_revenue_streams import RevenueStream, DepartmentRevenue
from .models_accounting_advanced import Revenue
from .decorators import finance_revenue_streams_access
from .views_accounting_advanced import is_accountant


def _parse_ymd(value, default_date):
    if not value:
        return default_date
    try:
        return timezone.datetime.strptime(value, '%Y-%m-%d').date()
    except Exception:
        return default_date


_SERVICE_TYPE_ALIASES = {
    # Consultation
    'consultation': 'consultation',
    'outpatient': 'consultation',
    'gp': 'consultation',
    'opd': 'consultation',
    # Lab
    'lab': 'lab',
    'lab_test': 'lab',
    'lab_result': 'lab',
    'laboratory': 'lab',
    'laboratory_test': 'lab',
    'blood_test': 'lab',
    'pathology': 'lab',
    'histology': 'lab',
    # Pharmacy
    'pharmacy': 'pharmacy',
    'pharmacy_prescription': 'pharmacy',
    'pharmacy_walkin': 'pharmacy',
    'medication': 'pharmacy',
    'drug': 'pharmacy',
    # Imaging
    'imaging': 'imaging',
    'imaging_study': 'imaging',
    'radiology': 'imaging',
    'scan': 'imaging',
    # Other explicit types (kept as their own KPIs)
    'dental': 'dental',
    'dentistry': 'dental',
    'gynecology': 'gynecology',
    'psychiatry': 'psychiatry',
    'psychiatric': 'psychiatry',
    'mental_health': 'psychiatry',
    'physiotherapy': 'physiotherapy',
    'physio': 'physiotherapy',
    'therapy': 'physiotherapy',
    'registration': 'registration',
    'registration_fee': 'registration',
    'surgery': 'surgery',
    'procedure': 'surgery',
    'emergency': 'emergency',
    'ambulance': 'ambulance',
    'admission': 'admission',
    'detainment': 'admission',
    'detention': 'admission',
    'bed': 'admission',
    # Reception / cashier preset keys (not in PaymentReceipt.SERVICE_TYPES choices but may be stored on JSON / legacy)
    'antenatal': 'gynecology',
    'gynae_special': 'gynecology',
    'gynae': 'gynecology',
    # 'invoice' / 'invoice_line' entries are expanded by drilling into their underlying
    # invoice in _alloc_from_combined_receipt; the alias is only used as a fallback if the
    # invoice cannot be resolved.
    'invoice': 'other',
    'invoice_line': 'other',
    'other': 'other',
}

# Service codes that bed-billing emits with category='accommodation'. We classify each
# code explicitly so consumables flow into the Pharmacy stream while the rest of the
# admission lines (bed/doctor/nursing) flow into the Admission stream.
_ADMISSION_ITEM_CODES = {'ADM-ACCOM', 'ADM-DOCTOR-CARE', 'ADM-NURSING-CARE', 'DETENTION'}
_PHARMACY_ITEM_CODES = {'ADM-CONSUMABLES'}

# Cashier / legacy codes that are laboratory tests but do not use LAB-* prefixes.
_LAB_NON_PREFIX_SERVICE_CODES = frozenset(
    {
        'URA001',  # Urinalysis (cashier quick add)
    }
)

# Description fragments for lines billed as "Clinical Services" / Other but clearly lab work.
_LAB_DESCRIPTION_FRAGMENTS = (
    'urinalysis',
    'urine analysis',
    'urine microscopy',
    'full blood count',
    'full blood',
    ' fbc',
    'fbc ',
    'fbc,',
    'fbc.',
    '(fbc',
    ' rbs ',
    ' fbs ',
    'fasting blood sugar',
    'random blood sugar',
    'fasting glucose',
    ' hba1c',
    ' hemoglobin',
    ' haemoglobin',
    ' hgb ',
    'platelet count',
    'blood film',
    'peripheral smear',
    'peripheral blood',
    'malaria rdt',
    'malaria parasite',
    'malaria smear',
    'malaria film',
    'mp slide',
    ' widal',
    ' typhoid',
    'hepatitis b',
    'hepatitis c',
    ' hbsag',
    'anti-hcv',
    ' hiv ',
    'hiv test',
    'hiv screening',
    ' vdrl',
    ' syphilis',
    'blood culture',
    'culture and sensitivity',
    ' c/s ',
    ' c&s ',
    'culture & sensitivity',
    'sensitivity pattern',
    ' urea ',
    ' creatinine',
    'electrolytes',
    ' u&e',
    ' u and e',
    'liver function',
    ' lft ',
    'lipid profile',
    ' cholesterol',
    'triglyceride',
    'prothrombin time',
    ' aptt ',
    ' inr',
    ' d-dimer',
    ' ferritin',
    ' tsh ',
    'thyroid function',
    'troponin',
    'stool analysis',
    'stool culture',
    'occult blood',
    'semen analysis',
    ' pap smear',
    'pap test',
    ' sputum ',
    ' gram stain',
    'afb smear',
    'gene expert',
    ' cbnaat',
    ' crp ',
    ' esr ',
    'psa ',
    'cea ',
    'afp ',
    ' estradiol',
    'oestradiol',
    ' testosterone',
    ' beta hcg',
    ' b-hcg',
    'hcg qualit',
    'procalcitonin',
    ' amylase',
    ' lipase',
    ' ldh ',
    ' ck-mb',
    ' ck mb',
    ' blood group',
    'group & save',
    'cross match',
    'crossmatch',
    'g&s ',
    'renal profile',
    'liver profile',
    'renal function test',
    'liver function test',
    'electrolyte panel',
    'metabolic screen',
    'renal panel',
    'urine dipstick',
    'csf analysis',
    'fluid analysis',
    'ascitic fluid',
    'pleural fluid',
    ' hb electrophoresis',
    'protein electrophoresis',
    'lactate dehydrogenase',
    'sputum for afb',
    'afb culture',
)


def _invoice_line_is_laboratory(line, display_cat_key=None):
    """
    True when the line should count toward the Laboratory revenue stream even if
    ServiceCode.category / display category are generic (e.g. Clinical Services).
    """
    sc = getattr(line, 'service_code', None)
    if not sc:
        return False
    code = (getattr(sc, 'code', None) or '').strip().upper()
    desc = (getattr(line, 'description', '') or '').casefold()
    cat = (getattr(sc, 'category', None) or '').strip().casefold()
    dcat = (display_cat_key or '').strip().casefold()
    blob = f'{desc} {cat} {dcat}'

    # Hard exclusion: pharmacy-like lines must never be classified as laboratory,
    # even if their description contains a generic word (e.g. "vitamin", "hcg").
    if getattr(line, 'prescription_id', None):
        return False
    if code in _PHARMACY_ITEM_CODES:
        return False
    if code.startswith(('WALKIN-', 'DRUG-', 'PHARM', 'RX-', 'MED-', 'PHARMA')):
        return False
    if any(k in cat for k in ('pharm', 'drug', 'medication', 'dispens')):
        return False

    if code.startswith(
        ('LAB-', 'LABTEST-', 'PATH-', 'BIO-', 'HEMA-', 'MIC-', 'MICRO-', 'SERO-', 'IMMUNO-')
    ):
        return True
    if code in _LAB_NON_PREFIX_SERVICE_CODES:
        return True
    if code.startswith('LAB') and len(code) >= 6 and not code.startswith('LABEL'):
        return True
    if len(code) >= 4 and code.startswith('SL') and code[2:3].isdigit():
        return True

    if any(
        x in dcat
        for x in (
            'laborat',
            'pathology',
            'histology',
            'microbiology',
            'biochemistry',
            'hematology',
            'serology',
            'immunology',
        )
    ):
        return True
    if any(
        x in cat
        for x in (
            'laborat',
            'pathology',
            'histology',
            'microbiology',
            'biochemistry',
            'hematology',
            'serology',
            'immunology',
        )
    ):
        return True
    if 'investigation' in cat and 'imag' not in cat and 'radio' not in cat and 'scan' not in cat:
        return True
    if 'investigation' in dcat and 'imag' not in dcat and 'radio' not in dcat and 'scan' not in dcat:
        return True

    for frag in _LAB_DESCRIPTION_FRAGMENTS:
        if frag in desc:
            return True

    return False


def _bucket_for_invoice_line(line):
    """
    Return the canonical revenue bucket for an InvoiceLine, preferring the line's
    ServiceCode.code over substring matching on display category. Returns None if no
    explicit code-level rule applies (caller should fall back to category substrings).
    """
    sc = getattr(line, 'service_code', None)
    code = (getattr(sc, 'code', '') or '').strip().upper()
    cat_raw = ((getattr(sc, 'category', None) or '') if sc else '').strip().casefold()
    desc_early = (getattr(line, 'description', '') or '').casefold()

    # Pharmacy short-circuit FIRST: a line with a prescription, a pharmacy code, or a
    # pharmacy ServiceCode.category must always land in Pharmacy regardless of any
    # description/keyword matches that might otherwise pull it into Lab/Gynae.
    if getattr(line, 'prescription_id', None):
        return 'pharmacy'
    if code:
        if code in _PHARMACY_ITEM_CODES:
            return 'pharmacy'
        if code.startswith(('WALKIN-', 'DRUG-', 'PHARM', 'RX-', 'MED-', 'PHARMA')):
            return 'pharmacy'
    if any(k in cat_raw for k in ('pharm', 'drug', 'medication', 'dispens')):
        return 'pharmacy'

    # Now safe to run gynae description heuristics (pharmacy lines already returned).
    if any(
        x in desc_early
        for x in (
            'antenatal (special payment)',
            'gynae / special (special payment)',
            'gynae / special',
            'gynaecology (special)',
        )
    ):
        return 'gynecology'
    if 'special payment' in desc_early and (
        'antenatal' in desc_early or 'gynae' in desc_early or 'gynaec' in desc_early
    ):
        return 'gynecology'
    if code:
        if code in _ADMISSION_ITEM_CODES or code.startswith('BED-') or code.startswith('ADM-ACCOM'):
            return 'admission'
        if code == 'REG' or code.startswith('REG-'):
            return 'registration'
        if code == 'PHYSIO' or code.startswith('PHYSIO'):
            return 'physiotherapy'
        # Ward/clinical consumables billed as inventory-style lines (not bed ADM-CONSUMABLES)
        if code.startswith('CS-') or 'clinical consum' in cat_raw:
            return 'consumables'
        if code == 'CASH-MISC' or code.startswith('CASH-MISC'):
            desc0 = (getattr(line, 'description', '') or '').casefold()
            if 'dressing' in desc0 or 'misc' in desc0:
                return 'consultation'
        if code.startswith('LAB-') or code.startswith('LABTEST-') or (code.startswith('LAB') and len(code) >= 6):
            if not code.startswith('LABEL'):
                return 'lab'
        if len(code) >= 4 and code.startswith('SL') and code[2:3].isdigit():
            return 'lab'
        if code.startswith(('IMG-', 'IMGCAT-', 'RAD-', 'ECG')) or (code.startswith('IMG') and len(code) >= 5):
            return 'imaging'
        if code.startswith('DRS'):
            return 'consultation'
        if code.startswith('PROC') and not code.startswith('PROCUREMENT'):
            return 'surgery'
        if code == 'MAT-ANC':
            return 'gynecology'
        if code == 'CON002':
            d = (getattr(line, 'description', '') or '').casefold()
            if any(
                x in d
                for x in (
                    'gynae',
                    'gynaec',
                    'gynec',
                    'special payment',
                    '/ special',
                )
            ):
                return 'gynecology'
    if _invoice_line_is_laboratory(line):
        return 'lab'
    desc = (getattr(line, 'description', '') or '').casefold()
    if 'consumable' in desc:
        return 'consumables'
    return None


def _specialty_consult_stream_from_context(line, encounter):
    """
    Map specialist-style consultation revenue to the right stream using encounter type,
    provider department/specialization, and line description (matches views_specialists.py).
    Returns None only when nothing more specific than general consultation applies.
    """
    desc = (getattr(line, 'description', '') or '').casefold()
    et = (getattr(encounter, 'encounter_type', None) or '').strip().lower() if encounter else ''

    if et in ('antenatal', 'gynae'):
        return 'gynecology'
    if et == 'er':
        return 'emergency'
    if et == 'surgery':
        return 'surgery'

    prov = getattr(encounter, 'provider', None) if encounter else None
    dept_name = ''
    spec = ''
    if prov:
        dept = getattr(prov, 'department', None)
        dept_name = (getattr(dept, 'name', None) or '').casefold()
        spec = (getattr(prov, 'specialization', None) or '').casefold()
    blob = f'{dept_name} {spec} {desc}'

    if any(
        x in blob
        for x in (
            'dental',
            'dentistry',
            'odont',
            'oral surgery',
            ' oral ',
            'tooth',
        )
    ):
        return 'dental'
    if any(
        x in blob
        for x in (
            'psychiatric',
            'psychiatry',
            'mental health',
            'mental-health',
            'behavioural',
            'behavioral',
        )
    ):
        return 'psychiatry'
    if any(
        x in blob
        for x in (
            'gynae',
            'gynaec',
            'gynec',
            'obstetr',
            'ob-gyn',
            'obgyn',
            'matern',
            'antenatal',
            ' anc ',
            'midwif',
            'obstetric',
        )
    ):
        return 'gynecology'

    return None


def _bucket_for_consultation_like_line(line, encounter, raw_cat_label):
    """
    Route MAT-ANC / CON001 / CON002 (and other rows shown as Consultation) into the
    correct revenue stream instead of lumping everything into general Consultation.
    """
    try:
        from .utils_billing import CONSULTATION_LINE_SERVICE_CODES
    except Exception:
        return None

    consult_codes = frozenset(c.strip().upper() for c in CONSULTATION_LINE_SERVICE_CODES if c)
    sc = getattr(line, 'service_code', None)
    code = (getattr(sc, 'code', '') or '').strip().upper()
    cat_key = str(raw_cat_label or '').strip().casefold()

    if code == 'MAT-ANC':
        return 'gynecology'

    if code not in consult_codes:
        if 'consult' not in cat_key:
            return None
        return _specialty_consult_stream_from_context(line, encounter) or 'consultation'

    if code in ('CON001', 'CONS-GEN') or code == 'S00023':
        return 'consultation'

    # CON002: specialist visit fee — cashier "Gynae / Special" invoices often have no encounter;
    # route those (and explicit gynae encounter types) to Gynecology, not general Consultation.
    if code == 'CON002':
        desc = (getattr(line, 'description', '') or '').casefold()
        if any(
            x in desc
            for x in (
                'gynae',
                'gynaec',
                'gynec',
                'special (special payment)',
                'special payment',
                '/ special',
            )
        ):
            return 'gynecology'
        et = (getattr(encounter, 'encounter_type', None) or '').strip().lower() if encounter else ''
        if et == 'gynae':
            return 'gynecology'
        spec = _specialty_consult_stream_from_context(line, encounter)
        if spec:
            return spec
        return 'consultation'

    return _specialty_consult_stream_from_context(line, encounter) or 'consultation'


def _infer_stream_from_code_category_desc(line, cat_key):
    """
    Deep routing for invoice lines that still have no stream after display-category rules.
    Uses ServiceCode.code patterns (LAB-/SL*/IMG-/WALKIN-/…), ServiceCode.category, and
    common description keywords so revenue does not fall into Other by default.
    """
    sc = getattr(line, 'service_code', None)
    code = (getattr(sc, 'code', '') or '').strip().upper()
    desc = (getattr(line, 'description', '') or '').casefold()
    cat_sc = (getattr(sc, 'category', None) or '').strip().casefold() if sc else ''
    ck = (cat_key or '').strip().casefold()
    blob = f'{code} {desc} {cat_sc} {ck}'

    # Pharmacy code/category short-circuit must run before any lab/gynae heuristics
    # so a pharmacy line whose description contains a generic test/gynae word cannot
    # be diverted out of Pharmacy.
    if getattr(line, 'prescription_id', None):
        return 'pharmacy'
    if code in _PHARMACY_ITEM_CODES:
        return 'pharmacy'
    if code.startswith(('WALKIN-', 'DRUG-', 'PHARM', 'RX-', 'MED-', 'PHARMA')):
        return 'pharmacy'
    if any(k in cat_sc for k in ('pharm', 'drug', 'medication', 'dispens')):
        return 'pharmacy'

    if any(
        x in desc
        for x in (
            'antenatal (special payment)',
            'gynae / special (special payment)',
            'gynae / special',
            'gynaecology (special)',
        )
    ):
        return 'gynecology'

    if _invoice_line_is_laboratory(line, ck):
        return 'lab'

    if code.startswith('LAB-') or code.startswith('LABTEST-') or (code.startswith('LAB') and len(code) >= 6):
        if not code.startswith('LABEL'):
            return 'lab'
    if len(code) >= 4 and code.startswith('SL') and code[2:3].isdigit():
        return 'lab'
    if code.startswith(('IMG-', 'IMGCAT-', 'RAD-', 'ECG')) or (code.startswith('IMG') and len(code) >= 5):
        return 'imaging'
    if code.startswith('DRS'):
        return 'consultation'
    if code.startswith('PROC') and not code.startswith('PROCUREMENT'):
        return 'surgery'

    if 'laborat' in cat_sc or cat_sc in ('lab', 'pathology', 'histology', 'microbiology', 'biochemistry'):
        return 'lab'
    if 'pharm' in cat_sc or 'drug' in cat_sc or 'medication' in cat_sc or 'dispens' in cat_sc:
        return 'pharmacy'
    if 'radio' in cat_sc or 'imag' in cat_sc or ('scan' in cat_sc and 'path' not in cat_sc):
        return 'imaging'
    if 'dental' in cat_sc:
        return 'dental'
    if 'gyn' in cat_sc or 'obstet' in cat_sc or 'matern' in cat_sc or 'antenatal' in cat_sc:
        return 'gynecology'
    if 'psych' in cat_sc:
        return 'psychiatry'
    if 'surg' in cat_sc or 'theatre' in cat_sc or 'theater' in cat_sc or 'operation' in cat_sc:
        return 'surgery'
    if 'emerg' in cat_sc:
        return 'emergency'
    if 'ambul' in cat_sc:
        return 'ambulance'
    if 'accommod' in cat_sc or 'inpatient' in cat_sc:
        return 'admission'
    if 'consum' in cat_sc and 'clinical' in cat_sc:
        return 'pharmacy'
    if 'nursing' in cat_sc and 'midwif' not in cat_sc:
        return 'consultation'

    if 'pathology' in ck or 'histology' in ck or 'microbiology' in ck:
        return 'lab'
    if 'nursing' in ck:
        return 'consultation'
    if 'medical' in ck and 'record' not in ck:
        return 'consultation'
    if 'diagnostic' in ck:
        if any(x in blob for x in ('imag', 'radio', 'scan', 'x-ray', 'xray', 'ultrasound', 'mri', 'ct ')):
            return 'imaging'
        return 'lab'
    if any(x in ck or x in desc for x in ('endoscop', 'colonoscopy', 'bronchoscopy')):
        return 'surgery'
    if any(x in desc for x in ('ultrasound', 'x-ray', 'xray', 'echocardi', ' mri ', ' ct scan')):
        return 'imaging'
    if any(
        x in desc
        for x in (
            'infusion',
            'injection',
            'syrup',
            'suspension',
            'suppository',
            'tablet',
            'capsule',
            ' i.v',
            ' iv ',
        )
    ):
        return 'pharmacy'
    if any(x in desc for x in ('blood sugar', 'fbc', 'full blood', 'lft', 'kft', 'lipid', 'troponin', 'urine re')):
        return 'lab'

    return None


def _classify_invoice_line_to_stream(line, encounter):
    """
    Single line → revenue stream key (including 'other'). Used for allocation sums
    and invoice-line drilldown so the UI matches receipt splits.
    """
    # Hard pharmacy short-circuit: any line that is unambiguously pharmacy must always
    # land in Pharmacy regardless of how downstream rules read its description, so
    # walk-in / dispensed items can never bleed into Lab / Gynecology / Other.
    sc = getattr(line, 'service_code', None)
    code_u = (getattr(sc, 'code', '') or '').strip().upper()
    cat_u = (getattr(sc, 'category', '') or '').strip().casefold()
    if (
        getattr(line, 'prescription_id', None)
        or code_u in _PHARMACY_ITEM_CODES
        or code_u.startswith(('WALKIN-', 'DRUG-', 'PHARM', 'RX-', 'MED-', 'PHARMA'))
        or any(k in cat_u for k in ('pharm', 'drug', 'medication', 'dispens'))
    ):
        return 'pharmacy'

    st = _bucket_for_invoice_line(line)
    if st:
        return st
    try:
        from .utils_billing import invoice_line_display_category
    except Exception:
        return 'other'
    raw_cat = invoice_line_display_category(line) or 'Other'
    cat_key = str(raw_cat).strip().casefold()

    st = _bucket_for_consultation_like_line(line, encounter, raw_cat)
    if st:
        return st
    st = _infer_stream_from_code_category_desc(line, cat_key) or ''
    if st:
        return st

    desc_key = (getattr(line, 'description', '') or '').casefold()
    code_u = (
        (getattr(getattr(line, 'service_code', None), 'code', None) or '').strip().upper()
    )
    if _invoice_line_is_laboratory(line, cat_key):
        return 'lab'
    if 'consult' in cat_key:
        return 'consultation'
    if 'clinical' in cat_key and 'consum' in cat_key:
        return 'consumables'
    if 'clinical' in cat_key:
        return 'consultation'
    if 'treatment' in cat_key:
        if code_u == 'REG' or 'registration' in desc_key:
            return 'registration'
        return 'consultation'
    if cat_key == 'other' and ('dressing' in desc_key or code_u.startswith('CASH-MISC')):
        return 'consultation'
    if 'therapy' in cat_key:
        if 'psych' in cat_key or 'psych' in desc_key or 'mental' in desc_key:
            return 'psychiatry'
        if 'physio' in cat_key or 'physio' in desc_key or 'rehab' in desc_key:
            return 'physiotherapy'
        return 'physiotherapy'
    if 'lab' in cat_key or 'laborat' in cat_key:
        return 'lab'
    if 'pharm' in cat_key or 'drug' in cat_key or 'medic' in cat_key:
        return 'pharmacy'
    if any(
        x in cat_key
        for x in ('supply', 'supplies', 'inventory', 'procurement', 'reagent', 'consumable stock')
    ):
        return 'lab'
    if 'imag' in cat_key or 'radiol' in cat_key or 'scan' in cat_key or 'x-ray' in cat_key or 'xray' in cat_key:
        return 'imaging'
    if 'dental' in cat_key:
        return 'dental'
    if 'gyn' in cat_key or 'obst' in cat_key:
        return 'gynecology'
    if 'psych' in cat_key:
        return 'psychiatry'
    if 'surg' in cat_key or 'procedure' in cat_key or 'theatre' in cat_key:
        return 'surgery'
    if 'emerg' in cat_key:
        return 'emergency'
    if (
        'admission' in cat_key
        or 'ward' in cat_key
        or 'bed' in cat_key
        or 'accommod' in cat_key
        or 'detent' in cat_key
    ):
        return 'admission'
    return 'other'


_SERVICE_DISPLAY = {
    'consultation': 'Consultation',
    'lab': 'Laboratory',
    'pharmacy': 'Pharmacy',
    'imaging': 'Imaging',
    'dental': 'Dental',
    'gynecology': 'Gynecology',
    'psychiatry': 'Psychiatry',
    'surgery': 'Surgery',
    'emergency': 'Emergency',
    'ambulance': 'Ambulance',
    'admission': 'Admission',
    'physiotherapy': 'Physiotherapy / therapy',
    'registration': 'Registration & fees',
    'other': 'Other',
}

_SERVICE_COLOR = {
    'consultation': '#3b82f6',
    'lab': '#10b981',
    'pharmacy': '#f59e0b',
    'imaging': '#8b5cf6',
    'dental': '#06b6d4',
    'gynecology': '#ec4899',
    'psychiatry': '#a855f7',
    'surgery': '#ef4444',
    'emergency': '#f97316',
    'ambulance': '#dc2626',
    'admission': '#64748b',
    'physiotherapy': '#14b8a6',
    'registration': '#78716c',
    'other': '#6b7280',
}

# Canonical service stream → department label for "Revenue by department" (same totals
# as the receipt allocation logic; not the accounting Revenue journal table).
_SERVICE_STREAM_TO_DEPARTMENT = {
    'consultation': 'Outpatient & Consultation',
    'lab': 'Laboratory',
    'pharmacy': 'Pharmacy',
    'imaging': 'Radiology & Imaging',
    'dental': 'Dental',
    'gynecology': 'Obstetrics & Gynaecology',
    'psychiatry': 'Psychiatry & Mental Health',
    'surgery': 'Surgery & Theatre',
    'emergency': 'Emergency',
    'ambulance': 'Ambulance',
    'admission': 'Wards & Inpatient',
    'physiotherapy': 'Physiotherapy',
    'registration': 'Administration & Records',
    'other': 'Unallocated & Other',
}


def _department_for_service_stream(stream_key):
    sk = stream_key if stream_key in _SERVICE_DISPLAY else 'other'
    return _SERVICE_STREAM_TO_DEPARTMENT.get(sk, _SERVICE_STREAM_TO_DEPARTMENT['other'])


def _normalize_service_type(raw):
    if not raw:
        return 'other'
    key = str(raw).strip().lower()
    key = key.replace(' ', '_').replace('-', '_')
    if key == 'combined':
        return 'combined'
    normalized = _SERVICE_TYPE_ALIASES.get(key)
    if normalized:
        return normalized
    return key if key in _SERVICE_DISPLAY else 'other'


def _resolve_invoice_for_entry(svc):
    """Resolve an Invoice from a service_details.services[] entry by id, or by parsing the
    invoice number out of the entry name (e.g. 'Invoice INV20260306899')."""
    try:
        from .models import Invoice
    except Exception:
        return None
    sid = svc.get('service_id') or svc.get('id')
    if sid:
        try:
            inv = Invoice.objects.filter(id=sid).first()
        except Exception:
            inv = None
        if inv is not None:
            return inv
    name = str(svc.get('name') or '')
    if name:
        import re
        m = re.search(r'(INV[A-Za-z0-9_-]+)', name)
        if m:
            try:
                return Invoice.objects.filter(invoice_number=m.group(1)).first()
            except Exception:
                return None
    return None


# Fixed nightly ratio fallback when the encounter invoice has no ADM-* lines (rare;
# mostly applies to live admissions). Mirrors BedBillingService nightly constants:
# accom 150 + doctor 80 + nursing 70 + consumables 50 = 350.
_BED_FALLBACK_TOTAL = Decimal('350')
_BED_FALLBACK_CONSUMABLES = Decimal('50')


def _split_bed_entry(svc, entry_amt):
    """
    Resolve the Admission from a 'bed' entry and split entry_amt across canonical buckets
    using ADM-* line totals on the encounter invoice (preferred) or the nightly-rate
    ratio (fallback). ADM-CONSUMABLES portion -> pharmacy; the rest -> admission.
    Returns None if the admission cannot be resolved.
    """
    if entry_amt <= 0:
        return None
    try:
        from .models import Admission, InvoiceLine
    except Exception:
        return None

    sid = svc.get('service_id') or svc.get('id')
    admission = None
    if sid:
        try:
            admission = Admission.objects.filter(id=sid).only('id', 'encounter_id').first()
        except Exception:
            admission = None
    if admission is None or not getattr(admission, 'encounter_id', None):
        return None

    code_totals = {}
    try:
        adm_codes = list(_ADMISSION_ITEM_CODES) + list(_PHARMACY_ITEM_CODES)
        lines = (
            InvoiceLine.objects.filter(
                invoice__encounter_id=admission.encounter_id,
                is_deleted=False,
                waived_at__isnull=True,
                service_code__code__in=adm_codes,
            )
            .select_related('service_code')
            .only('line_total', 'service_code__code')
        )
        for line in lines:
            try:
                code = (line.service_code.code or '').strip().upper()
                lt = Decimal(str(line.line_total or 0))
            except Exception:
                continue
            if code and lt > 0:
                code_totals[code] = code_totals.get(code, Decimal('0.00')) + lt
    except Exception:
        code_totals = {}

    if not code_totals:
        # Fallback: split using the standard nightly ratio so consumables still flow to pharmacy.
        consumables_share = entry_amt * (_BED_FALLBACK_CONSUMABLES / _BED_FALLBACK_TOTAL)
        admission_share = entry_amt - consumables_share
        return {'admission': admission_share, 'pharmacy': consumables_share}

    total = sum(code_totals.values(), Decimal('0.00'))
    if total <= 0:
        return {'admission': entry_amt}

    out = {}
    allocated = Decimal('0.00')
    for code, sub in code_totals.items():
        share = entry_amt * (sub / total)
        bucket = 'pharmacy' if code in _PHARMACY_ITEM_CODES else 'admission'
        out[bucket] = out.get(bucket, Decimal('0.00')) + share
        allocated += share
    remainder = entry_amt - allocated
    if remainder != 0:
        out['admission'] = out.get('admission', Decimal('0.00')) + remainder
    return out


def _expand_combined_entry(svc, raw_type_key, entry_amt):
    """
    Drill into a single service_details.services[] entry and return a dict mapping
    canonical bucket keys to portions of entry_amt. Returns None when no expansion
    applies (caller falls back to the alias map).
    """
    if not svc or entry_amt <= 0:
        return None
    if raw_type_key in ('invoice', 'invoice_line'):
        invoice = _resolve_invoice_for_entry(svc)
        if invoice is None:
            return None
        sub, sub_total = _infer_alloc_from_invoice(invoice, entry_amt)
        if sub and sub_total > 0:
            remainder = entry_amt - sub_total
            if remainder != 0:
                sub['other'] = sub.get('other', Decimal('0.00')) + remainder
            return sub
        return None
    if raw_type_key == 'bed':
        return _split_bed_entry(svc, entry_amt)
    return None


def _alloc_from_combined_receipt(rec):
    """
    Returns (allocations_dict, allocated_total).
    allocations_dict maps canonical service keys -> Decimal amount.
    Whole-invoice and bed-charge entries are expanded into their underlying buckets so
    a combined payment of multiple invoices is split per-line instead of dumped into
    'Other'.
    """
    alloc = {}
    services = (getattr(rec, 'service_details', None) or {}).get('services') or []
    if not services:
        return alloc, Decimal('0.00')

    total_services = Decimal('0.00')
    parsed = []  # list of (entry_amt, sub_alloc_dict {bucket: portion of entry_amt})
    for svc in services:
        try:
            amt = Decimal(str(svc.get('price', 0) or 0))
        except (TypeError, ValueError):
            continue
        if amt <= 0:
            continue

        raw_type = (svc.get('type') or '').strip().lower().replace(' ', '_').replace('-', '_')
        sub_alloc = _expand_combined_entry(svc, raw_type, amt)
        if not sub_alloc:
            stype = _normalize_service_type(svc.get('type') or '')
            if stype == 'combined':
                stype = 'other'
            sub_alloc = {stype: amt}

        parsed.append((amt, sub_alloc))
        total_services += amt

    if total_services <= 0:
        return alloc, Decimal('0.00')

    try:
        receipt_amt = Decimal(str(getattr(rec, 'amount_paid', 0) or 0))
    except (TypeError, ValueError):
        receipt_amt = total_services

    # Scale allocations so the sum matches the receipt amount paid.
    scale = (receipt_amt / total_services) if total_services > 0 else Decimal('1.00')

    allocated_total = Decimal('0.00')
    for _entry_amt, sub_alloc in parsed:
        for stype, sub_amt in sub_alloc.items():
            a = (sub_amt * scale)
            alloc[stype] = alloc.get(stype, Decimal('0.00')) + a
            allocated_total += a
    return alloc, allocated_total


def _infer_alloc_from_invoice(invoice, receipt_amt):
    """
    Infer service allocations from invoice lines and allocate receipt_amt proportionally.
    Returns (allocations_dict, allocated_total).
    """
    if not invoice:
        return {}, Decimal('0.00')

    try:
        from .models import InvoiceLine, Encounter
    except Exception:
        return {}, Decimal('0.00')

    encounter = None
    if getattr(invoice, 'encounter_id', None):
        try:
            encounter = (
                Encounter.objects.filter(pk=invoice.encounter_id)
                .select_related('provider__department')
                .first()
            )
        except Exception:
            encounter = getattr(invoice, 'encounter', None)

    # Pull minimal line data; display_category requires service_code and some fields.
    lines = (
        InvoiceLine.objects.filter(invoice_id=invoice.id, is_deleted=False)
        .select_related('service_code')
        .only(
            'id',
            'quantity',
            'unit_price',
            'discount_amount',
            'tax_amount',
            'line_total',
            'description',
            'prescription_id',
            'service_code__code',
            'service_code__category',
        )
    )

    cat_sums = {}  # canonical -> Decimal billed-like sum
    billed_total = Decimal('0.00')

    for line in lines:
        if getattr(line, 'waived_at', None):
            continue
        # Use display_line_total when available (handles stale repricing).
        try:
            line_total = Decimal(str(getattr(line, 'display_line_total', None) or line.line_total or 0))
        except Exception:
            line_total = Decimal('0.00')
        if line_total <= 0:
            continue

        st = _classify_invoice_line_to_stream(line, encounter)
        cat_sums[st] = cat_sums.get(st, Decimal('0.00')) + line_total
        billed_total += line_total

    if billed_total <= 0:
        return {}, Decimal('0.00')

    alloc = {}
    allocated_total = Decimal('0.00')
    for st, subtotal in cat_sums.items():
        if subtotal <= 0:
            continue
        amt = (receipt_amt * (subtotal / billed_total))
        alloc[st] = alloc.get(st, Decimal('0.00')) + amt
        allocated_total += amt

    return alloc, allocated_total


def _alloc_for_receipt(rec):
    """
    Canonical allocation for ONE receipt. Ensures allocations sum to amount_paid
    (remainder goes to 'other').
    """
    try:
        receipt_amt = Decimal(str(getattr(rec, 'amount_paid', 0) or 0))
    except Exception:
        receipt_amt = Decimal('0.00')

    raw = getattr(rec, 'service_type', None)
    st = _normalize_service_type(raw)

    if st != 'other' and st != 'combined' and st in _SERVICE_DISPLAY:
        return {st: receipt_amt}

    alloc, _ = _alloc_from_combined_receipt(rec)
    if not alloc:
        alloc, _ = _infer_alloc_from_invoice(getattr(rec, 'invoice', None), receipt_amt)

    if not alloc:
        return {'other': receipt_amt}

    out = {}
    for k, v in (alloc or {}).items():
        k2 = k if k in _SERVICE_DISPLAY else 'other'
        out[k2] = out.get(k2, Decimal('0.00')) + (v or Decimal('0.00'))
    remainder = receipt_amt - sum(out.values(), Decimal('0.00'))
    if remainder != 0:
        out['other'] = out.get('other', Decimal('0.00')) + remainder
    return out


def _invoice_line_category_debug(invoice):
    """
    Returns list of {'category': str, 'total': float} sorted desc for UI debug.
    """
    if not invoice:
        return []
    try:
        from .models import InvoiceLine
        from .utils_billing import invoice_line_display_category
    except Exception:
        return []

    lines = (
        InvoiceLine.objects.filter(invoice_id=invoice.id, is_deleted=False)
        .select_related('service_code')
        .only(
            'id',
            'quantity',
            'unit_price',
            'discount_amount',
            'tax_amount',
            'line_total',
            'description',
            'prescription_id',
            'service_code__code',
            'service_code__category',
        )
    )
    sums = {}
    for line in lines:
        if getattr(line, 'waived_at', None):
            continue
        try:
            line_total = Decimal(str(getattr(line, 'display_line_total', None) or line.line_total or 0))
        except Exception:
            line_total = Decimal('0.00')
        if line_total <= 0:
            continue
        cat = invoice_line_display_category(line) or 'Other'
        sums[cat] = sums.get(cat, Decimal('0.00')) + line_total

    out = [{'category': k, 'total': float(v)} for k, v in sums.items()]
    out.sort(key=lambda r: r['total'], reverse=True)
    return out


def _invoice_line_items_debug(invoice, limit=80):
    """
    Returns a list of invoice line items for UI details:
    [{'code','category','revenue_stream','description','qty','unit','total','waived','is_pharmacy'}]
    Uses display totals to handle stale repricing/sync.
    """
    if not invoice:
        return []
    try:
        from .models import InvoiceLine, Encounter
        from .utils_billing import invoice_line_display_category
    except Exception:
        return []

    encounter = None
    if getattr(invoice, 'encounter_id', None):
        try:
            encounter = (
                Encounter.objects.filter(pk=invoice.encounter_id)
                .select_related('provider__department')
                .first()
            )
        except Exception:
            encounter = getattr(invoice, 'encounter', None)

    lines = (
        InvoiceLine.objects.filter(invoice_id=invoice.id, is_deleted=False)
        .select_related('service_code')
        .only(
            'id',
            'quantity',
            'unit_price',
            'discount_amount',
            'tax_amount',
            'line_total',
            'description',
            'prescription_id',
            'waived_at',
            'service_code__code',
            'service_code__category',
        )
        .order_by('created')
    )

    out = []
    for line in lines[: max(1, int(limit or 80))]:
        sc = getattr(line, 'service_code', None)
        code = (getattr(sc, 'code', '') or '').strip()
        cat = invoice_line_display_category(line) or (getattr(sc, 'category', None) or 'Other')

        try:
            qty = Decimal(str(getattr(line, 'quantity', 0) or 0))
        except Exception:
            qty = Decimal('0.00')
        try:
            unit = Decimal(str(getattr(line, 'display_unit_price', None) or line.unit_price or 0))
        except Exception:
            unit = Decimal('0.00')
        try:
            total = Decimal(str(getattr(line, 'display_line_total', None) or line.line_total or 0))
        except Exception:
            total = Decimal('0.00')

        waived = bool(getattr(line, 'waived_at', None))
        if waived:
            rstream = 'Waived'
        else:
            sk = _classify_invoice_line_to_stream(line, encounter)
            rstream = _SERVICE_DISPLAY.get(sk, sk)

        out.append({
            'code': code,
            'category': cat,
            'revenue_stream': rstream,
            'description': (getattr(line, 'description', None) or ''),
            'qty': float(qty),
            'unit': float(unit),
            'total': float(total),
            'waived': waived,
            'is_pharmacy': bool(getattr(line, 'prescription_id', None)),
        })

    return out


def _build_revenue_streams_analytics(date_from, date_to):
    """
    Single source of truth for the Revenue Streams dashboard:
    - KPI totals (service buckets)
    - by-service breakdown (allocated combined)
    - daily trend totals and by-service stacked series
    - by-department breakdown (derived from the same receipt allocation → department labels)
    """
    try:
        from .models_accounting import PaymentReceipt
    except Exception:
        PaymentReceipt = None

    service_totals = {k: Decimal('0.00') for k in _SERVICE_DISPLAY.keys()}
    service_counts = {k: 0 for k in _SERVICE_DISPLAY.keys()}
    daily_totals = {}  # date -> Decimal
    daily_by_service = {}  # date -> {service_key -> Decimal}
    other_reason_totals = {}  # reason -> Decimal
    other_reason_counts = {}  # reason -> int
    raw_service_type_totals = {}  # raw service_type -> Decimal
    raw_service_type_counts = {}  # raw service_type -> int
    deposit_totals = Decimal('0.00')
    deposit_counts = 0
    dept_totals = defaultdict(Decimal)
    dept_aggregate_txns = defaultdict(int)
    dept_receipt_ids = defaultdict(set)

    if PaymentReceipt:
        base = PaymentReceipt.objects.filter(
            receipt_date__date__gte=date_from,
            receipt_date__date__lte=date_to,
            is_deleted=False,
        )

        known_raw = set(_SERVICE_TYPE_ALIASES.keys()) | set(_SERVICE_DISPLAY.keys())
        known_raw.add('combined')

        # Fast path: non-combined receipts (still normalized to canonical buckets)
        non_combined = base.exclude(service_type='combined').exclude(service_type__isnull=True).exclude(service_type='')
        day_rows = (
            non_combined
            .annotate(day=TruncDate('receipt_date'))
            .values('day', 'service_type')
            .annotate(total=Sum('amount_paid'), count=Count('id'))
        )
        for row in day_rows:
            day = row.get('day')
            raw_st = row.get('service_type')
            st = _normalize_service_type(raw_st)
            amt = row.get('total') or Decimal('0.00')
            cnt = row.get('count') or 0
            service_totals[st] = service_totals.get(st, Decimal('0.00')) + amt
            service_counts[st] = service_counts.get(st, 0) + cnt
            daily_totals[day] = daily_totals.get(day, Decimal('0.00')) + amt
            if day not in daily_by_service:
                daily_by_service[day] = {}
            daily_by_service[day][st] = daily_by_service[day].get(st, Decimal('0.00')) + amt

            raw_key = (str(raw_st).strip().lower() if raw_st else 'unclassified')
            raw_service_type_totals[raw_key] = raw_service_type_totals.get(raw_key, Decimal('0.00')) + amt
            raw_service_type_counts[raw_key] = raw_service_type_counts.get(raw_key, 0) + cnt

            st2 = st if st in _SERVICE_DISPLAY else 'other'
            dept = _department_for_service_stream(st2)
            dept_totals[dept] += amt
            dept_aggregate_txns[dept] += int(cnt or 0)

        combined_qs = base.filter(service_type='combined').only(
            'id', 'receipt_date', 'amount_paid', 'service_details'
        )
        for rec in combined_qs:
            dt = getattr(rec, 'receipt_date', None)
            day = dt.date() if dt else None
            if not day:
                continue

            alloc, allocated_total = _alloc_from_combined_receipt(rec)
            try:
                receipt_amt = Decimal(str(getattr(rec, 'amount_paid', 0) or 0))
            except (TypeError, ValueError):
                receipt_amt = allocated_total

            if day not in daily_by_service:
                daily_by_service[day] = {}

            # If combined receipt has no service_details, infer from invoice/bill lines.
            if not alloc:
                alloc, allocated_total = _infer_alloc_from_invoice(getattr(rec, 'invoice', None), receipt_amt)

            if alloc:
                touched_depts = set()
                for st, amt in alloc.items():
                    st2 = st if st in _SERVICE_DISPLAY else 'other'
                    service_totals[st2] = service_totals.get(st2, Decimal('0.00')) + amt
                    daily_by_service[day][st2] = daily_by_service[day].get(st2, Decimal('0.00')) + amt
                    daily_totals[day] = daily_totals.get(day, Decimal('0.00')) + amt
                    dlab = _department_for_service_stream(st2)
                    dept_totals[dlab] += amt
                    touched_depts.add(dlab)

                remainder = receipt_amt - allocated_total
                if remainder != 0:
                    service_totals['other'] += remainder
                    daily_by_service[day]['other'] = daily_by_service[day].get('other', Decimal('0.00')) + remainder
                    daily_totals[day] = daily_totals.get(day, Decimal('0.00')) + remainder
                    other_reason_totals['combined_rounding_or_unmapped'] = other_reason_totals.get('combined_rounding_or_unmapped', Decimal('0.00')) + remainder
                    other_reason_counts['combined_rounding_or_unmapped'] = other_reason_counts.get('combined_rounding_or_unmapped', 0) + 1
                    d_other = _department_for_service_stream('other')
                    dept_totals[d_other] += remainder
                    touched_depts.add(d_other)
                for d in touched_depts:
                    dept_receipt_ids[d].add(rec.id)
            else:
                service_totals['other'] += receipt_amt
                daily_by_service[day]['other'] = daily_by_service[day].get('other', Decimal('0.00')) + receipt_amt
                daily_totals[day] = daily_totals.get(day, Decimal('0.00')) + receipt_amt
                service_counts['other'] = service_counts.get('other', 0) + 1
                other_reason_totals['combined_missing_details_and_invoice'] = other_reason_totals.get('combined_missing_details_and_invoice', Decimal('0.00')) + receipt_amt
                other_reason_counts['combined_missing_details_and_invoice'] = other_reason_counts.get('combined_missing_details_and_invoice', 0) + 1
                d_other = _department_for_service_stream('other')
                dept_totals[d_other] += receipt_amt
                dept_receipt_ids[d_other].add(rec.id)

            raw_service_type_totals['combined'] = raw_service_type_totals.get('combined', Decimal('0.00')) + receipt_amt
            raw_service_type_counts['combined'] = raw_service_type_counts.get('combined', 0) + 1

        # Improve classification of "Other"/unknown receipts:
        # - If service_details has services[] even when service_type='other', split it
        # - Else infer based on invoice line mix and allocate proportionally
        unknown_qs = (
            base.exclude(service_type='combined')
            .filter(
                Q(service_type__isnull=True)
                | Q(service_type='')
                | Q(service_type='other')
                | ~Q(service_type__in=list(known_raw))
            )
            .select_related('invoice')
            .only(
                'id',
                'receipt_date',
                'amount_paid',
                'service_type',
                'service_details',
                'invoice_id',
                'payment_method',
            )
        )
        for rec in unknown_qs:
            dt = getattr(rec, 'receipt_date', None)
            day = dt.date() if dt else None
            if not day:
                continue
            try:
                receipt_amt = Decimal(str(getattr(rec, 'amount_paid', 0) or 0))
            except Exception:
                receipt_amt = Decimal('0.00')
            if receipt_amt <= 0:
                continue

            if day not in daily_by_service:
                daily_by_service[day] = {}

            # Prefer explicit service_details.services allocation when present.
            alloc, allocated_total = _alloc_from_combined_receipt(rec)
            inferred_reason = None
            if not alloc:
                alloc, allocated_total = _infer_alloc_from_invoice(getattr(rec, 'invoice', None), receipt_amt)
                inferred_reason = 'inferred_from_invoice_lines' if alloc else None
            else:
                inferred_reason = 'inferred_from_service_details'

            if alloc:
                touched_depts = set()
                for st, amt in alloc.items():
                    st2 = st if st in _SERVICE_DISPLAY else 'other'
                    service_totals[st2] = service_totals.get(st2, Decimal('0.00')) + amt
                    daily_by_service[day][st2] = daily_by_service[day].get(st2, Decimal('0.00')) + amt
                    daily_totals[day] = daily_totals.get(day, Decimal('0.00')) + amt
                    dlab = _department_for_service_stream(st2)
                    dept_totals[dlab] += amt
                    touched_depts.add(dlab)

                remainder = receipt_amt - allocated_total
                if remainder != 0:
                    service_totals['other'] += remainder
                    daily_by_service[day]['other'] = daily_by_service[day].get('other', Decimal('0.00')) + remainder
                    daily_totals[day] = daily_totals.get(day, Decimal('0.00')) + remainder
                    other_reason_totals['unknown_rounding_or_unmapped'] = other_reason_totals.get('unknown_rounding_or_unmapped', Decimal('0.00')) + remainder
                    other_reason_counts['unknown_rounding_or_unmapped'] = other_reason_counts.get('unknown_rounding_or_unmapped', 0) + 1
                    d_other = _department_for_service_stream('other')
                    dept_totals[d_other] += remainder
                    touched_depts.add(d_other)
                for d in touched_depts:
                    dept_receipt_ids[d].add(rec.id)
            else:
                # Could not infer; keep as Other.
                service_totals['other'] += receipt_amt
                daily_by_service[day]['other'] = daily_by_service[day].get('other', Decimal('0.00')) + receipt_amt
                daily_totals[day] = daily_totals.get(day, Decimal('0.00')) + receipt_amt
                if not getattr(rec, 'invoice_id', None):
                    reason = 'no_invoice_link'
                elif not (getattr(rec, 'service_details', None) or {}).get('services'):
                    reason = 'no_service_details'
                else:
                    reason = 'unclassified'
                other_reason_totals[reason] = other_reason_totals.get(reason, Decimal('0.00')) + receipt_amt
                other_reason_counts[reason] = other_reason_counts.get(reason, 0) + 1
                d_other = _department_for_service_stream('other')
                dept_totals[d_other] += receipt_amt
                dept_receipt_ids[d_other].add(rec.id)

            if inferred_reason:
                other_reason_totals[inferred_reason] = other_reason_totals.get(inferred_reason, Decimal('0.00')) + Decimal('0.00')
                other_reason_counts[inferred_reason] = other_reason_counts.get(inferred_reason, 0) + 1

            raw_key = (str(getattr(rec, 'service_type', '')).strip().lower() or 'unclassified')
            raw_service_type_totals[raw_key] = raw_service_type_totals.get(raw_key, Decimal('0.00')) + receipt_amt
            raw_service_type_counts[raw_key] = raw_service_type_counts.get(raw_key, 0) + 1

        # Deposits: every receipt in range using the deposit payment rail or combined bills
        # with deposit_applied (was only counted on "unknown" receipts before).
        deposit_totals = Decimal('0.00')
        deposit_counts = 0
        for rec in base.only('amount_paid', 'payment_method', 'service_details').iterator(chunk_size=500):
            try:
                ra = Decimal(str(getattr(rec, 'amount_paid', 0) or 0))
            except Exception:
                ra = Decimal('0.00')
            if ra <= 0:
                continue
            try:
                sd = getattr(rec, 'service_details', None) or {}
                if getattr(rec, 'payment_method', None) == 'deposit' or sd.get('deposit_applied'):
                    deposit_totals += ra
                    deposit_counts += 1
            except Exception:
                pass

    dept_rows = []
    for dept_name in sorted(dept_totals.keys(), key=lambda k: dept_totals[k], reverse=True):
        tot = dept_totals[dept_name]
        if tot <= 0:
            continue
        txn = int(dept_aggregate_txns.get(dept_name, 0)) + len(dept_receipt_ids.get(dept_name, ()))
        dept_rows.append(
            {
                # Use 'department' in templates: Django parses item.department__name as
                # nested lookups (department, then __name), not a single dict key.
                'department': dept_name,
                'department__name': dept_name,
                'total': tot,
                'count': txn,
            }
        )

    total_revenue = sum((v or Decimal('0.00')) for v in service_totals.values())

    for r in dept_rows:
        amt = r.get('total') or Decimal('0.00')
        r['percentage'] = (amt / total_revenue * Decimal('100')) if total_revenue > 0 else Decimal('0.0')

    by_service_rows = []
    for st, label in _SERVICE_DISPLAY.items():
        amt = service_totals.get(st, Decimal('0.00')) or Decimal('0.00')
        if amt <= 0 and st != 'other':
            continue
        by_service_rows.append({
            'service_type': st,
            'service_name': label,
            'total': amt,
            'count': service_counts.get(st, 0),
            'percentage': (amt / total_revenue * Decimal('100')) if total_revenue > 0 else Decimal('0.0'),
            'color': _SERVICE_COLOR.get(st, '#6b7280'),
        })
    by_service_rows.sort(key=lambda r: r['total'], reverse=True)

    labels = []
    series_total = []
    series_by_service = {k: [] for k in _SERVICE_DISPLAY.keys()}
    cursor = date_from
    while cursor <= date_to:
        labels.append(cursor.isoformat())
        t = daily_totals.get(cursor, Decimal('0.00'))
        series_total.append(float(t))
        day_map = daily_by_service.get(cursor, {})
        for st in series_by_service.keys():
            series_by_service[st].append(float(day_map.get(st, Decimal('0.00'))))
        cursor = cursor + timedelta(days=1)

    dept_chart = []
    for r in dept_rows[:12]:
        dept_chart.append({
            'label': r.get('department') or r.get('department__name') or 'Unassigned',
            'total': float(r.get('total') or Decimal('0.00')),
            'count': int(r.get('count') or 0),
        })

    analytics = {
        'meta': {
            'date_from': date_from.isoformat(),
            'date_to': date_to.isoformat(),
            'days': (date_to - date_from).days + 1,
        },
        'kpis': {
            'total': float(total_revenue),
            'deposit_total': float(deposit_totals),
            'deposit_count': int(deposit_counts),
            **{k: float(service_totals.get(k, Decimal('0.00')) or Decimal('0.00')) for k in _SERVICE_DISPLAY.keys()},
        },
        'breakdowns': {
            'other_reasons': [
                {
                    'reason': k,
                    'count': int(other_reason_counts.get(k, 0)),
                    'total': float(other_reason_totals.get(k, Decimal('0.00')) or Decimal('0.00')),
                }
                for k in sorted(other_reason_totals.keys())
            ],
            'raw_service_types': [
                {
                    'service_type': k,
                    'count': int(raw_service_type_counts.get(k, 0)),
                    'total': float(raw_service_type_totals.get(k, Decimal('0.00')) or Decimal('0.00')),
                }
                for k in sorted(raw_service_type_totals.keys())
            ],
        },
        'tables': {
            'by_service': [
                {**r, 'total': float(r['total']), 'percentage': float(r['percentage'])}
                for r in by_service_rows
            ],
            'by_department': [
                {
                    'department': r.get('department') or r.get('department__name') or 'Unassigned',
                    'total': float(r.get('total') or Decimal('0.00')),
                    'count': int(r.get('count') or 0),
                    'percentage': float(((r.get('total') or Decimal('0.00')) / total_revenue * Decimal('100')) if total_revenue > 0 else 0),
                }
                for r in dept_rows
            ],
        },
        'charts': {
            'trend_total': {
                'labels': labels,
                'totals': series_total,
            },
            'trend_by_service': {
                'labels': labels,
                'series': series_by_service,
                'display': {k: _SERVICE_DISPLAY[k] for k in _SERVICE_DISPLAY.keys()},
                'colors': {k: _SERVICE_COLOR.get(k, '#6b7280') for k in _SERVICE_DISPLAY.keys()},
            },
            'service_share': {
                'labels': [r['service_name'] for r in by_service_rows],
                'totals': [float(r['total']) for r in by_service_rows],
                'colors': [r.get('color') for r in by_service_rows],
            },
            'top_departments': {
                'labels': [d['label'] for d in dept_chart],
                'totals': [d['total'] for d in dept_chart],
            },
        },
    }

    return analytics, by_service_rows, dept_rows, total_revenue


@login_required
@user_passes_test(finance_revenue_streams_access)
def revenue_streams_dashboard(request):
    """
    Main dashboard for revenue stream monitoring
    Shows breakdown by department and service type
    """
    # Date range (defaults to current month)
    today = timezone.now().date()
    default_from = today.replace(day=1)
    date_from = _parse_ymd(request.GET.get('date_from'), default_from)
    date_to = _parse_ymd(request.GET.get('date_to'), today)

    analytics_data, revenue_by_service, revenue_by_dept, total_revenue = _build_revenue_streams_analytics(date_from, date_to)
    
    # Get top revenue streams
    try:
        revenue_streams = RevenueStream.objects.filter(is_active=True)
        stream_performance = []
        
        for stream in revenue_streams:
            stream_revenue = Revenue.objects.filter(
                revenue_stream=stream,
                revenue_date__gte=date_from,
                revenue_date__lte=date_to,
                is_deleted=False
            ).aggregate(
                total=Sum('amount'),
                count=Count('id')
            )
            
            stream_performance.append({
                'stream': stream,
                'total': stream_revenue['total'] or Decimal('0.00'),
                'count': stream_revenue['count'] or 0,
                'target': stream.monthly_target,
                'achievement': ((stream_revenue['total'] or Decimal('0.00')) / stream.monthly_target * 100) if stream.monthly_target > 0 else 0
            })
        
        stream_performance.sort(key=lambda x: x['total'], reverse=True)
    except:
        stream_performance = []
    
    kpis = analytics_data.get('kpis') or {}
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'total_revenue': total_revenue,
        'revenue_by_service': revenue_by_service,
        'revenue_by_dept': revenue_by_dept,
        'stream_performance': stream_performance,
        'analytics_data': analytics_data,
        'kpis': kpis,
    }
    
    return render(request, 'hospital/revenue/streams_dashboard.html', context)


@login_required
@user_passes_test(finance_revenue_streams_access)
def revenue_streams_service_details(request):
    """
    Drilldown: list receipts contributing to a canonical service bucket.
    """
    today = timezone.now().date()
    default_from = today.replace(day=1)
    date_from = _parse_ymd(request.GET.get('date_from'), default_from)
    date_to = _parse_ymd(request.GET.get('date_to'), today)
    bucket = (request.GET.get('service') or '').strip().lower()
    if bucket not in _SERVICE_DISPLAY:
        bucket = 'other'
    raw_filter = (request.GET.get('raw') or '').strip().lower()

    try:
        from .models_accounting import PaymentReceipt
    except Exception:
        PaymentReceipt = None

    rows = []
    total_bucket = Decimal('0.00')
    limit = max(50, min(1000, int(request.GET.get('limit', 400) or 400)))

    if PaymentReceipt:
        qs = (
            PaymentReceipt.objects.filter(
                receipt_date__date__gte=date_from,
                receipt_date__date__lte=date_to,
                is_deleted=False,
            )
            .select_related('patient', 'invoice')
            .only(
                'id',
                'receipt_number',
                'receipt_date',
                'amount_paid',
                'payment_method',
                'service_type',
                'service_details',
                'patient__id',
                'patient__first_name',
                'patient__last_name',
                'invoice__id',
                'invoice__invoice_number',
            )
            .order_by('-receipt_date')
        )

        for rec in qs.iterator(chunk_size=500):
            raw_st = (getattr(rec, 'service_type', None) or 'unclassified')
            raw_st_key = str(raw_st).strip().lower() if raw_st else 'unclassified'
            if raw_filter and raw_filter != raw_st_key:
                continue

            alloc = _alloc_for_receipt(rec)
            amt = alloc.get(bucket) or Decimal('0.00')
            if amt <= 0:
                continue

            try:
                paid = Decimal(str(getattr(rec, 'amount_paid', 0) or 0))
            except Exception:
                paid = Decimal('0.00')

            total_bucket += amt

            details = getattr(rec, 'service_details', None) or {}
            deposit_flag = False
            try:
                deposit_flag = (getattr(rec, 'payment_method', None) == 'deposit') or bool(details.get('deposit_applied'))
            except Exception:
                deposit_flag = False

            patient = getattr(rec, 'patient', None)
            patient_name = None
            if patient:
                patient_name = (getattr(patient, 'full_name', None) or f"{getattr(patient, 'first_name', '')} {getattr(patient, 'last_name', '')}").strip()

            invoice = getattr(rec, 'invoice', None)
            invoice_no = getattr(invoice, 'invoice_number', None) if invoice else None
            svc_list = (details or {}).get('services') if isinstance(details, dict) else None
            if not isinstance(svc_list, list):
                svc_list = []

            services_annotated = []
            for s in svc_list:
                if not isinstance(s, dict):
                    continue
                s_type = s.get('type')
                s_name = s.get('name')
                s_price = s.get('price')
                # Compute the revenue-stream label for this entry using the same helpers
                # used for receipt allocation, so the user can see which bucket the row
                # contributes to instead of guessing from "type".
                raw_t = (str(s_type) if s_type else '').strip().lower().replace(' ', '_').replace('-', '_')
                try:
                    amt_dec = Decimal(str(s_price or 0))
                except Exception:
                    amt_dec = Decimal('0.00')
                stream_label = '-'
                try:
                    sub = _expand_combined_entry(s, raw_t, amt_dec) if amt_dec > 0 else None
                except Exception:
                    sub = None
                if sub:
                    stream_label = ', '.join(
                        sorted({_SERVICE_DISPLAY.get(k, str(k).title()) for k in sub.keys()})
                    )
                else:
                    norm = _normalize_service_type(s_type or '')
                    if norm == 'combined':
                        norm = 'other'
                    stream_label = _SERVICE_DISPLAY.get(norm, str(norm).title())
                services_annotated.append({
                    'type': s_type,
                    'name': s_name,
                    'price': s_price,
                    'revenue_stream': stream_label,
                })

            rows.append({
                'receipt_number': getattr(rec, 'receipt_number', None),
                'receipt_date': getattr(rec, 'receipt_date', None),
                'patient_name': patient_name,
                'invoice_number': invoice_no,
                'payment_method': getattr(rec, 'payment_method', None),
                'raw_service_type': raw_st_key,
                'amount_paid': paid,
                'allocated_amount': amt,
                'is_deposit': deposit_flag,
                'allocation': {k: float(v) for k, v in alloc.items()},
                'services': services_annotated,
                'invoice_category_mix': _invoice_line_category_debug(invoice),
                'invoice_lines': _invoice_line_items_debug(invoice),
            })

            if len(rows) >= limit:
                break

    context = {
        'date_from': date_from,
        'date_to': date_to,
        'bucket': bucket,
        'bucket_label': _SERVICE_DISPLAY.get(bucket, bucket.title()),
        'raw_filter': raw_filter,
        'rows': rows,
        'total_bucket': total_bucket,
        'row_count': len(rows),
        'limit': limit,
    }
    return render(request, 'hospital/revenue/streams_service_details.html', context)


@login_required
@user_passes_test(finance_revenue_streams_access)
def revenue_by_department_report(request):
    """
    Detailed report of revenue by department
    """
    # Get date range
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from:
        today = timezone.now().date()
        date_from = today.replace(day=1)
    else:
        date_from = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
    
    if not date_to:
        date_to = timezone.now().date()
    else:
        date_to = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Get all departments
    departments = Department.objects.filter(is_deleted=False)
    
    dept_data = []
    total_all = Decimal('0.00')
    
    for dept in departments:
        try:
            dept_revenue = Revenue.objects.filter(
                department=dept,
                revenue_date__gte=date_from,
                revenue_date__lte=date_to,
                is_deleted=False
            ).aggregate(
                total=Sum('amount'),
                count=Count('id')
            )
            
            total = dept_revenue['total'] or Decimal('0.00')
            count = dept_revenue['count'] or 0
            
            if total > 0 or count > 0:
                dept_data.append({
                    'department': dept,
                    'total': total,
                    'count': count,
                    'average': total / count if count > 0 else Decimal('0.00')
                })
                total_all += total
        except:
            pass
    
    # Calculate percentages
    for item in dept_data:
        if total_all > 0:
            item['percentage'] = (item['total'] / total_all) * 100
        else:
            item['percentage'] = 0
    
    # Sort by total
    dept_data.sort(key=lambda x: x['total'], reverse=True)
    
    avg_per_department = (total_all / len(dept_data)) if dept_data else Decimal('0.00')
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'dept_data': dept_data,
        'total_revenue': total_all,
        'average_per_department': avg_per_department,
    }
    
    return render(request, 'hospital/revenue/department_report.html', context)


@login_required
@user_passes_test(finance_revenue_streams_access)
def revenue_streams_api(request):
    """
    API endpoint for revenue stream data (for charts)
    """
    try:
        today = timezone.now().date()
        if request.GET.get('date_from') or request.GET.get('date_to'):
            default_from = today.replace(day=1)
            start_date = _parse_ymd(request.GET.get('date_from'), default_from)
            end_date = _parse_ymd(request.GET.get('date_to'), today)
        else:
            days = max(1, int(request.GET.get('days', 30)))
            end_date = today
            start_date = end_date - timedelta(days=days)

        analytics_data, _, _, _ = _build_revenue_streams_analytics(start_date, end_date)
        return JsonResponse({
            'success': True,
            **analytics_data,
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
        }, status=500)


@login_required
@user_passes_test(finance_revenue_streams_access)
def revenue_streams_dashboard_print(request):
    today = timezone.now().date()
    default_from = today.replace(day=1)
    date_from = _parse_ymd(request.GET.get('date_from'), default_from)
    date_to = _parse_ymd(request.GET.get('date_to'), today)
    analytics_data, revenue_by_service, revenue_by_dept, total_revenue = _build_revenue_streams_analytics(date_from, date_to)
    return render(request, 'hospital/revenue/streams_dashboard_print.html', {
        'date_from': date_from,
        'date_to': date_to,
        'total_revenue': total_revenue,
        'revenue_by_service': revenue_by_service,
        'revenue_by_dept': revenue_by_dept,
        'analytics_data': analytics_data,
        'kpis': analytics_data.get('kpis') or {},
    })


@login_required
@user_passes_test(finance_revenue_streams_access)
def revenue_streams_dashboard_export_excel(request):
    today = timezone.now().date()
    default_from = today.replace(day=1)
    date_from = _parse_ymd(request.GET.get('date_from'), default_from)
    date_to = _parse_ymd(request.GET.get('date_to'), today)
    analytics_data, revenue_by_service, revenue_by_dept, total_revenue = _build_revenue_streams_analytics(date_from, date_to)

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Summary'
    title = f"Revenue Streams ({date_from.isoformat()} to {date_to.isoformat()})"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])
    ws.append(['Total revenue (GHS)', float(total_revenue)])
    ws['A3'].font = Font(bold=True)
    ws['B3'].number_format = '#,##0.00'
    ws.append([])
    ws.append(['KPI', 'Amount (GHS)'])
    hdr = ws[5]
    for c in hdr:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1f2937')
    for key in ['consultation', 'lab', 'pharmacy', 'imaging', 'dental', 'gynecology', 'surgery', 'emergency', 'ambulance', 'admission', 'other']:
        ws.append([_SERVICE_DISPLAY.get(key, key.title()), float((analytics_data.get('kpis') or {}).get(key, 0.0) or 0.0)])
        ws[f'B{ws.max_row}'].number_format = '#,##0.00'

    ws2 = wb.create_sheet('By Service')
    ws2.append(['Service type', 'Amount (GHS)', 'Transactions', 'Share %'])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for r in revenue_by_service:
        ws2.append([
            r.get('service_name') or _SERVICE_DISPLAY.get(r.get('service_type'), 'Other'),
            float(r.get('total') or 0),
            int(r.get('count') or 0),
            float(r.get('percentage') or 0),
        ])
    for i in range(2, ws2.max_row + 1):
        ws2[f'B{i}'].number_format = '#,##0.00'
        ws2[f'D{i}'].number_format = '0.0'

    ws3 = wb.create_sheet('By Department')
    ws3.append(['Department', 'Amount (GHS)', 'Transactions', 'Share %'])
    for c in ws3[1]:
        c.font = Font(bold=True)
    for r in revenue_by_dept:
        dept = (
            (r.get('department') or r.get('department__name'))
            if isinstance(r, dict)
            else getattr(r, 'department__name', None)
        )
        ws3.append([
            dept or 'Unassigned',
            float((r.get('total') if isinstance(r, dict) else getattr(r, 'total', 0)) or 0),
            int((r.get('count') if isinstance(r, dict) else getattr(r, 'count', 0)) or 0),
            float((r.get('percentage') if isinstance(r, dict) else getattr(r, 'percentage', 0)) or 0),
        ])
    for i in range(2, ws3.max_row + 1):
        ws3[f'B{i}'].number_format = '#,##0.00'
        ws3[f'D{i}'].number_format = '0.0'

    ws4 = wb.create_sheet('Daily Trend')
    trend = (analytics_data.get('charts') or {}).get('trend_total') or {}
    ws4.append(['Date', 'Total (GHS)'])
    for c in ws4[1]:
        c.font = Font(bold=True)
    for d, t in zip(trend.get('labels') or [], trend.get('totals') or []):
        ws4.append([d, float(t or 0)])
        ws4[f'B{ws4.max_row}'].number_format = '#,##0.00'

    # Other breakdown sheets
    ws5 = wb.create_sheet('Other Breakdown')
    ws5.append(['Reason', 'Count', 'Amount (GHS)'])
    for c in ws5[1]:
        c.font = Font(bold=True)
    for r in (analytics_data.get('breakdowns') or {}).get('other_reasons') or []:
        ws5.append([r.get('reason'), int(r.get('count') or 0), float(r.get('total') or 0)])
        ws5[f'C{ws5.max_row}'].number_format = '#,##0.00'

    ws6 = wb.create_sheet('Raw Service Types')
    ws6.append(['service_type', 'Count', 'Amount (GHS)'])
    for c in ws6[1]:
        c.font = Font(bold=True)
    for r in (analytics_data.get('breakdowns') or {}).get('raw_service_types') or []:
        ws6.append([r.get('service_type'), int(r.get('count') or 0), float(r.get('total') or 0)])
        ws6[f'C{ws6.max_row}'].number_format = '#,##0.00'

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                v = '' if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            sheet.column_dimensions[col_letter].width = min(60, max(12, max_len + 2))

    buf = io.BytesIO()
    wb.save(buf)
    fname = f"revenue_streams_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename=\"{fname}\"'
    return resp


# ==================== STREAM REPORTING (SENDABLE EXCEL) ====================

def _parse_choice(value, allowed, default=None):
    v = (value or '').strip().lower()
    return v if v in allowed else default


def _payer_type_for_receipt(rec):
    try:
        inv = getattr(rec, 'invoice', None)
        payer = getattr(inv, 'payer', None) if inv else None
        pt = getattr(payer, 'payer_type', None)
        return (str(pt).strip().lower() if pt else 'unknown')
    except Exception:
        return 'unknown'


def _patient_name_for_receipt(rec):
    patient = getattr(rec, 'patient', None)
    if not patient:
        return ''
    return (getattr(patient, 'full_name', None) or f"{getattr(patient, 'first_name', '')} {getattr(patient, 'last_name', '')}").strip()


def _split_consultation_alloc(rec, consultation_amount):
    """
    Split a consultation allocation into general vs specialist using invoice lines when available.
    Returns dict with keys: consultation_general, consultation_specialist (values sum to consultation_amount).
    If we cannot infer, returns all as consultation_general.
    """
    try:
        from .models import InvoiceLine
        from .utils_billing import GENERAL_OPD_LINE_SERVICE_CODES
    except Exception:
        return {'consultation_general': consultation_amount, 'consultation_specialist': Decimal('0.00')}

    invoice = getattr(rec, 'invoice', None)
    if not invoice or not getattr(invoice, 'id', None):
        return {'consultation_general': consultation_amount, 'consultation_specialist': Decimal('0.00')}

    try:
        lines = (
            InvoiceLine.objects.filter(invoice_id=invoice.id, is_deleted=False)
            .select_related('service_code')
            .only(
                'id',
                'quantity',
                'unit_price',
                'discount_amount',
                'tax_amount',
                'line_total',
                'service_code__code',
            )
        )
    except Exception:
        return {'consultation_general': consultation_amount, 'consultation_specialist': Decimal('0.00')}

    general_total = Decimal('0.00')
    specialist_total = Decimal('0.00')

    for line in lines:
        if getattr(line, 'waived_at', None):
            continue
        sc = getattr(line, 'service_code', None)
        code = (getattr(sc, 'code', '') or '').strip().upper()
        if not code:
            continue
        try:
            line_total = Decimal(str(getattr(line, 'display_line_total', None) or line.line_total or 0))
        except Exception:
            line_total = Decimal('0.00')
        if line_total <= 0:
            continue

        # Specialist consultation line convention
        if code == 'CON002':
            specialist_total += line_total
        # General consultation-like lines (CON001, CONS-GEN, S00023, etc.)
        elif code in GENERAL_OPD_LINE_SERVICE_CODES:
            general_total += line_total

    denom = general_total + specialist_total
    if denom <= 0:
        return {'consultation_general': consultation_amount, 'consultation_specialist': Decimal('0.00')}

    # Allocate the receipt's consultation portion proportionally to the billed consultation mix.
    gen = (consultation_amount * (general_total / denom)) if general_total > 0 else Decimal('0.00')
    spec = (consultation_amount * (specialist_total / denom)) if specialist_total > 0 else Decimal('0.00')
    remainder = consultation_amount - (gen + spec)
    if remainder != 0:
        gen += remainder
    return {'consultation_general': gen, 'consultation_specialist': spec}


def _alloc_for_receipt_with_specialist_split(rec):
    """
    Like _alloc_for_receipt(rec) but splits consultation into:
    - consultation_general
    - consultation_specialist
    """
    alloc = _alloc_for_receipt(rec)
    if not alloc:
        return alloc

    consult_amt = alloc.get('consultation')
    if consult_amt is None:
        return alloc
    try:
        consult_amt = Decimal(str(consult_amt or 0))
    except Exception:
        consult_amt = Decimal('0.00')

    # Remove original consultation bucket and replace with split keys.
    alloc2 = dict(alloc)
    alloc2.pop('consultation', None)

    split = _split_consultation_alloc(rec, consult_amt)
    for k, v in split.items():
        if v and v != 0:
            alloc2[k] = alloc2.get(k, Decimal('0.00')) + v
    return alloc2


def _stream_label(key):
    if key == 'consultation_general':
        return 'Consultation (General)'
    if key == 'consultation_specialist':
        return 'Consultation (Specialist)'
    return _SERVICE_DISPLAY.get(key, key.replace('_', ' ').title())


def _mgmt_xlsx_header_style():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill('solid', fgColor='1E40AF')
    font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return font, fill, align, border


def _mgmt_style_header_row(ws, row_idx, n_cols):
    font, fill, align, border = _mgmt_xlsx_header_style()
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row_idx, column=col)
        c.font = font
        c.fill = fill
        c.alignment = align
        c.border = border


def _mgmt_body_border(ws, row_idx, n_cols):
    from openpyxl.styles import Border, Side, Alignment

    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(vertical='top', wrap_text=True)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row_idx, column=col)
        c.border = border
        c.alignment = align


def _mgmt_autosize_sheet(ws, max_scan_rows=5000, min_w=10, max_w=55):
    from openpyxl.utils import get_column_letter

    if not ws.max_column:
        return
    last_row = min(ws.max_row or 1, max_scan_rows)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = min_w
        for r in range(1, last_row + 1):
            val = ws.cell(row=r, column=col).value
            if val is not None:
                maxlen = max(maxlen, min(90, len(str(val))))
        ws.column_dimensions[letter].width = min(max_w, maxlen + 2)


def _mgmt_period_gl_totals(date_from, date_to):
    """Best-effort accrued revenue / paid expenses for the same window (advanced accounting)."""
    revenue = Decimal('0.00')
    expense = Decimal('0.00')
    try:
        from .models_accounting_advanced import Revenue, Expense
        from django.db.models import Sum

        r = Revenue.objects.filter(revenue_date__gte=date_from, revenue_date__lte=date_to).aggregate(t=Sum('amount'))['t']
        revenue = Decimal(str(r or 0))
        e = Expense.objects.filter(
            expense_date__gte=date_from,
            expense_date__lte=date_to,
            status='paid',
        ).aggregate(t=Sum('amount'))['t']
        expense = Decimal(str(e or 0))
    except Exception:
        pass
    return revenue, expense


@login_required
@user_passes_test(is_accountant)
def accounting_stream_reports_export_excel(request):
    """
    Sendable, detailed stream workbook (cash received headline) with optional filters.

    Query params:
      - date_from, date_to (Y-m-d)
      - stream: one of canonical stream keys, or 'all' (default)
      - payer_type: cash/nhis/private/corporate/all (default all)
      - payment_method: cash/card/bank_transfer/mobile_money/cheque/insurance/deposit/all (default all)
    """
    try:
        from .models_accounting import PaymentReceipt
    except Exception:
        PaymentReceipt = None

    today = timezone.now().date()
    default_from = today.replace(day=1)
    date_from = _parse_ymd(request.GET.get('date_from'), default_from)
    date_to = _parse_ymd(request.GET.get('date_to'), today)

    stream_filter = _parse_choice(
        request.GET.get('stream') or 'all',
        allowed={
            'all',
            'consultation_general',
            'consultation_specialist',
            'lab',
            'pharmacy',
            'imaging',
            'dental',
            'gynecology',
            'surgery',
            'emergency',
            'ambulance',
            'admission',
            'other',
        },
        default='all',
    )
    payer_type_filter = _parse_choice(
        request.GET.get('payer_type') or 'all',
        allowed={'all', 'cash', 'nhis', 'private', 'corporate'},
        default='all',
    )
    payment_method_filter = _parse_choice(
        request.GET.get('payment_method') or 'all',
        allowed={'all', 'cash', 'card', 'bank_transfer', 'mobile_money', 'cheque', 'insurance', 'deposit'},
        default='all',
    )

    if not PaymentReceipt:
        return HttpResponse("PaymentReceipt model not available.", status=500)

    qs = (
        PaymentReceipt.objects.filter(
            receipt_date__date__gte=date_from,
            receipt_date__date__lte=date_to,
            is_deleted=False,
        )
        .select_related('patient', 'invoice', 'invoice__payer')
        .only(
            'id',
            'receipt_number',
            'receipt_date',
            'amount_paid',
            'payment_method',
            'service_type',
            'service_details',
            'notes',
            'patient__id',
            'patient__first_name',
            'patient__last_name',
            'invoice__id',
            'invoice__invoice_number',
            'invoice__payer__id',
            'invoice__payer__name',
            'invoice__payer__payer_type',
        )
        .order_by('-receipt_date')
    )

    if payment_method_filter and payment_method_filter != 'all':
        qs = qs.filter(payment_method=payment_method_filter)

    # We'll filter payer_type at row-build time because payer_type depends on invoice link presence.

    # Build detailed rows (one row per receipt allocation).
    rows = []
    totals_by_stream = {}
    totals_by_payer_type = {}
    totals_by_payment_method = {}
    receipt_count = 0
    total_cash_in_scope = Decimal('0.00')

    for rec in qs.iterator(chunk_size=500):
        payer_type = _payer_type_for_receipt(rec)
        if payer_type_filter != 'all' and payer_type != payer_type_filter:
            continue

        try:
            paid = Decimal(str(getattr(rec, 'amount_paid', 0) or 0))
        except Exception:
            paid = Decimal('0.00')

        receipt_count += 1
        total_cash_in_scope += paid

        pm = (getattr(rec, 'payment_method', None) or 'unknown')
        totals_by_payment_method[pm] = totals_by_payment_method.get(pm, Decimal('0.00')) + paid

        alloc = _alloc_for_receipt_with_specialist_split(rec)
        if not alloc:
            continue

        inv = getattr(rec, 'invoice', None)
        payer = getattr(inv, 'payer', None) if inv else None
        payer_name = getattr(payer, 'name', '') if payer else ''
        invoice_no = getattr(inv, 'invoice_number', '') if inv else ''

        # Totals by payer type: allocated across streams, but payer bucket is receipt-level
        totals_by_payer_type[payer_type] = totals_by_payer_type.get(payer_type, Decimal('0.00')) + paid

        for k, v in alloc.items():
            if v is None:
                continue
            try:
                amt = Decimal(str(v or 0))
            except Exception:
                amt = Decimal('0.00')
            if amt == 0:
                continue
            if stream_filter != 'all' and k != stream_filter:
                continue

            totals_by_stream[k] = totals_by_stream.get(k, Decimal('0.00')) + amt

            rows.append({
                'receipt_date': getattr(rec, 'receipt_date', None),
                'receipt_number': getattr(rec, 'receipt_number', ''),
                'invoice_number': invoice_no,
                'patient_name': _patient_name_for_receipt(rec),
                'payer_name': payer_name,
                'payer_type': payer_type,
                'payment_method': pm,
                'stream': k,
                'stream_label': _stream_label(k),
                'allocated_amount': amt,
                'receipt_amount': paid,
                'notes': (getattr(rec, 'notes', None) or ''),
            })

    # Sort rows for readability
    rows.sort(key=lambda r: (r['stream_label'], r['receipt_date'] or timezone.now()), reverse=False)

    sum_allocated = sum((Decimal(str(r.get('allocated_amount') or 0)) for r in rows), Decimal('0.00'))
    headline_total = (
        sum(totals_by_stream.values(), Decimal('0.00'))
        if stream_filter != 'all'
        else sum(totals_by_payer_type.values(), Decimal('0.00'))
    )
    gl_revenue, gl_expense = _mgmt_period_gl_totals(date_from, date_to)
    net_indicator = gl_revenue - gl_expense

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    thin_line = openpyxl.styles.Side(style='thin', color='CBD5E1')
    grid_border = openpyxl.styles.Border(
        left=thin_line, right=thin_line, top=thin_line, bottom=thin_line
    )

    # --- Cover (management pack) ---
    ws_cov = wb.active
    ws_cov.title = 'Cover'
    hospital_name = getattr(settings, 'HOSPITAL_NAME', 'Hospital Management System')
    ws_cov.merge_cells('A1:F1')
    ws_cov['A1'] = hospital_name
    ws_cov['A1'].font = Font(name='Calibri', size=22, bold=True, color='1E3A8A')
    ws_cov['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_cov.row_dimensions[1].height = 34
    ws_cov.merge_cells('A2:F2')
    ws_cov['A2'] = 'Management revenue pack'
    ws_cov['A2'].font = Font(name='Calibri', size=13, italic=True, color='475569')
    ws_cov['A2'].alignment = Alignment(horizontal='center')
    ws_cov.merge_cells('A3:F3')
    stream_title = _stream_label(stream_filter) if stream_filter != 'all' else 'All clinical streams'
    ws_cov['A3'] = f'{stream_title} · Cash collections & billing intelligence'
    ws_cov['A3'].font = Font(name='Calibri', size=11, color='64748B')
    ws_cov['A3'].alignment = Alignment(horizontal='center', wrap_text=True)

    param_row = 5
    ws_cov.cell(param_row, 1, 'Report parameters').font = Font(bold=True, size=12, color='1E40AF')
    param_row += 1
    param_labels = [
        ('Reporting period', f'{date_from.isoformat()} → {date_to.isoformat()}'),
        ('Stream filter', stream_title),
        ('Payer type filter', payer_type_filter.replace('_', ' ').title()),
        ('Payment method filter', payment_method_filter.replace('_', ' ').title()),
        ('Prepared for', (request.user.get_full_name() or '').strip() or request.user.username),
        ('Generated (server time)', timezone.now().strftime('%Y-%m-%d %H:%M %Z')),
    ]
    for label, val in param_labels:
        ws_cov.cell(param_row, 1, label).font = Font(bold=True)
        ws_cov.cell(param_row, 2, val).font = Font(size=11)
        ws_cov.cell(param_row, 1).border = grid_border
        ws_cov.cell(param_row, 2).border = grid_border
        param_row += 1
    param_row += 1
    ws_cov.merge_cells(start_row=param_row, start_column=1, end_row=param_row, end_column=6)
    ws_cov.cell(param_row, 1, 'Confidential — for internal management use only.')
    ws_cov.cell(param_row, 1).font = Font(size=9, italic=True, color='94A3B8')
    for col in range(1, 7):
        ws_cov.column_dimensions[get_column_letter(col)].width = 22
    ws_cov.column_dimensions['B'].width = 40

    # --- Executive summary ---
    ws_ex = wb.create_sheet('Executive summary')
    ws_ex.append([f'Executive snapshot · {date_from} to {date_to}'])
    ws_ex.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws_ex['A1'].font = Font(size=16, bold=True, color='1E3A8A')
    ws_ex['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws_ex.row_dimensions[1].height = 26
    ws_ex.append([])

    kpi_specs = [
        ('Receipts in scope (count)', receipt_count, None),
        ('Total cash received — filtered receipts (GHS)', float(total_cash_in_scope), '#,##0.00'),
        ('Headline total — as on hub (GHS)', float(headline_total), '#,##0.00'),
        ('Sum of allocation lines in this export (GHS)', float(sum_allocated), '#,##0.00'),
        ('Accrued revenue (GL) same window (GHS)', float(gl_revenue), '#,##0.00'),
        ('Paid expenses (GL) same window (GHS)', float(gl_expense), '#,##0.00'),
        ('Net (GL revenue − expenses) (GHS)', float(net_indicator), '#,##0.00'),
        ('Detail allocation rows', len(rows), None),
    ]
    ws_ex.append(['KPI', 'Value', 'Notes'])
    _mgmt_style_header_row(ws_ex, ws_ex.max_row, 3)
    for label, val, numfmt in kpi_specs:
        ws_ex.append([label, val, ''])
        r = ws_ex.max_row
        ws_ex.cell(r, 1).font = Font(bold=True)
        if numfmt:
            ws_ex.cell(r, 2).number_format = numfmt
            if val is not None and isinstance(val, (int, float)) and label.startswith('Net (GL'):
                ws_ex.cell(r, 2).font = Font(bold=True, color='047857' if val >= 0 else 'B91C1C')
        _mgmt_body_border(ws_ex, r, 3)
    ws_ex.append([])
    ws_ex.append(
        [
            'Cash rows = every receipt matching filters; allocation rows split combined receipts by stream.',
            '',
            '',
        ]
    )
    _mgmt_body_border(ws_ex, ws_ex.max_row, 3)
    ws_ex.cell(ws_ex.max_row, 1).font = Font(size=9, italic=True, color='64748B')
    note_row = ws_ex.max_row + 1
    ws_ex.cell(note_row, 1, 'Note: GL revenue/expense uses advanced accounting tables when present; otherwise amounts may be zero.')
    ws_ex.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    ws_ex.cell(note_row, 1).font = Font(size=9, italic=True, color='64748B')

    # --- Summary (collections mix) ---
    ws = wb.create_sheet('Collections summary')
    title = f"Cash received by payer & method — {date_from.isoformat()} to {date_to.isoformat()}"
    if stream_filter != 'all':
        title = f"{_stream_label(stream_filter)} — {title}"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws['A1'].font = Font(size=14, bold=True, color='1E3A8A')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.append([])
    ws.append(['Headline total (GHS)', float(headline_total), '', '', ''])
    ws['A3'].font = Font(bold=True, size=12)
    ws['B3'].number_format = '#,##0.00'
    ws['B3'].font = Font(bold=True, size=12, color='047857')
    ws.append([])

    ws.append(['Payer type', 'Amount (GHS)', '% of payer mix', '', ''])
    _mgmt_style_header_row(ws, ws.max_row, 5)
    payer_grand = sum(totals_by_payer_type.values(), Decimal('0.00')) or Decimal('1')
    for pt in ['cash', 'nhis', 'private', 'corporate', 'unknown']:
        if pt not in totals_by_payer_type:
            continue
        amt = totals_by_payer_type.get(pt, Decimal('0.00')) or Decimal('0.00')
        ratio = (amt / payer_grand) if payer_grand else Decimal('0')
        ws.append([pt, float(amt), float(ratio), '', ''])
        r = ws.max_row
        ws[f'B{r}'].number_format = '#,##0.00'
        ws[f'C{r}'].number_format = '0.0%'
        _mgmt_body_border(ws, r, 5)

    ws.append([])
    ws.append(['Payment method', 'Amount (GHS)', '% of methods', '', ''])
    _mgmt_style_header_row(ws, ws.max_row, 5)
    pm_grand = sum(totals_by_payment_method.values(), Decimal('0.00')) or Decimal('1')
    for pm, amt in sorted(totals_by_payment_method.items(), key=lambda kv: kv[1], reverse=True):
        amt = amt or Decimal('0.00')
        ratio = (amt / pm_grand) if pm_grand else Decimal('0')
        ws.append([pm, float(amt), float(ratio), '', ''])
        r = ws.max_row
        ws[f'B{r}'].number_format = '#,##0.00'
        ws[f'C{r}'].number_format = '0.0%'
        _mgmt_body_border(ws, r, 5)

    # --- Streams ---
    ws2 = wb.create_sheet('Streams breakdown')
    ws2.append(['Stream', 'Allocated amount (GHS)', '% of streams', 'Rank'])
    _mgmt_style_header_row(ws2, 1, 4)
    stream_grand = sum(totals_by_stream.values(), Decimal('0.00')) or Decimal('1')
    sorted_streams = sorted(totals_by_stream.items(), key=lambda kv: kv[1], reverse=True)
    for rank, (k, amt) in enumerate(sorted_streams, start=1):
        amt = amt or Decimal('0.00')
        ratio = (amt / stream_grand) if stream_grand else Decimal('0')
        ws2.append([_stream_label(k), float(amt), float(ratio), rank])
        r = ws2.max_row
        ws2[f'B{r}'].number_format = '#,##0.00'
        ws2[f'C{r}'].number_format = '0.0%'
        _mgmt_body_border(ws2, r, 4)

    # --- Detail ---
    ws3 = wb.create_sheet('Receipt allocations')
    hdr = [
        'Receipt date',
        'Receipt #',
        'Invoice #',
        'Patient',
        'Payer',
        'Payer type',
        'Payment method',
        'Stream',
        'Allocated amount (GHS)',
        'Receipt amount (GHS)',
        'Notes',
    ]
    ws3.append(hdr)
    _mgmt_style_header_row(ws3, 1, len(hdr))
    zebra = PatternFill('solid', fgColor='F8FAFC')
    for r in rows:
        dt = r.get('receipt_date')
        ws3.append([
            dt.strftime('%Y-%m-%d %H:%M') if dt else '',
            r.get('receipt_number') or '',
            r.get('invoice_number') or '',
            r.get('patient_name') or '',
            r.get('payer_name') or '',
            r.get('payer_type') or '',
            r.get('payment_method') or '',
            r.get('stream_label') or r.get('stream') or '',
            float(r.get('allocated_amount') or Decimal('0.00')),
            float(r.get('receipt_amount') or Decimal('0.00')),
            r.get('notes') or '',
        ])
        rr = ws3.max_row
        ws3[f'I{rr}'].number_format = '#,##0.00'
        ws3[f'J{rr}'].number_format = '#,##0.00'
        _mgmt_body_border(ws3, rr, len(hdr))
        if rr % 2 == 0:
            for c in range(1, len(hdr) + 1):
                ws3.cell(rr, c).fill = zebra
    ws3.freeze_panes = 'A2'
    if ws3.max_row >= 1:
        ws3.auto_filter.ref = f'A1:{get_column_letter(len(hdr))}{ws3.max_row}'

    # --- Notes ---
    ws_notes = wb.create_sheet('Definitions & notes')
    notes = [
        'CASH RECEIVED (headline): Total amount paid on receipts in the selected period, after payer and payment-method filters.',
        'ALLOCATIONS: Each receipt is split into clinical streams (consultation general/specialist, lab, pharmacy, etc.) using receipt metadata and invoice line mix where applicable.',
        'COMBINED RECEIPTS: When one receipt covers multiple services, the Detail sheet shows one row per stream allocation; amounts sum to the receipt where allocation is complete.',
        'STREAM FILTER: Restricts Detail and stream totals to a single department bucket; payer and payment tables still reflect all receipts matching other filters.',
        'GL REVENUE / EXPENSE: Pulled from advanced accounting Revenue and Expense (paid) for the same calendar dates — management view of accrual/cash expense vs cashier collections.',
        'BILLED vs NET LINES: Billed Lines uses stored invoice line amounts; Net Lines uses display_line_total / display_unit_price where repricing applies.',
    ]
    ws_notes.append(['Methodology & definitions'])
    ws_notes['A1'].font = Font(size=14, bold=True, color='1E3A8A')
    for n in notes:
        ws_notes.append([n])
        ws_notes.cell(ws_notes.max_row, 1).alignment = Alignment(wrap_text=True, vertical='top')
    ws_notes.column_dimensions['A'].width = 110

    # Reconciliation sheets: billed vs net invoice lines
    try:
        from .models import InvoiceLine
    except Exception:
        InvoiceLine = None

    if InvoiceLine:
        line_qs = (
            InvoiceLine.objects.filter(
                created__date__gte=date_from,
                created__date__lte=date_to,
                is_deleted=False,
            )
            .select_related('invoice', 'invoice__payer', 'invoice__patient', 'service_code')
            .only(
                'id',
                'created',
                'invoice_id',
                'invoice__invoice_number',
                'invoice__patient__first_name',
                'invoice__patient__last_name',
                'invoice__payer__name',
                'invoice__payer__payer_type',
                'prescription_id',
                'service_code__code',
                'service_code__category',
                'description',
                'quantity',
                'unit_price',
                'tax_amount',
                'discount_amount',
                'line_total',
                'waived_at',
            )
            .order_by('created')
        )

        ws4 = wb.create_sheet('Billed Lines')
        ws4.append([
            'Billed date',
            'Invoice #',
            'Patient',
            'Payer',
            'Payer type',
            'Service code',
            'Category',
            'Description',
            'Is pharmacy line',
            'Qty',
            'Unit price (GHS)',
            'Tax (GHS)',
            'Discount (GHS)',
            'Line total (GHS)',
            'Waived',
        ])
        _mgmt_style_header_row(ws4, 1, 14)

        ws5 = wb.create_sheet('Net Lines')
        ws5.append([
            'Billed date',
            'Invoice #',
            'Patient',
            'Payer',
            'Payer type',
            'Service code',
            'Category',
            'Description',
            'Is pharmacy line',
            'Qty',
            'Display unit (GHS)',
            'Display line total (GHS)',
            'Waived',
        ])
        _mgmt_style_header_row(ws5, 1, 12)

        for line in line_qs.iterator(chunk_size=1000):
            inv = getattr(line, 'invoice', None)
            payer = getattr(inv, 'payer', None) if inv else None
            patient = getattr(inv, 'patient', None) if inv else None

            patient_name = ''
            if patient:
                patient_name = (getattr(patient, 'full_name', None) or f"{getattr(patient, 'first_name', '')} {getattr(patient, 'last_name', '')}").strip()

            payer_name = getattr(payer, 'name', '') if payer else ''
            payer_type = (getattr(payer, 'payer_type', None) or 'unknown') if payer else 'unknown'

            sc = getattr(line, 'service_code', None)
            sc_code = (getattr(sc, 'code', '') or '').strip()
            sc_cat = (getattr(sc, 'category', '') or '').strip()

            billed_dt = getattr(line, 'created', None)
            billed_str = billed_dt.strftime('%Y-%m-%d %H:%M') if billed_dt else ''
            invoice_no = getattr(inv, 'invoice_number', '') if inv else ''

            qty = getattr(line, 'quantity', None) or 0
            unit = getattr(line, 'unit_price', None) or 0
            tax = getattr(line, 'tax_amount', None) or 0
            disc = getattr(line, 'discount_amount', None) or 0
            lt = getattr(line, 'line_total', None) or 0
            waived = bool(getattr(line, 'waived_at', None))
            is_pharmacy_line = bool(getattr(line, 'prescription_id', None))

            ws4.append([
                billed_str,
                invoice_no,
                patient_name,
                payer_name,
                payer_type,
                sc_code,
                sc_cat,
                getattr(line, 'description', '') or '',
                'Yes' if is_pharmacy_line else 'No',
                float(qty),
                float(unit),
                float(tax),
                float(disc),
                float(lt),
                'Yes' if waived else 'No',
            ])
            for col_letter in ['K', 'L', 'M', 'N']:
                ws4[f'{col_letter}{ws4.max_row}'].number_format = '#,##0.00'

            # Net sheet uses display_* pricing (handles stale repricing)
            try:
                d_unit = getattr(line, 'display_unit_price', None) or unit
            except Exception:
                d_unit = unit
            try:
                d_total = getattr(line, 'display_line_total', None) or lt
            except Exception:
                d_total = lt

            ws5.append([
                billed_str,
                invoice_no,
                patient_name,
                payer_name,
                payer_type,
                sc_code,
                sc_cat,
                getattr(line, 'description', '') or '',
                'Yes' if is_pharmacy_line else 'No',
                float(qty),
                float(d_unit or 0),
                float(d_total or 0),
                'Yes' if waived else 'No',
            ])
            ws5[f'K{ws5.max_row}'].number_format = '#,##0.00'
            ws5[f'L{ws5.max_row}'].number_format = '#,##0.00'

    for sheet in wb.worksheets:
        _mgmt_autosize_sheet(sheet)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffix = stream_filter if stream_filter != 'all' else 'all_streams'
    fname = f"management_stream_report_{suffix}_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
    return FileResponse(
        buf,
        as_attachment=True,
        filename=fname,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

